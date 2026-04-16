from celery import shared_task
from import_export import resources
from .models import Ubicacion, Activo, Plano
import time

def try_decode(content, encodings=['utf-8-sig', 'iso-8859-1', 'windows-1252', 'utf-8']):
    """Intenta decodificar el contenido usando una lista de encodings prioritarios."""
    for encoding in encodings:
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    # Si ninguno funciona, forzar utf-8 ignorando errores para que al menos no rompa la tarea
    return content.decode('utf-8', errors='ignore')

@shared_task(bind=True)
def import_ubicaciones_task(self, file_path, file_format):
    """
    Tarea Celery para importar ubicaciones con seguimiento de progreso.
    
    Args:
        self: Instancia de la tarea (bind=True)
        file_path: Ruta al archivo temporal de importación
        file_format: Formato del archivo ('csv', 'xlsx', etc.)
    
    Returns:
        dict: Resultado de la importación con estadísticas
    """
    from tablib import Dataset
    import os
    
    # Inicializar el resource y cachés
    from .resources import UbicacionResource
    resource = UbicacionResource()
    
    from django.core.files.storage import default_storage
    # Leer el archivo
    with default_storage.open(file_path, 'rb') as f:
        file_content = f.read()
        if file_format == 'csv':
            dataset = Dataset().load(try_decode(file_content), format='csv')
        elif file_format in ['xls', 'xlsx']:
            dataset = Dataset().load(file_content, format=file_format)
        else:
            raise ValueError(f"Formato no soportado: {file_format}")
    
    total_rows = len(dataset)
    resource.before_import(dataset)
    
    # Actualizar estado inicial
    self.update_state(
        state='PROGRESS',
        meta={
            'current': 0,
            'total': total_rows,
            'status': 'Iniciando importación...',
            'current_row': None
        }
    )
    
    # Procesar fila por fila con seguimiento
    result = resources.Result()
    for i, row in enumerate(dataset.dict, start=1):
        try:
            # Obtener nombre de la fila para mostrar
            row_name = row.get('nombre', f'Fila {i}')
            
            # Actualizar progreso
            self.update_state(
                state='PROGRESS',
                meta={
                    'current': i,
                    'total': total_rows,
                    'status': f'Procesando {i}/{total_rows}',
                    'current_row': row_name,
                    'percent': int((i / total_rows) * 100)
                }
            )
            
            # Procesar la fila (usando ModelInstanceLoader para IE 4.x)
            from import_export.instance_loaders import ModelInstanceLoader
            instance_loader = ModelInstanceLoader(resource, dataset)
            row_result = resource.import_row(
                row,
                instance_loader,
                row_number=i,
                dry_run=False
            )
            result.append_row_result(row_result)
            
        except Exception as e:
            # Registrar error pero continuar
            result.append_base_error(
                resources.Error(
                    error=e,
                    traceback=str(e),
                    row=row
                )
            )
    
    # Limpiar archivo temporal
    try:
        if default_storage.exists(file_path):
            default_storage.delete(file_path)
    except:
        pass
    
    # Retornar resultado final
    return {
        'status': 'completed',
        'total': total_rows,
        'new': result.totals.get('new', 0),
        'updated': result.totals.get('update', 0),
        'skipped': result.totals.get('skip', 0),
        'errors': len(result.base_errors) + len(result.row_errors()),
        'error_messages': [str(e.error) for e in result.base_errors[:10]]  # Primeros 10 errores
    }

@shared_task(bind=True)
def import_activos_task(self, file_path, file_format, user_id=None, import_name="Importación sin nombre", verification_mode=False, dry_run=False):
    """
    Tarea Celery para importar activos con seguimiento de progreso, soporte para verificación y dry-run.
    """
    from tablib import Dataset
    import os
    import json
    from django.contrib.auth.models import User
    from django.core.cache import cache
    from .models import RegistroImportacion, Activo
    from django.core.files.storage import default_storage
    
    user = User.objects.get(id=user_id) if user_id else None
    cache_key = f"import_activos_progress_{user_id}" if user_id else "import_activos_progress_system"
    
    # Crear registro de importación solo si no es verificación pura
    registro = None
    if not verification_mode:
        registro = RegistroImportacion.objects.create(
            nombre=import_name,
            usuario=user,
            estado='PROCESANDO'
        )
    
    print(f"[CELERY] Iniciando importación: {import_name} (User: {user_id})")
    print(f"[CELERY] Cargando archivo: {file_path}")
    
    # Leer el archivo
    try:
        with default_storage.open(file_path, 'rb') as f:
            file_content = f.read()
            if file_format == 'csv':
                dataset = Dataset().load(try_decode(file_content), format='csv')
            elif file_format in ['xls', 'xlsx']:
                dataset = Dataset().load(file_content, format=file_format)
            else:
                raise ValueError(f"Formato no soportado: {file_format}")
    except Exception as e:
        if registro:
            registro.estado = 'ERROR'
            registro.detalles_error = f"Error al leer archivo: {str(e)}"
            registro.save()
        error_res = {'status': 'error', 'message': str(e)}
        cache.set(cache_key, error_res, 3600)
        return error_res

    total_rows = len(dataset)
    if registro:
        registro.total_rows = total_rows
        registro.save()

    from .resources import ActivoResource
    from import_export import resources
    resource = ActivoResource()
    
    # Estado inicial
    progress_info = {
        'current': 0, 
        'total': total_rows, 
        'status': 'Verificando archivo...' if verification_mode else 'Procesando importación...', 
        'percent': 0,
        'new': 0,
        'updated': 0,
        'skipped': 0,
        'errors': 0,
        'verification_mode': verification_mode,
        'dry_run': dry_run
    }
    cache.set(cache_key, progress_info, 3600)
    self.update_state(state='PROGRESS', meta=progress_info)

    # Unificar lógica: Modo Verificación fuerza dry_run
    if verification_mode:
        dry_run = True

    resource.before_import(dataset, user=user)
    
    print(f"[CELERY] Iniciando import_data para {total_rows} filas...")
    try:
        # Ejecutar importación (simulada si dry_run/verificación)
        # raise_errors=False para capturar errores por fila en el reporte
        result = resource.import_data(dataset, dry_run=dry_run, raise_errors=False)
        print(f"[CELERY] import_data completado. Procesando reporte...")

        # Recopilar resultados detallados
        detailed_messages = []
        detailed_errors = []
        
        # Procesar errores globales
        for error in result.base_errors:
            msg = f"Error General: {str(error.error)}"
            detailed_errors.append(msg)
            detailed_messages.append(f"[ERROR CRÍTICO] {msg}")

        # Procesar resultados por fila para el reporte detallado
        # OPTIMIZACIÓN CRÍTICA: dataset.dict es una propiedad costosa en tablib que recrea la lista.
        # La extraemos una sola vez fuera del loop.
        dataset_dict = dataset.dict
        
        for i, row_result in enumerate(result.rows):
            row_idx = i + 1
            data_row = dataset_dict[i]
            nombre = str(data_row.get('nombre') or 'S/N').strip()
            codigo = str(data_row.get('codigo_interno') or '---').strip()
            
            action = "SIN CAMBIOS"
            details = ""
            status_tag = "SKIP"
            
            if row_result.import_type == 'new':
                action = "NUEVO"
                status_tag = "NEW"
                details = "Se creará un nuevo activo."
            elif row_result.import_type == 'update':
                action = "ACTUALIZAR"
                status_tag = "UPDATE"
                # Usar el atributo 'changed_fields' que inyectamos en ActivoResource.import_row
                if hasattr(row_result, 'changed_fields') and row_result.changed_fields:
                    details = f"Cambios en: {', '.join(row_result.changed_fields)}"
                else:
                    details = "Actualización detectada."
            elif row_result.import_type == 'skip':
                action = "OMITIDO"
                status_tag = "SKIP"
                details = "Sin cambios o fila vacía."
            elif row_result.import_type == 'error':
                action = "ERROR"
                status_tag = "ERROR"
                err_msg = str(row_result.errors[0].error) if row_result.errors else "Error desconocido"
                details = f"{err_msg}"
                detailed_errors.append(f"Fila {row_idx}: {err_msg}")
            
            # Formato de línea para el log de verificación
            if verification_mode or dry_run:
                msg = f"Fila {row_idx}: {codigo} ({nombre}) -> [{action}] {details}"
                detailed_messages.append(msg)
        
        # Actualizar registro si existe (solo en modo no-verificación o dry-run confirmado)
        if registro:
            registro.filas_nuevas = result.totals.get('new', 0)
            registro.filas_actualizadas = result.totals.get('update', 0)
            registro.filas_omitidas = result.totals.get('skip', 0)
            registro.filas_error = len(detailed_errors)
            registro.estado = 'COMPLETADO' if not dry_run else 'PROCESANDO'
            if detailed_errors:
                registro.detalles_error = "\n".join(detailed_errors[:20]) # Guardar primeros 20 errores
            registro.save()

        final_res = {
            'status': 'completed',
            'status_code': 'completed',
            'total': total_rows,
            'new': result.totals.get('new', 0),
            'updated': result.totals.get('update', 0),
            'skipped': result.totals.get('skip', 0),
            'errors': len(detailed_errors),
            'error_list': detailed_errors,
            'results': detailed_messages, # Para el modal de verificación
            'verification_mode': verification_mode,
            'dry_run': dry_run,
            'file_path': file_path
        }
        
    except Exception as e:
        error_msg = f"Error crítico durante el procesamiento: {str(e)}"
        if registro:
            registro.estado = 'ERROR'
            registro.detalles_error = error_msg
            registro.save()
        final_res = {'status': 'error', 'message': error_msg}

    # Limpiar archivo original SOLO si no es dry_run y no es verificación
    if not dry_run and not verification_mode:
        try:
            if default_storage.exists(file_path):
                default_storage.delete(file_path)
        except:
            pass
        
    cache.set(cache_key, final_res, 3600)
    return final_res

@shared_task(bind=True)
def revertir_importacion_task(self, registro_id):
    """
    Tarea para eliminar los activos creados en una importación específica.
    """
    from .models import RegistroImportacion, Activo
    import json
    
    try:
        registro = RegistroImportacion.objects.get(id=registro_id)
        if registro.estado != 'COMPLETADO':
            return {'status': 'error', 'message': 'Solo se pueden revertir importaciones completadas.'}
        
        ids = json.loads(registro.ids_creados) if registro.ids_creados else []
        if not ids:
            registro.estado = 'REVERTIDO'
            registro.save()
            return {'status': 'completed', 'message': 'No había activos nuevos para eliminar.'}
        
        total = len(ids)
        # Eliminar activos por ID
        # Usamos filter().delete() para mayor eficiencia
        borrados, _ = Activo.objects.filter(id__in=ids).delete()
        
        registro.estado = 'REVERTIDO'
        registro.save()
        
        return {'status': 'completed', 'deleted_count': borrados, 'expected_count': total}
    except Exception as e:
        return {'status': 'error', 'message': str(e)}
@shared_task(bind=True)
def import_planos_task(self, file_path, file_format):
    """
    Tarea Celery para importar PLANOS con seguimiento de progreso real.
    """
    from tablib import Dataset
    import os
    from .resources import PlanoResource
    
    # Inicializar resource
    resource = PlanoResource()
    
    from django.core.files.storage import default_storage
    # Leer archivo
    try:
        with default_storage.open(file_path, 'rb') as f:
            file_content = f.read()
            if file_format == 'csv':
                dataset = Dataset().load(try_decode(file_content), format='csv')
            elif file_format in ['xls', 'xlsx']:
                dataset = Dataset().load(file_content, format=file_format)
            else:
                raise ValueError(f"Formato no soportado: {file_format}")
    except FileNotFoundError:
        return {'status': 'error', 'message': 'Archivo temporal no encontrado'}

    total_rows = len(dataset)
    resource.before_import(dataset)
    
    # Estado inicial
    self.update_state(
        state='PROGRESS',
        meta={'current': 0, 'total': total_rows, 'status': 'Iniciando importación...', 'percent': 0}
    )
    
    result = resources.Result()
    
    for i, row in enumerate(dataset.dict, start=1):
        try:
            # Feedback en Redis cada 5 filas para que la barra se mueva fluido
            if i % 5 == 0 or i == total_rows:
                self.update_state(
                    state='PROGRESS',
                    meta={
                        'current': i, 
                        'total': total_rows, 
                        'status': f'Procesando plano {i}/{total_rows}: {row.get("nombre", "")}', 
                        'percent': int((i / total_rows) * 100)
                    }
                )
            
            # Importar fila (usando ModelInstanceLoader para IE 4.x)
            from import_export.instance_loaders import ModelInstanceLoader
            instance_loader = ModelInstanceLoader(resource, dataset)
            row_result = resource.import_row(row, instance_loader, row_number=i, dry_run=False)
            result.append_row_result(row_result)
            
        except Exception as e:
            result.append_base_error(resources.Error(error=e, traceback=str(e), row=row))
            
    # Limpiar archivo
    try:
        from django.core.files.storage import default_storage
        if default_storage.exists(file_path):
            default_storage.delete(file_path)
    except:
        pass
        
    return {
        'status': 'completed',
        'total': total_rows,
        'new': result.totals.get('new', 0),
        'updated': result.totals.get('update', 0),
        'skipped': result.totals.get('skip', 0),
        'errors': len(result.base_errors) + len(result.row_errors()),
    }

@shared_task(bind=True)
def import_bienes_afectos_task(self, file_path, file_format, user_id=None, import_name="Importación Bienes Afectos", verification_mode=False, dry_run=False):
    """
    Tarea Celery para importar bienes afectos con seguimiento de progreso, soporte para verificación y dry-run.
    """
    from tablib import Dataset
    import os
    from django.contrib.auth.models import User
    from django.core.cache import cache
    from .models import BienAfecto
    from django.core.files.storage import default_storage
    from .resources import BienAfectoResource

    user = User.objects.get(id=user_id) if user_id else None
    cache_key = f"import_bienes_progress_{user_id}"

    # Leer el archivo
    try:
        with default_storage.open(file_path, 'rb') as f:
            file_content = f.read()
            if file_format == 'csv':
                dataset = Dataset().load(try_decode(file_content), format='csv')
            elif file_format in ['xls', 'xlsx']:
                dataset = Dataset().load(file_content, format=file_format)
            else:
                raise ValueError(f"Formato no soportado: {file_format}")
    except Exception as e:
        error_res = {'status': 'error', 'message': str(e)}
        cache.set(cache_key, error_res, 3600)
        return error_res

    total_rows = len(dataset)
    resource = BienAfectoResource()
    
    # Estado inicial
    progress_info = {
        'status': 'Verificando archivo...' if verification_mode else 'Procesando importación...', 
        'percent': 0,
        'verification_mode': verification_mode,
        'dry_run': dry_run,
        'total': total_rows
    }
    cache.set(cache_key, progress_info, 3600)
    self.update_state(state='PROGRESS', meta=progress_info)

    if verification_mode:
        dry_run = True

    resource.before_import(dataset, user=user)
    
    try:
        result = resource.import_data(dataset, dry_run=dry_run, raise_errors=False)

        detailed_messages = []
        detailed_errors = []
        
        # Procesar errores globales
        for error in result.base_errors:
            detailed_errors.append(f"Error General: {str(error.error)}")

        dataset_dict = dataset.dict
        
        for i, row_result in enumerate(result.rows):
            row_idx = i + 1
            data_row = dataset_dict[i]
            nombre = str(data_row.get('nombre') or 'S/N').strip()
            codigo = str(data_row.get('codigo_interno') or '---').strip()
            activo_vinculado = str(data_row.get('activo_actual_codigo') or '').strip()
            
            action = "SIN CAMBIOS"
            details = ""
            
            if row_result.import_type == 'new':
                action = "NUEVO"
                details = "Se creará un nuevo bien afecto."
            elif row_result.import_type == 'update':
                action = "ACTUALIZAR"
                if hasattr(row_result, 'changed_fields') and row_result.changed_fields:
                    details = f"Cambios en: {', '.join(row_result.changed_fields)}"
                else:
                    details = "Actualización detectada."
            elif row_result.import_type == 'error':
                action = "ERROR"
                err_msg = str(row_result.errors[0].error) if row_result.errors else "Error desconocido"
                details = f"{err_msg}"
                detailed_errors.append(f"Fila {row_idx}: {err_msg}")
            
            if verification_mode or dry_run:
                vinculo = f" [Activo: {activo_vinculado}]" if activo_vinculado else ""
                msg = f"Fila {row_idx}: {codigo} ({nombre}){vinculo} -> [{action}] {details}"
                detailed_messages.append(msg)
        
        from celery import states
        final_res = {
            'status': 'completed',
            'state': states.SUCCESS,
            'total': total_rows,
            'new': result.totals.get('new', 0),
            'updated': result.totals.get('update', 0),
            'skipped': result.totals.get('skip', 0),
            'errors': len(detailed_errors),
            'error_list': detailed_errors,
            'results': detailed_messages,
            'verification_mode': verification_mode,
            'dry_run': dry_run,
            'file_path': file_path
        }
        
    except Exception as e:
        final_res = {'status': 'error', 'message': f"Error crítico: {str(e)}"}

    if not dry_run and not verification_mode:
        try:
            if default_storage.exists(file_path):
                default_storage.delete(file_path)
        except:
            pass
        
    cache.set(cache_key, final_res, 3600)
    return final_res
@shared_task(bind=True)
def import_submittals_task(self, file_path, file_format, user_id=None, import_name="Importación Submittals"):
    """
    Tarea Celery para importar CONTROL DE SUBMITTALS con seguimiento de progreso real y registro histórico.
    """
    from tablib import Dataset
    from .resources import ControlSubmittalResource
    from import_export import resources
    from django.core.files.storage import default_storage
    from .models import RegistroImportacion
    from django.contrib.auth.models import User
    from django.core.cache import cache

    user = User.objects.get(id=user_id) if user_id else None
    cache_key = f"import_submittals_progress_{user_id}" if user_id else "import_submittals_progress_system"
    
    registro = RegistroImportacion.objects.create(
        nombre=import_name,
        tipo='Submittals',
        usuario=user,
        estado='PROCESANDO'
    )
    
    # Leer archivo
    try:
        with default_storage.open(file_path, 'rb') as f:
            file_content = f.read()
            if file_format == 'csv':
                dataset = Dataset().load(try_decode(file_content), format='csv')
            elif file_format in ['xls', 'xlsx']:
                dataset = Dataset().load(file_content, format=file_format)
            elif file_format == 'json':
                dataset = Dataset().load(file_content, format='json')
            else:
                raise ValueError(f"Formato no soportado: {file_format}")
    except Exception as e:
        registro.estado = 'ERROR'
        registro.detalles_error = str(e)
        registro.save()
        return {'status': 'error', 'message': f'Error al leer archivo: {str(e)}'}

    total_rows = len(dataset)
    registro.total_rows = total_rows
    registro.save()
    
    resource = ControlSubmittalResource()
    result = resources.Result()
    
    for i, row in enumerate(dataset.dict, start=1):
        try:
            if i % 5 == 0 or i == total_rows:
                progress = int((i / total_rows) * 100)
                meta = {
                    'current': i, 
                    'total': total_rows, 
                    'status': f'Procesando submittal {i}/{total_rows}', 
                    'percent': progress
                }
                self.update_state(state='PROGRESS', meta=meta)
                cache.set(cache_key, meta, 3600)
            
            from import_export.instance_loaders import ModelInstanceLoader
            instance_loader = ModelInstanceLoader(resource, dataset)
            row_result = resource.import_row(row, instance_loader, row_number=i, dry_run=False)
            result.append_row_result(row_result)
        except Exception as e:
            result.append_base_error(resources.Error(error=e, traceback=str(e), row=row))
            
    # Actualizar registro final
    registro.filas_nuevas = result.totals.get('new', 0)
    registro.filas_actualizadas = result.totals.get('update', 0)
    registro.filas_omitidas = result.totals.get('skip', 0)
    registro.filas_error = len(result.base_errors) + len(result.row_errors())
    registro.estado = 'COMPLETADO'
    registro.save()

    # Limpiar archivo temporal
    try:
        if default_storage.exists(file_path):
            default_storage.delete(file_path)
    except:
        pass
        
    final_res = {
        'status': 'completed',
        'total': total_rows,
        'new': result.totals.get('new', 0),
        'updated': result.totals.get('update', 0),
        'skipped': result.totals.get('skip', 0),
        'errors': registro.filas_error,
    }
    cache.set(cache_key, final_res, 3600)
    return final_res

@shared_task(bind=True)
def import_categorias_task(self, file_path, file_format, user_id=None, import_name="Importación Categorías Activos"):
    """
    Tarea Celery para importar CATEGORIAS con seguimiento de progreso real.
    """
    from tablib import Dataset
    from django.core.files.storage import default_storage
    from django.core.cache import cache
    from .resources import CategoriaResource
    from .models import RegistroImportacion
    from django.contrib.auth.models import User
    
    user = User.objects.get(id=user_id) if user_id else None
    
    # Crear registro de importación
    registro = RegistroImportacion.objects.create(
        nombre=import_name,
        tipo='Categorías (Activos)',
        usuario=user,
        estado='PROCESANDO'
    )
    
    # Inicializar resource
    resource = CategoriaResource()
    
    # Marcador de progreso en caché
    cache_key = f"import_categorias_progress_{user_id}" if user_id else "import_categorias_progress_system"

    # Leer archivo
    try:
        with default_storage.open(file_path, 'rb') as f:
            file_content = f.read()
            if file_format == 'csv':
                dataset = Dataset().load(try_decode(file_content), format='csv')
            elif file_format in ['xls', 'xlsx']:
                dataset = Dataset().load(file_content, format=file_format)
            else:
                raise ValueError(f"Formato no soportado: {file_format}")
    except Exception as e:
        registro.estado = 'ERROR'
        registro.detalles_error = f'Error al leer archivo: {str(e)}'
        registro.save()
        error_res = {'status': 'error', 'message': f'Error al leer archivo: {str(e)}'}
        cache.set(cache_key, error_res, 3600)
        return error_res

    total_rows = len(dataset)
    registro.total_filas = total_rows
    registro.save()
    
    # Estado inicial
    progress_info = {
        'current': 0, 
        'total': total_rows, 
        'status': 'Iniciando importacion...', 
        'percent': 0,
        'new': 0,
        'updated': 0,
        'skipped': 0,
        'errors': 0
    }
    cache.set(cache_key, progress_info, 3600)
    self.update_state(state='PROGRESS', meta=progress_info)

    try:
        result = resource.import_data(dataset, dry_run=False, raise_errors=False)
        
        detailed_errors = []
        for error in result.base_errors:
            detailed_errors.append(f"Error General: {str(error.error)}")
        for line, errors in result.row_errors():
            for error in errors:
                detailed_errors.append(f"Fila {line}: {str(error.error)}")

        # Actualizar registro con éxito
        registro.filas_nuevas = result.totals.get('new', 0)
        registro.filas_actualizadas = result.totals.get('update', 0)
        registro.filas_omitidas = result.totals.get('skip', 0)
        registro.filas_error = len(detailed_errors)
        registro.estado = 'COMPLETADO'
        if detailed_errors:
            registro.detalles_error = "\n".join(detailed_errors[:50])
        registro.save()

        final_res = {
            'status': 'completed',
            'status_code': 'completed',
            'total': total_rows,
            'new': registro.filas_nuevas,
            'updated': registro.filas_actualizadas,
            'skipped': registro.filas_omitidas,
            'errors': registro.filas_error,
            'error_list': detailed_errors
        }
    except Exception as e:
        error_msg = f"Error crítico durante el procesamiento: {str(e)}"
        registro.estado = 'ERROR'
        registro.detalles_error = error_msg
        registro.save()
        final_res = {'status': 'error', 'message': error_msg}

    # Limpiar archivo original
    try:
        if default_storage.exists(file_path):
            default_storage.delete(file_path)
    except:
        pass
        
    cache.set(cache_key, final_res, 3600)
    return final_res

@shared_task(bind=True)
def generar_reporte_activos_task(self, reporte_id, filtros=None, query_ids=None):
    import io
    import json
    from django.utils import timezone
    from django.core.files.base import ContentFile
    from .models import Activo, ReporteGenerado
    
    try:
        reporte = ReporteGenerado.objects.get(id=reporte_id)
    except ReporteGenerado.DoesNotExist:
        return {'status': 'error', 'message': 'Reporte no encontrado'}
        
    reporte.estado = 'PROCESANDO'
    reporte.save()
    
    try:
        qs = Activo.objects.select_related('modelo__marca', 'ubicacion', 'familia').all()
        
        if query_ids:
            qs = qs.filter(id__in=query_ids)
        elif filtros:
            # Replicar lógica del Super Filtro
            if type(filtros) == str:
                filtros = json.loads(filtros)
            if filtros.get('estado'):
                qs = qs.filter(estado__in=filtros['estado'])
            if filtros.get('ubicacion'):
                from .views_celery import _get_descendant_ids
                from .models import Ubicacion
                expanded_ids = _get_descendant_ids(Ubicacion, filtros['ubicacion'])
                qs = qs.filter(ubicacion_id__in=expanded_ids)
            if filtros.get('familia'):
                qs = qs.filter(familia_id__in=filtros['familia'])
            if filtros.get('marca'):
                qs = qs.filter(modelo__marca_id__in=filtros['marca'])
            if filtros.get('modelo'):
                qs = qs.filter(modelo_id__in=filtros['modelo'])
            if filtros.get('busqueda'):
                q = filtros['busqueda']
                from django.db.models import Q
                qs = qs.filter(Q(nombre__icontains=q) | Q(codigo_interno__icontains=q) | Q(serie__icontains=q))
                
        qs = qs.order_by('-creado_en')
        
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Activos Filtrados"
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        headers = ['Código', 'Nombre', 'Estado', 'Ubicación', 'Familia', 'Marca', 'Modelo', 'Serie']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        for row_idx, a in enumerate(qs.iterator(), 2):
            ws.cell(row=row_idx, column=1, value=a.codigo_interno).border = thin_border
            ws.cell(row=row_idx, column=2, value=a.nombre).border = thin_border
            ws.cell(row=row_idx, column=3, value=a.get_estado_display()).border = thin_border
            # En Celery a veces "get_ruta_completa" hace llamadas a DB extra, cuidado con N+1.
            # a.ubicacion y a.familia van con _str_ local.
            ws.cell(row=row_idx, column=4, value=a.ubicacion.ruta_completa if getattr(a.ubicacion, 'ruta_completa', None) else (a.ubicacion.nombre if a.ubicacion else '')).border = thin_border
            ws.cell(row=row_idx, column=5, value=str(a.familia) if a.familia else '').border = thin_border
            ws.cell(row=row_idx, column=6, value=a.modelo.marca.nombre if a.modelo and a.modelo.marca else '').border = thin_border
            ws.cell(row=row_idx, column=7, value=a.modelo.nombre if a.modelo else '').border = thin_border
            ws.cell(row=row_idx, column=8, value=a.serie or '').border = thin_border
            
        # Auto-width
        for col in ws.columns:
            max_len = max((min(len(str(cell.value or '')), 50) for cell in col), default=10) + 2
            ws.column_dimensions[col[0].column_letter].width = max_len
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Save file to MinIO
        file_name = f"exportacion_{reporte.id}.xlsx"
        reporte.archivo.save(file_name, ContentFile(output.read()))
        
        reporte.estado = 'COMPLETADO'
        reporte.completado_en = timezone.now()
        reporte.save()
        return {'status': 'completed', 'reporte_id': reporte.id}
    except Exception as e:
        reporte.estado = 'ERROR'
        reporte.detalles_error = str(e)
        reporte.save()
        return {'status': 'error', 'message': str(e)}
