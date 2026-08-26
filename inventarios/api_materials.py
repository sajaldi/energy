from decimal import Decimal
from datetime import datetime as _datetime
from collections import defaultdict
from django.http import JsonResponse
from django.db.models import Q
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.shortcuts import get_object_or_404
from inventarios.models import Material, CategoriaMaterial, StockRecord
from presupuestos.models import ArticuloRequisicion

@login_required
def api_list_materials(request):
    """
    API endpoint to list materials with pagination and filtering.
    """
    query = request.GET.get('q', '')
    category_id = request.GET.get('category')
    ubicacion_id = request.GET.get('ubicacion')
    page_number = request.GET.get('page', 1)
    
    from django.db.models import Sum

    materials = Material.objects.only(
        'id', 'nombre', 'sku', 'descripcion', 'unidad_medida', 'precio_estimado',
        'categoria', 'tipo_material', 'imagen', 'ancho', 'alto', 'peso', 'profundidad'
    )

    if query:
        # Búsqueda avanzada:
        # - ? entre términos = AND (debe contener TODOS, orden no importa)
        # - , entre términos = OR (contiene cualquiera)
        # - // antes de un término = excluir
        # - Insensible a acentos (unaccent)
        # Ejemplo: "etiqueta ? térmica" busca items que contengan ambos
        # Ejemplo: "cable, tubo" busca items que contengan cable O tubo
        # Ejemplo: "PVC ? 40 //codo" busca PVC y 40 pero excluye "codo"
        import re
        import unicodedata

        def strip_accents(text):
            """Quita acentos de un string para búsqueda insensible."""
            nfkd = unicodedata.normalize('NFKD', text)
            return ''.join(c for c in nfkd if not unicodedata.category(c).startswith('M'))

        def term_q(term):
            """Genera Q para un término, buscando con y sin acentos."""
            t = term.strip()
            t_no_accent = strip_accents(t)
            q = Q(sku__icontains=t) | Q(nombre__icontains=t) | Q(descripcion__icontains=t)
            if t != t_no_accent:
                q |= Q(sku__icontains=t_no_accent) | Q(nombre__icontains=t_no_accent) | Q(descripcion__icontains=t_no_accent)
            return q

        # Primero separar exclusiones (delimitadas por //)
        parts = re.split(r'//', query)
        main_query = parts[0].strip()
        exclusions = [p.strip() for p in parts[1:] if p.strip()]

        # Construir filtro de exclusión
        exclude_q = Q()
        for ex in exclusions:
            for sub in re.split(r'[,?]', ex):
                sub = sub.strip()
                if sub:
                    exclude_q |= term_q(sub)

        # Procesar la query principal
        if ',' in main_query:
            # Coma = OR entre grupos
            or_groups = main_query.split(',')
            include_q = Q()
            for group in or_groups:
                group = group.strip()
                if not group:
                    continue
                if '?' in group:
                    # Dentro del grupo, ? = AND
                    and_terms = [t.strip() for t in group.split('?') if t.strip()]
                    group_q = Q()
                    for term in and_terms:
                        group_q &= term_q(term)
                    include_q |= group_q
                else:
                    include_q |= term_q(group)
        elif '?' in main_query:
            # Solo ? = AND entre todos los términos
            and_terms = [t.strip() for t in main_query.split('?') if t.strip()]
            include_q = Q()
            for term in and_terms:
                include_q &= term_q(term)
        else:
            # Búsqueda simple
            include_q = term_q(main_query)

        if include_q:
            materials = materials.filter(include_q)
        if exclude_q:
            materials = materials.exclude(exclude_q)
        
    if category_id:
        materials = materials.filter(categoria_id=category_id)

    materials = materials.order_by('nombre')

    # Paginación eficiente: solo contamos IDs (sin joins ni anotaciones)
    page_size = 30
    try:
        page_num = int(page_number)
    except (TypeError, ValueError):
        page_num = 1

    # Obtener un item más para saber si hay siguiente página
    offset = (page_num - 1) * page_size
    page_ids = list(materials.values_list('id', flat=True)[offset:offset + page_size + 1])
    has_next = len(page_ids) > page_size
    page_ids = page_ids[:page_size]

    # Ahora traemos los materiales completos solo para los IDs de la página
    materials_list = list(
        Material.objects.filter(id__in=page_ids)
        .select_related('categoria', 'unidad_medida')
        .only('id', 'nombre', 'sku', 'descripcion', 'unidad_medida', 'precio_estimado',
              'categoria', 'tipo_material', 'imagen', 'ancho', 'alto', 'peso', 'profundidad')
        .order_by('nombre')
    )
    material_ids = [m.id for m in materials_list]

    # Stock por lote (solo para los materiales de esta página)
    from django.db.models import Sum as DSum
    stock_map = {}
    stock_qs = StockRecord.objects.filter(material_id__in=material_ids)
    if ubicacion_id:
        stock_qs = stock_qs.filter(ubicacion_id=ubicacion_id)
    for row in stock_qs.values('material_id').annotate(total=DSum('cantidad')):
        stock_map[row['material_id']] = float(row['total'] or 0)
    
    # Batch: departamentos
    dept_map = defaultdict(set)
    for mid, did in Material.departamentos.through.objects.filter(
        material_id__in=material_ids
    ).values_list('material_id', 'departamento_id'):
        dept_map[mid].add(did)
    
    # Batch: bodegas con stock
    bodegas_map = defaultdict(list)
    for sr in StockRecord.objects.filter(
        material_id__in=material_ids, cantidad__gt=0
    ).values('material_id', 'ubicacion__nombre', 'ubicacion_especifica'):
        label = f"{sr['ubicacion__nombre']} · {sr['ubicacion_especifica']}" if sr['ubicacion_especifica'] else sr['ubicacion__nombre']
        bodegas_map[sr['material_id']].append(label)
    
    data = []
    
    user_depto_id = None
    if hasattr(request.user, 'perfil') and request.user.perfil.departamento_id:
        user_depto_id = request.user.perfil.departamento_id
        
    for m in materials_list:
        if m.imagen:
             image_url = m.imagen.url
        else:
             image_url = "data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='100' height='100' viewBox='0 0 24 24' fill='none' stroke='%233b82f6' stroke-width='1.5' stroke-linecap='round' stroke-linejoin='round'%3E%3Cpath d='M21 16V8a2 2 0 0 0-1-1.73l-7-4a2 2 0 0 0-2 0l-7 4A2 2 0 0 0 3 8v8a2 2 0 0 0 1 1.73l7 4a2 2 0 0 0 2 0l7-4A2 2 0 0 0 21 16z'%3E%3C/path%3E%3Cpolyline points='3.27 6.96 12 12.01 20.73 6.96'%3E%3C/polyline%3E%3Cline x1='12' y1='22.08' x2='12' y2='12'%3E%3C/line%3E%3C/svg%3E"

        dept_ids = dept_map.get(m.id, set())
        is_allowed = True
        if dept_ids and user_depto_id and user_depto_id not in dept_ids:
            is_allowed = False

        unique_bodegas = sorted(set(bodegas_map.get(m.id, [])))

        data.append({
            'id': m.id,
            'nombre': m.nombre,
            'sku': m.sku,
            'descripcion': m.descripcion or '',
            'unidad': m.unidad_medida.nombre if m.unidad_medida else 'Unidad',
            'precio_estimado': float(m.precio_estimado),
            'stock': stock_map.get(m.id, 0),
            'ancho': float(m.ancho) if m.ancho else None,
            'alto': float(m.alto) if m.alto else None,
            'peso': float(m.peso) if m.peso else None,
            'profundidad': float(m.profundidad) if m.profundidad else None,
            'categoria': m.categoria.nombre if m.categoria else 'General',
            'tipo_material': m.get_tipo_material_display() if hasattr(m, 'get_tipo_material_display') else m.tipo_material,
            'image_url': image_url,
            'is_allowed': is_allowed,
            'bodegas': unique_bodegas,
            'es_tecnico': m.es_tecnico,
        })
        
    return JsonResponse({
        'results': data,
        'has_next': has_next,
        'num_pages': 0,
        'current_page': page_num,
        'total': len(data)
    })

@login_required
def api_list_categories(request):
    categories = list(CategoriaMaterial.objects.all().order_by('nombre'))
    
    def build_tree(parent_id):
        nodes = []
        for cat in categories:
            if cat.padre_id == parent_id:
                children = build_tree(cat.id)
                nodes.append({
                    'id': cat.id,
                    'nombre': cat.nombre,
                    'children': children
                })
        return nodes

    tree = build_tree(None)
    return JsonResponse({'results': tree})

@login_required
def api_master_sync(request):
    """
    Consolidates ALL necessary data for offline mode (Basement Mode).
    Returns materials, categories, units, and warehouse locations.
    """
    from activos.models import Ubicacion
    from .models import UnidadMedida, StockRecord
    from django.db.models import Sum, OuterRef, Subquery, DecimalField
    from django.db.models.functions import Coalesce
    from decimal import Decimal

    # 1. Materials with stock
    stock_subquery = StockRecord.objects.filter(material=OuterRef('pk'))
    stock_total_expr = Subquery(
        stock_subquery.values('material').annotate(total=Sum('cantidad')).values('total'),
        output_field=DecimalField()
    )

    materials_qs = Material.objects.select_related('categoria', 'unidad_medida').annotate(
        stock_total=Coalesce(stock_total_expr, Decimal('0.00'))
    ).order_by('nombre')

    # Filtrado por departamento si aplica
    user_depto_id = None
    if hasattr(request.user, 'perfil') and request.user.perfil.departamento_id:
        user_depto_id = request.user.perfil.departamento_id

    materials_data = []
    for m in materials_qs:
        # Lógica de restricción básica
        dept_ids = [d.id for d in m.departamentos.all()]
        if dept_ids and user_depto_id and user_depto_id not in dept_ids:
            continue

        materials_data.append({
            'id': m.id,
            'nombre': m.nombre,
            'sku': m.sku,
            'desc': m.descripcion or '',
            'cat_id': m.categoria_id,
            'uni': m.unidad_medida.abreviatura if m.unidad_medida else 'UND',
            'stock': float(m.stock_total or 0)
        })

    # 2. Categories
    categories = list(CategoriaMaterial.objects.values('id', 'nombre', 'padre_id'))

    # 3. Units
    units = list(UnidadMedida.objects.values('id', 'nombre', 'abreviatura'))

    # 4. Locations (Warehouse only)
    locations = list(Ubicacion.objects.filter(tipo='ALMACEN').values('id', 'nombre'))

    return JsonResponse({
        'status': 'success',
        'materials': materials_data,
        'categories': categories,
        'units': units,
        'locations': locations,
        'timestamp': timezone.now().isoformat() if 'timezone' in globals() else None
    })


def _normalizar_texto(texto):
    """Minúsculas y sin acentos para comparar descripciones y nombres."""
    import unicodedata
    if not texto:
        return ''
    texto = unicodedata.normalize('NFKD', str(texto))
    texto = ''.join(c for c in texto if not unicodedata.combining(c))
    return texto.lower().strip()


_STOPWORDS = {
    'de', 'del', 'la', 'el', 'los', 'las', 'un', 'una', 'unos', 'unas',
    'para', 'por', 'con', 'sin', 'sobre', 'en', 'y', 'o', 'a', 'al', 'e',
    'material', 'nuevo', 'servicio', 'instalacion', 'instalación', 'se', 'su',
}


def _tokens_significativos(texto):
    return {t for t in texto.split() if len(t) >= 3 and t not in _STOPWORDS}


def _material_es_placeholder(material):
    if not material:
        return False
    nombre_norm = _normalizar_texto(material.nombre)
    return (
        'fuera de cat' in nombre_norm
        or 'sin cat' in nombre_norm
    )


def _articulo_coincide_con_material(descripcion, material):
    """Determina si un artículo 'huérfano' corresponde al material dado."""
    desc_norm = _normalizar_texto(descripcion)
    if not desc_norm:
        return False

    sku_norm = _normalizar_texto(material.sku)
    if sku_norm and sku_norm in desc_norm:
        return True

    nombre_norm = _normalizar_texto(material.nombre)
    if nombre_norm and nombre_norm in desc_norm:
        return True

    tokens_material = _tokens_significativos(nombre_norm)
    if not tokens_material:
        return False
    tokens_desc = _tokens_significativos(desc_norm)
    coinciden = tokens_material & tokens_desc
    return len(coinciden) / len(tokens_material) >= 0.5


@login_required
def api_precios_historicos(request, material_id):
    material = get_object_or_404(Material, pk=material_id)

    base = ArticuloRequisicion.objects.filter(
        cr8ca_costoaproximado__isnull=False
    ).exclude(
        cr8ca_costoaproximado=0
    )

    articulos_vinculados = base.filter(material=material)

    # Artículos 'huérfanos' (material NULL o vinculado a un placeholder) que
    # corresponden a este material por nombre/SKU.
    ids_placeholder = [
        m.id for m in Material.objects.filter(
            Q(nombre__icontains='fuera de cat') | Q(nombre__icontains='sin cat')
        ) if _material_es_placeholder(m)
    ]
    huérfanos_qs = base.filter(
        Q(material__isnull=True) | Q(material_id__in=ids_placeholder)
    ).select_related('requisicion', 'proveedor')

    articulos = articulos_vinculados.select_related('requisicion', 'proveedor')
    por_requisicion = []
    for a in huérfanos_qs:
        if _articulo_coincide_con_material(a.cr8ca_articulo, material):
            por_requisicion.append(a)

    por_requisicion.sort(
        key=lambda a: (a.createdon or getattr(a.requisicion, 'fecha', None)) or _datetime.min,
        reverse=True,
    )
    articulos = list(articulos) + por_requisicion

    # Deduplicar por PK y limitar a 50
    vistos = set()
    articulos_unicos = []
    for a in articulos:
        pk = str(a.pk)
        if pk in vistos:
            continue
        vistos.add(pk)
        articulos_unicos.append(a)
        if len(articulos_unicos) >= 50:
            break

    data = []
    for a in articulos_unicos:
        req = a.requisicion
        data.append({
            'req_num': req.cr8ca_requisicion if req else '—',
            'req_id': str(req.pk) if req else None,
            'fecha': a.createdon.strftime('%d/%m/%Y') if a.createdon else '—',
            'proveedor': a.proveedor.nombre if a.proveedor else '—',
            'cantidad': float(a.cr8ca_cantidad),
            'precio_unitario': float(a.cr8ca_costoaproximado),
            'subtotal': float(a.subtotal),
            'estado': req.get_estado_requisicion_display() if req else '—',
        })

    return JsonResponse({
        'material': material.nombre,
        'material_sku': material.sku,
        'precio_actual': float(material.precio_estimado or 0),
        'historial': data,
        'total_registros': len(data),
    })


@login_required
def api_create_material(request):
    """Crea un material rápido desde la requisición (nombre, sku opcional, precio)."""
    import json
    import uuid

    if request.method != 'POST':
        return JsonResponse({'error': 'POST requerido'}, status=405)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'JSON inválido'}, status=400)

    nombre = data.get('nombre', '').strip()
    if not nombre:
        return JsonResponse({'error': 'El nombre es requerido.'}, status=400)

    sku = data.get('sku', '').strip()
    precio = data.get('precio_estimado', 0)

    try:
        precio = float(precio) if precio else 0
    except (ValueError, TypeError):
        precio = 0

    # Generar SKU automático si no se proporcionó
    if not sku:
        sku = f"MAT-{uuid.uuid4().hex[:8].upper()}"

    # Verificar que el SKU no exista
    if Material.objects.filter(sku=sku).exists():
        return JsonResponse({'error': f'El SKU "{sku}" ya existe.'}, status=400)

    material = Material.objects.create(
        nombre=nombre,
        sku=sku,
        precio_estimado=precio,
    )

    return JsonResponse({
        'id': material.id,
        'nombre': material.nombre,
        'sku': material.sku,
        'precio_estimado': float(material.precio_estimado),
    })


@login_required
def api_material_detail(request, material_id):
    """Retorna todos los datos de un material para edición en modal."""
    material = get_object_or_404(Material, pk=material_id)
    
    data = {
        'id': material.id,
        'nombre': material.nombre,
        'sku': material.sku,
        'descripcion': material.descripcion or '',
        'precio_estimado': float(material.precio_estimado or 0),
        'stock_minimo': float(material.stock_minimo or 0),
        'tipo_material': material.tipo_material,
        'marca': material.marca.nombre if material.marca else '',
        'marca_id': material.marca_id,
        'categoria': material.categoria.nombre if material.categoria else '',
        'categoria_id': material.categoria_id,
        'unidad_medida': material.unidad_medida.nombre if material.unidad_medida else '',
        'unidad_medida_id': material.unidad_medida_id,
        'alto': float(material.alto) if material.alto else None,
        'ancho': float(material.ancho) if material.ancho else None,
        'peso': float(material.peso) if material.peso else None,
        'profundidad': float(material.profundidad) if material.profundidad else None,
        'no_afecta_stock': material.no_afecta_stock,
        'imagen_url': material.imagen.url if material.imagen else '',
        'stock_total': float(material.get_stock_total()),
        'creado_en': material.creado_en.strftime('%d/%m/%Y %H:%M') if material.creado_en else '',
        'actualizado_en': material.actualizado_en.strftime('%d/%m/%Y %H:%M') if material.actualizado_en else '',
        # Código de exoneración
        'codigo_exoneracion': material.codigo_exoneracion is not None,
        'codigo_exoneracion_id': material.codigo_exoneracion_id,
        'codigo_exoneracion_codigo': material.codigo_exoneracion.codigo if material.codigo_exoneracion else '',
        'codigo_exoneracion_desc': material.codigo_exoneracion.descripcion[:80] if material.codigo_exoneracion else '',
        'codigo_exoneracion_dai': float(material.codigo_exoneracion.dai) if material.codigo_exoneracion else 0,
        'codigo_exoneracion_isc': float(material.codigo_exoneracion.isc) if material.codigo_exoneracion else 0,
        'codigo_exoneracion_ipc': float(material.codigo_exoneracion.ipc) if material.codigo_exoneracion else 0,
        'codigo_exoneracion_isv': float(material.codigo_exoneracion.isv) if material.codigo_exoneracion else 0,
    }
    
    # Opciones para selects
    from activos.models import Marca
    data['opciones_tipo'] = Material.TIPO_MATERIAL_CHOICES
    data['opciones_categoria'] = list(CategoriaMaterial.objects.values('id', 'nombre').order_by('nombre'))
    data['opciones_unidad'] = list(
        Material.unidad_medida.field.related_model.objects.values('id', 'nombre').order_by('nombre')
    )
    
    return JsonResponse(data)


@login_required 
def api_material_update(request, material_id):
    """Actualiza un material desde el modal de edición."""
    import json
    
    if request.method != 'POST':
        return JsonResponse({'error': 'POST requerido'}, status=405)
    
    material = get_object_or_404(Material, pk=material_id)
    
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'JSON inválido'}, status=400)
    
    # Actualizar campos
    if 'nombre' in data and data['nombre'].strip():
        material.nombre = data['nombre'].strip()
    if 'sku' in data and data['sku'].strip():
        # Verificar unicidad
        if Material.objects.filter(sku=data['sku']).exclude(pk=material.pk).exists():
            return JsonResponse({'error': f'El SKU "{data["sku"]}" ya existe en otro material.'}, status=400)
        material.sku = data['sku'].strip()
    if 'descripcion' in data:
        material.descripcion = data['descripcion']
    if 'precio_estimado' in data:
        try:
            material.precio_estimado = float(data['precio_estimado']) if data['precio_estimado'] else 0
        except (ValueError, TypeError):
            pass
    if 'stock_minimo' in data:
        try:
            material.stock_minimo = float(data['stock_minimo']) if data['stock_minimo'] else 0
        except (ValueError, TypeError):
            pass
    if 'tipo_material' in data:
        material.tipo_material = data['tipo_material']
    if 'categoria_id' in data:
        material.categoria_id = data['categoria_id'] or None
    if 'unidad_medida_id' in data:
        material.unidad_medida_id = data['unidad_medida_id'] or None
    if 'alto' in data:
        material.alto = float(data['alto']) if data['alto'] else None
    if 'ancho' in data:
        material.ancho = float(data['ancho']) if data['ancho'] else None
    if 'peso' in data:
        material.peso = float(data['peso']) if data['peso'] else None
    if 'profundidad' in data:
        material.profundidad = float(data['profundidad']) if data['profundidad'] else None
    if 'no_afecta_stock' in data:
        material.no_afecta_stock = bool(data['no_afecta_stock'])
    if 'codigo_exoneracion_id' in data:
        material.codigo_exoneracion_id = data['codigo_exoneracion_id'] or None
    
    material.save()
    
    return JsonResponse({
        'success': True,
        'id': material.id,
        'nombre': material.nombre,
        'sku': material.sku,
        'precio_estimado': float(material.precio_estimado),
    })


@login_required
def api_search_codigos_exoneracion(request):
    """Busca códigos de exoneración por código o descripción, con totales de materiales."""
    query = request.GET.get('q', '').strip()
    
    from presupuestos.models import CodigoExoneracion, MaterialExoneracion
    from django.db.models import Sum, Count, F, DecimalField
    from django.db.models.functions import Coalesce
    
    if query == '*' or query == '':
        qs = CodigoExoneracion.objects.filter(activo=True)
    elif len(query) < 2:
        return JsonResponse({'results': []})
    else:
        qs = CodigoExoneracion.objects.filter(
            Q(codigo__icontains=query) | Q(descripcion__icontains=query),
            activo=True
        )

    results = qs.annotate(
        total_materiales=Count('materiales_solicitud'),
        total_cantidad=Coalesce(Sum('materiales_solicitud__cantidad'), 0, output_field=DecimalField()),
    ).order_by('codigo')[:50]

    # Calcular monto total (cantidad * precio del material)
    data = []
    for c in results:
        monto = 0
        if c.total_materiales > 0:
            items = MaterialExoneracion.objects.filter(codigo_exoneracion=c).select_related('material')
            monto = sum(float(i.cantidad * (i.material.precio_estimado or 0)) for i in items)
        
        data.append({
            'id': c.id,
            'codigo': c.codigo,
            'descripcion': c.descripcion[:100],
            'dai': float(c.dai),
            'isc': float(c.isc),
            'ipc': float(c.ipc),
            'isv': float(c.isv),
            'total_materiales': c.total_materiales,
            'total_cantidad': float(c.total_cantidad),
            'total_monto': monto,
        })

    return JsonResponse({'results': data})


@login_required
def api_export_materials_excel(request):
    """
    Exporta materiales a Excel (.xlsx) con los mismos filtros que el catálogo.
    Soporta: ?q=búsqueda&category=id
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.http import HttpResponse
    from django.db.models import Sum

    query = request.GET.get('q', '')
    category_id = request.GET.get('category')

    materials = Material.objects.select_related('categoria', 'unidad_medida', 'marca').all()

    if query:
        import re
        import unicodedata

        def strip_accents(text):
            nfkd = unicodedata.normalize('NFKD', text)
            return ''.join(c for c in nfkd if not unicodedata.category(c).startswith('M'))

        def term_q(term):
            t = term.strip()
            t_no_accent = strip_accents(t)
            q = Q(sku__icontains=t) | Q(nombre__icontains=t) | Q(descripcion__icontains=t)
            if t != t_no_accent:
                q |= Q(sku__icontains=t_no_accent) | Q(nombre__icontains=t_no_accent) | Q(descripcion__icontains=t_no_accent)
            return q

        # Misma lógica de búsqueda avanzada que api_list_materials
        parts = re.split(r'//', query)
        main_query = parts[0].strip()
        exclusions = [p.strip() for p in parts[1:] if p.strip()]

        exclude_q = Q()
        for ex in exclusions:
            for sub in re.split(r'[,?]', ex):
                sub = sub.strip()
                if sub:
                    exclude_q |= term_q(sub)

        if ',' in main_query:
            or_groups = main_query.split(',')
            include_q = Q()
            for group in or_groups:
                group = group.strip()
                if not group:
                    continue
                if '?' in group:
                    and_terms = [t.strip() for t in group.split('?') if t.strip()]
                    group_q = Q()
                    for term in and_terms:
                        group_q &= term_q(term)
                    include_q |= group_q
                else:
                    include_q |= term_q(group)
        elif '?' in main_query:
            and_terms = [t.strip() for t in main_query.split('?') if t.strip()]
            include_q = Q()
            for term in and_terms:
                include_q &= term_q(term)
        else:
            include_q = term_q(main_query)

        if include_q:
            materials = materials.filter(include_q)
        if exclude_q:
            materials = materials.exclude(exclude_q)

    if category_id:
        materials = materials.filter(categoria_id=category_id)

    materials = materials.order_by('nombre')

    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Materiales"

    # Estilos
    header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='0070F2', end_color='0070F2', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers
    headers = ['SKU', 'Nombre', 'Categoría', 'Tipo', 'Unidad', 'Marca', 'Precio Estimado', 'Stock Mínimo', 'Descripción']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data rows
    for row_idx, mat in enumerate(materials[:10000], 2):  # Limitar a 10k para evitar timeout
        ws.cell(row=row_idx, column=1, value=mat.sku or '')
        ws.cell(row=row_idx, column=2, value=mat.nombre or '')
        ws.cell(row=row_idx, column=3, value=mat.categoria.nombre if mat.categoria else '')
        ws.cell(row=row_idx, column=4, value=mat.get_tipo_material_display() if mat.tipo_material else '')
        ws.cell(row=row_idx, column=5, value=mat.unidad_medida.abreviatura if mat.unidad_medida else '')
        ws.cell(row=row_idx, column=6, value=mat.marca.nombre if mat.marca else '')
        ws.cell(row=row_idx, column=7, value=float(mat.precio_estimado) if mat.precio_estimado else 0)
        ws.cell(row=row_idx, column=8, value=float(mat.stock_minimo) if mat.stock_minimo else 0)
        ws.cell(row=row_idx, column=9, value=(mat.descripcion or '')[:200])

        for col in range(1, 10):
            ws.cell(row=row_idx, column=col).border = thin_border

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    # Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f'materiales_catalogo.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
