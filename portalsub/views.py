import calendar
import io
import re
import openpyxl
from datetime import date
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_POST
from django.utils import timezone
from django.db.models import Count, Q
from mantenimiento.models import Empresa, DocumentoEmpresa, TecnicoPuesto, PuestoTrabajo
from presupuestos.models import OrdenCompra
from notificaciones.utils import crear_notificacion, notificar_a_grupo
from .models import (
    PerfilContratista, ExpedienteMensual, DocumentoOrdenCompra, DocumentoPersonal,
    TipoDocumentoPersonal, TipoEntregable, EntregableContratista, DocumentoEntregable,
    HistorialPersonal,
)
from .decorators import contratista_required, get_empresa


def login_view(request):
    if request.user.is_authenticated:
        if hasattr(request.user, 'perfil_contratista'):
            return redirect('portalsub:dashboard')
        return redirect('admin:index')

    error = None
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            if hasattr(user, 'perfil_contratista') and user.perfil_contratista.activo:
                login(request, user)
                return redirect('portalsub:dashboard')
            error = 'No tienes acceso al portal de subcontratistas.'
        else:
            error = 'Usuario o contraseña incorrectos.'

    return render(request, 'portalsub/login.html', {'error': error})


def logout_view(request):
    logout(request)
    return redirect('portalsub:login')


@contratista_required
def dashboard(request):
    empresa = get_empresa(request)
    today = date.today()

    expediente_actual, _ = ExpedienteMensual.objects.get_or_create(
        empresa=empresa,
        mes=today.month,
        anio=today.year,
        defaults={'estado': 'BORRADOR'}
    )

    docs = DocumentoEmpresa.objects.filter(empresa=empresa)
    tipos_requeridos = ['RTN', 'PLANILLA_IHSS', 'ALTAS_BAJAS', 'EXPEDIENTE_MENSUAL', 'REPORTES']
    docs_completos = docs.filter(es_valido=True).values_list('tipo_documento', flat=True).distinct()

    docs_mes_actual = docs.filter(es_valido=True, mes=today.month, anio=today.year)
    tipos_mes_subidos = set(docs_mes_actual.values_list('tipo_documento', flat=True))
    tipos_mes_requeridos = ['PLANILLA_IHSS', 'ALTAS_BAJAS', 'EXPEDIENTE_MENSUAL', 'REPORTES']
    progreso_mes = int(len(tipos_mes_subidos & set(tipos_mes_requeridos)) / len(tipos_mes_requeridos) * 100) if tipos_mes_requeridos else 0

    total_personal = TecnicoPuesto.objects.filter(empresa=empresa).count()
    vigentes = TecnicoPuesto.objects.filter(empresa=empresa, esta_vigente=True).count()

    meses = ExpedienteMensual.objects.filter(empresa=empresa).order_by('-anio', '-mes')[:12]

    configs = EntregableContratista.objects.filter(
        empresa=empresa, activo=True, tipo_entregable__activo=True
    ).select_related('tipo_entregable')

    docs_entregables = DocumentoEntregable.objects.filter(empresa=empresa, anio=today.year)
    doc_map = {}
    for d in docs_entregables:
        doc_map[(d.tipo_entregable_id, d.mes)] = d

    total_aplican = 0
    total_completos = 0
    for c in configs:
        for m in c.get_meses():
            total_aplican += 1
            doc = doc_map.get((c.tipo_entregable_id, m))
            if doc and doc.es_valido:
                total_completos += 1

    progreso_anual = int(total_completos / total_aplican * 100) if total_aplican else 0

    context = {
        'active_tab': 'dashboard',
        'empresa': empresa,
        'expediente_actual': expediente_actual,
        'progreso_mes': progreso_mes,
        'progreso_anual': progreso_anual,
        'entregables_completos': total_completos,
        'entregables_totales': total_aplican,
        'docs_completos': len(set(docs_completos) & set(tipos_requeridos)),
        'docs_totales': len(tipos_requeridos),
        'total_personal': total_personal,
        'personal_vigente': vigentes,
        'personal_no_vigente': total_personal - vigentes,
        'meses': meses,
        'mes_actual': today.month,
        'anio_actual': today.year,
    }
    return render(request, 'portalsub/dashboard.html', context)


@contratista_required
def expediente(request, mes=None, anio=None):
    empresa = get_empresa(request)
    today = date.today()
    anio = anio or today.year

    if not EntregableContratista.objects.filter(empresa=empresa).exists():
        tipos = TipoEntregable.objects.filter(activo=True)
        EntregableContratista.objects.bulk_create([
            EntregableContratista(empresa=empresa, tipo_entregable=t)
            for t in tipos
        ], ignore_conflicts=True)

    configs = EntregableContratista.objects.filter(
        empresa=empresa, activo=True,
        tipo_entregable__activo=True,
    ).select_related('tipo_entregable').order_by('tipo_entregable__nombre')

    meses_nombre = ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun', 'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']

    docs_subidos = DocumentoEntregable.objects.filter(
        empresa=empresa, anio=anio
    ).select_related('tipo_entregable')

    doc_map = {}
    for d in docs_subidos:
        key = (d.tipo_entregable_id, d.mes)
        doc_map[key] = d

    entregas = []
    for c in configs:
        meses = c.get_meses()
        fila = {'config': c, 'meses': {}}
        for m in range(1, 13):
            aplica = m in meses
            key = (c.tipo_entregable_id, m)
            doc = doc_map.get(key)
            fila['meses'][m] = {
                'aplica': aplica,
                'completo': doc is not None and doc.es_valido,
            }
        entregas.append(fila)

    total_aplican = sum(1 for f in entregas for m in f['meses'].values() if m['aplica'])
    total_completos = sum(1 for f in entregas for m in f['meses'].values() if m['aplica'] and m['completo'])
    progreso = int(total_completos / total_aplican * 100) if total_aplican else 0

    context = {
        'active_tab': 'expediente',
        'empresa': empresa,
        'anio': anio,
        'meses_nombre': meses_nombre,
        'entregas': entregas,
        'progreso': progreso,
        'total_aplican': total_aplican,
        'total_completos': total_completos,
    }
    return render(request, 'portalsub/expediente.html', context)


@contratista_required
def expediente_mes(request, mes, anio):
    empresa = get_empresa(request)
    today = date.today()
    anio = anio or today.year

    expediente_obj, _ = ExpedienteMensual.objects.get_or_create(
        empresa=empresa, mes=mes, anio=anio,
        defaults={'estado': 'BORRADOR'}
    )

    if not EntregableContratista.objects.filter(empresa=empresa).exists():
        tipos = TipoEntregable.objects.filter(activo=True)
        EntregableContratista.objects.bulk_create([
            EntregableContratista(empresa=empresa, tipo_entregable=t)
            for t in tipos
        ], ignore_conflicts=True)

    configs = EntregableContratista.objects.filter(
        empresa=empresa, activo=True,
        tipo_entregable__activo=True,
    ).select_related('tipo_entregable').order_by('tipo_entregable__nombre')

    items = []
    for c in configs:
        if mes not in c.get_meses():
            continue
        doc = DocumentoEntregable.objects.filter(
            empresa=empresa, tipo_entregable=c.tipo_entregable, mes=mes, anio=anio
        ).first()
        items.append({
            'config': c,
            'documento': doc,
            'completo': doc is not None and (doc.es_valido or doc.no_aplica),
            'no_aplica': doc.no_aplica if doc else False,
        })

    total_count = len(items)
    complete_count = sum(1 for i in items if i['completo'])
    progreso = int(complete_count / total_count * 100) if total_count else 0

    context = {
        'active_tab': 'expediente',
        'empresa': empresa,
        'expediente': expediente_obj,
        'items': items,
        'mes': mes,
        'anio': anio,
        'mes_nombre': calendar.month_name[mes].capitalize(),
        'progreso': progreso,
        'total_count': total_count,
        'complete_count': complete_count,
        'historial': expediente_obj.historial.all().order_by('-fecha')[:20],
    }
    return render(request, 'portalsub/expediente_mes.html', context)


@contratista_required
@require_POST
def subir_documento(request, mes, anio):
    empresa = get_empresa(request)
    tipo = request.POST.get('tipo_documento')
    archivo = request.FILES.get('archivo')
    descripcion = request.POST.get('descripcion', '')

    if not tipo or not archivo:
        return JsonResponse({'status': 'error', 'message': 'Faltan datos requeridos.'}, status=400)

    if tipo in dict(DocumentoEmpresa.TIPO_DOC_CHOICES):
        doc, created = DocumentoEmpresa.objects.update_or_create(
            empresa=empresa,
            tipo_documento=tipo,
            mes=mes if tipo in ['PLANILLA_IHSS', 'ALTAS_BAJAS', 'EXPEDIENTE_MENSUAL', 'REPORTES', 'OTRO'] else None,
            anio=anio if tipo in ['PLANILLA_IHSS', 'ALTAS_BAJAS', 'EXPEDIENTE_MENSUAL', 'REPORTES', 'OTRO'] else None,
            defaults={
                'archivo': archivo,
                'es_valido': True,
                'descripcion': descripcion,
            }
        )
        return JsonResponse({
            'status': 'success',
            'message': 'Documento subido correctamente.',
            'documento': {
                'id': doc.id,
                'tipo': doc.tipo_documento,
                'nombre': doc.archivo.name,
                'url': doc.archivo.url,
                'es_valido': doc.es_valido,
            }
        })

    return JsonResponse({'status': 'error', 'message': 'Tipo de documento inválido.'}, status=400)


@contratista_required
@require_POST
def eliminar_documento(request, doc_id):
    empresa = get_empresa(request)
    doc = get_object_or_404(DocumentoEmpresa, id=doc_id, empresa=empresa)
    doc.archivo.delete(save=False)
    doc.delete()
    return JsonResponse({'status': 'success', 'message': 'Documento eliminado.'})


@contratista_required
@require_POST
def subir_entregable(request):
    empresa = get_empresa(request)
    tipo_id = request.POST.get('tipo_entregable_id')
    mes = request.POST.get('mes')
    anio = request.POST.get('anio')
    archivo = request.FILES.get('archivo')

    if not tipo_id or not anio or not archivo:
        return JsonResponse({'status': 'error', 'message': 'Faltan datos requeridos.'}, status=400)

    try:
        anio = int(anio)
    except (ValueError, TypeError):
        return JsonResponse({'status': 'error', 'message': 'Año inválido.'}, status=400)

    tipo = get_object_or_404(TipoEntregable, pk=tipo_id)
    config = get_object_or_404(EntregableContratista, empresa=empresa, tipo_entregable=tipo, activo=True)

    try:
        mes_val = int(mes) if mes else None
    except (ValueError, TypeError):
        mes_val = None

    doc, created = DocumentoEntregable.objects.update_or_create(
        empresa=empresa,
        tipo_entregable=tipo,
        mes=mes_val,
        anio=anio,
        defaults={
            'archivo': archivo,
            'subido_por': request.user,
            'es_valido': True,
        }
    )
    return JsonResponse({
        'status': 'success',
        'message': 'Documento subido correctamente.',
        'documento': {
            'id': doc.id,
            'nombre': doc.archivo.name.split('/')[-1],
            'url': doc.archivo.url,
            'tipo': doc.tipo_entregable.nombre,
            'mes': doc.mes,
        }
    })


@contratista_required
@require_POST
def eliminar_entregable(request, doc_id):
    empresa = get_empresa(request)
    doc = get_object_or_404(DocumentoEntregable, pk=doc_id, empresa=empresa)
    doc.archivo.delete(save=False)
    doc.delete()
    return JsonResponse({'status': 'success', 'message': 'Documento eliminado.'})


@contratista_required
@require_POST
def enviar_expediente(request, mes, anio):
    empresa = get_empresa(request)
    expediente_obj = get_object_or_404(ExpedienteMensual, empresa=empresa, mes=mes, anio=anio)
    if expediente_obj.estado not in ('BORRADOR', 'RECHAZADO'):
        messages.warning(request, 'El expediente ya fue enviado o está en otro estado.')
        return redirect('portalsub:expediente_mes', mes=mes, anio=anio)

    # Validar que todos los entregables aplicables estén completos
    configs = EntregableContratista.objects.filter(
        empresa=empresa, activo=True,
        tipo_entregable__activo=True,
    ).select_related('tipo_entregable')
    incompletos = []
    for c in configs:
        if mes not in c.get_meses():
            continue
        doc = DocumentoEntregable.objects.filter(
            empresa=empresa, tipo_entregable=c.tipo_entregable,
            mes=mes, anio=anio,
        ).first()
        if not doc or (not doc.es_valido and not doc.no_aplica):
            incompletos.append(c.tipo_entregable.nombre)

    if incompletos:
        messages.error(
            request,
            f'No se puede enviar el expediente. Faltan los siguientes entregables: {", ".join(incompletos)}.'
        )
        return redirect('portalsub:expediente_mes', mes=mes, anio=anio)

    expediente_obj.estado = 'ENVIADO'
    expediente_obj.fecha_envio = timezone.now()
    expediente_obj.save()

    # Registrar en historial
    from .models import HistorialExpediente
    evento_tipo = 'REENVIADO' if request.POST.get('_reenvio') or expediente_obj.historial.filter(evento='ENVIADO').exists() else 'ENVIADO'
    HistorialExpediente.objects.create(
        expediente=expediente_obj,
        evento=evento_tipo,
        usuario=request.user,
        observaciones=f"Expediente {'reenviado' if evento_tipo == 'REENVIADO' else 'enviado'} para revisión."
    )

    notificar_a_grupo(
        grupo_nombre='Administradores',
        titulo="Expediente Mensual Enviado",
        mensaje=f"{empresa.nombre} ha enviado el expediente de {mes}/{anio} para revisión.",
        tipo='INFO',
        modulo='PORTAL_SUB',
        enlace=f"/admin/portalsub/expedientemensual/",
        icono='document-text-outline',
    )

    # Webhook Power Automate - notificar envío de expediente
    try:
        import requests as http_requests
        webhook_url = (
            "https://ce675e3ed2704594af019ed8d7d5f6.d7.environment.api.powerplatform.com:443"
            "/powerautomate/automations/direct/cu/09/workflows/946fb7325d714fdf87893231a5c7b6cb"
            "/triggers/manual/paths/invoke?api-version=1"
            "&sp=%2Ftriggers%2Fmanual%2Frun&sv=1.0"
            "&sig=S9C7DbiVeAms87D5ZOLkM0uj17TPdfLS-Vc68c0PpVE"
        )
        # Obtener correo del departamento del usuario
        correo_departamento = ''
        try:
            perfil = request.user.perfil
            if perfil.departamento and perfil.departamento.correo:
                correo_departamento = perfil.departamento.correo
        except Exception:
            pass

        payload = {
            "empresa": empresa.nombre,
            "mes": mes,
            "anio": anio,
            "estado": "ENVIADO",
            "expediente_url": f"https://softcom.ccg.hn/portalsub/expediente/{mes}/{anio}/",
            "fecha_envio": expediente_obj.fecha_envio.isoformat(),
            "enviado_por": request.user.get_full_name() or request.user.username,
            "correo_departamento": correo_departamento,
            "correo_usuario": request.user.email or '',
        }
        http_requests.post(webhook_url, json=payload, timeout=10)
    except Exception:
        pass  # No bloquear el flujo si el webhook falla

    return redirect('portalsub:expediente_mes', mes=mes, anio=anio)


@contratista_required
@require_POST
def toggle_no_aplica(request, mes, anio):
    """Toggle 'No Aplica' para un entregable específico."""
    import json
    from django.http import JsonResponse
    empresa = get_empresa(request)
    
    try:
        data = json.loads(request.body)
        tipo_id = data.get('tipo_entregable_id')
        no_aplica = data.get('no_aplica', True)
    except (json.JSONDecodeError, ValueError):
        tipo_id = request.POST.get('tipo_entregable_id')
        no_aplica = request.POST.get('no_aplica') == 'true'

    if not tipo_id:
        return JsonResponse({'status': 'error', 'message': 'Tipo de entregable requerido'}, status=400)

    doc, created = DocumentoEntregable.objects.get_or_create(
        empresa=empresa,
        tipo_entregable_id=tipo_id,
        mes=mes,
        anio=anio,
        defaults={'no_aplica': no_aplica, 'es_valido': True, 'subido_por': request.user}
    )
    if not created:
        doc.no_aplica = no_aplica
        doc.save()

    return JsonResponse({'status': 'success', 'no_aplica': doc.no_aplica})


@contratista_required
def personal_list(request):
    empresa = get_empresa(request)
    today = date.today()
    personal = TecnicoPuesto.objects.filter(empresa=empresa).order_by('-esta_vigente', 'apellido', 'nombre')

    q = request.GET.get('q', '').strip()
    if q:
        personal = personal.filter(
            Q(nombre__icontains=q) | Q(apellido__icontains=q) | Q(dni__icontains=q)
        )

    altas = personal.filter(
        fecha_alta__year=today.year, fecha_alta__month=today.month
    ).order_by('-fecha_alta')

    bajas = HistorialPersonal.objects.filter(
        tecnico__empresa=empresa, tipo='BAJA',
        fecha__year=today.year, fecha__month=today.month
    ).select_related('tecnico').order_by('-fecha')

    return render(request, 'portalsub/personal_list.html', {
        'active_tab': 'personal',
        'empresa': empresa,
        'personal': personal,
        'total': personal.count(),
        'vigentes': personal.filter(esta_vigente=True).count(),
        'altas': altas,
        'bajas': bajas,
        'total_tipos_doc': TipoDocumentoPersonal.objects.filter(activo=True).count(),
    })


@contratista_required
def personal_plantilla_dni(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'DNI'
    ws['A1'] = 'DNI'
    ws['A2'] = '0801-1990-12345'
    ws['A3'] = '0801-1985-67890'
    ws['A4'] = '0801-2000-11111'
    ws.column_dimensions['A'].width = 22
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="plantilla_dni.xlsx"'
    wb.save(response)
    return response


@contratista_required
def personal_verificar_pdf(request):
    empresa = get_empresa(request)
    resultado = None
    paso = 1

    if request.method == 'POST':
        archivo = request.FILES.get('archivo')
        if not archivo:
            messages.error(request, 'Debes seleccionar un archivo.')
            return redirect('portalsub:personal_verificar_pdf')

        if not archivo.name.lower().endswith('.pdf'):
            messages.error(request, 'Solo se admiten archivos .pdf')
            return redirect('portalsub:personal_verificar_pdf')

        try:
            import pypdf
            reader = pypdf.PdfReader(io.BytesIO(archivo.read()))
            texto = ''
            for page in reader.pages:
                texto += (page.extract_text() or '') + '\n'

            if not texto.strip():
                messages.error(request, 'No se pudo extraer texto del PDF.')
                return redirect('portalsub:personal_verificar_pdf')

            if 'INSTITUTO HONDUREÑO DE SEGURIDAD SOCIAL' not in texto.upper() or 'PLANILLA MENSUAL DE COTIZACION' not in texto.upper():
                messages.error(request, 'El PDF subido no corresponde a una planilla del IHSS.')
                return redirect('portalsub:personal_verificar_pdf')

            def norm(val):
                return val.replace('-', '').replace(' ', '').lower()

            periodo = ''
            lines = texto.split('\n')
            for i, line in enumerate(lines):
                if 'PERÍODO' in line.upper():
                    for j in range(i, min(i + 10, len(lines))):
                        candidate = lines[j].replace(' ', '')
                        if not candidate.strip():
                            continue
                        m = re.search(r'\d{1,2}[/\-\.]\d{1,2}[/\-\.]\d{2,4}', candidate)
                        if m:
                            periodo = m.group()
                            break
                    break
            dnis_pdf = []
            for match in re.finditer(r'\d{4}-?\d{4}-?\d{5}', texto):
                raw = match.group()
                dnis_pdf.append(raw)

            if not dnis_pdf:
                messages.error(request, 'No se encontraron números de IHSS en el PDF.')
                return redirect('portalsub:personal_verificar_pdf')

            dnis_pdf = list(dict.fromkeys(dnis_pdf))
            pdf_norm = {norm(d): d for d in dnis_pdf}

            vigentes_db = TecnicoPuesto.objects.filter(
                empresa=empresa, esta_vigente=True, dni__isnull=False
            ).exclude(dni__exact='')
            db_map = {}
            for p in vigentes_db:
                db_map[norm(p.dni)] = p

            coinciden = []
            no_encontrados = []
            en_otra_empresa = []
            for dni in dnis_pdf:
                key = norm(dni)
                if key in db_map:
                    coinciden.append(db_map[key])
                else:
                    no_encontrados.append(dni)

            # Verificar si los "no encontrados" están en otra empresa
            if no_encontrados:
                all_tecnicos = TecnicoPuesto.objects.filter(
                    dni__isnull=False
                ).exclude(empresa=empresa).select_related('empresa')
                otros_map = {}
                for t in all_tecnicos:
                    if t.dni:
                        otros_map[norm(t.dni)] = t
                
                no_encontrados_final = []
                for dni in no_encontrados:
                    key = norm(dni)
                    if key in otros_map:
                        tec = otros_map[key]
                        en_otra_empresa.append({
                            'dni': dni,
                            'nombre': f"{tec.nombre} {tec.apellido}",
                            'empresa': tec.empresa.nombre if tec.empresa else 'Desconocida',
                            'vigente': tec.esta_vigente,
                        })
                    else:
                        no_encontrados_final.append(dni)
                no_encontrados = no_encontrados_final

            faltantes = [p for p in vigentes_db if norm(p.dni) not in pdf_norm]

            resultado = {
                'total_pdf': len(dnis_pdf),
                'coinciden': len(coinciden),
                'coinciden_lista': coinciden,
                'no_encontrados': no_encontrados,
                'en_otra_empresa': en_otra_empresa,
                'faltantes': faltantes,
                'total_faltantes': len(faltantes),
            }

            request.session['pdf_dnis'] = dnis_pdf
            request.session['pdf_periodo'] = periodo
            paso = 3

        except Exception as e:
            messages.error(request, f'Error al leer el PDF: {e}')
            return redirect('portalsub:personal_verificar_pdf')

    pdf_dnis = request.session.get('pdf_dnis')
    if pdf_dnis and not resultado:
        paso = 2

    return render(request, 'portalsub/personal_verificar_pdf.html', {
        'active_tab': 'personal',
        'empresa': empresa,
        'resultado': resultado,
        'paso': paso,
        'periodo': request.session.get('pdf_periodo', ''),
    })


@contratista_required
def personal_reporte_pdf(request):
    empresa = get_empresa(request)
    dnis_pdf = request.session.get('pdf_dnis')
    if not dnis_pdf:
        messages.error(request, 'No hay datos de verificación. Debes verificar un PDF primero.')
        return redirect('portalsub:personal_verificar_pdf')

    def norm(val):
        return val.replace('-', '').replace(' ', '').lower()

    pdf_norm = {norm(d): d for d in dnis_pdf}

    vigentes_db = TecnicoPuesto.objects.filter(
        empresa=empresa, esta_vigente=True, dni__isnull=False
    ).exclude(dni__exact='')
    db_map = {}
    for p in vigentes_db:
        db_map[norm(p.dni)] = p

    coinciden = []
    no_encontrados = []
    for dni in dnis_pdf:
        key = norm(dni)
        if key in db_map:
            coinciden.append(db_map[key])
        else:
            no_encontrados.append(dni)

    faltantes = [p for p in vigentes_db if norm(p.dni) not in pdf_norm]

    wb = openpyxl.Workbook()
    ws_resumen = wb.active
    ws_resumen.title = 'Resumen'
    ws_resumen.append(['Indicador', 'Cantidad'])
    ws_resumen.append(['Total DNI en PDF', len(dnis_pdf)])
    ws_resumen.append(['Coincidencias', len(coinciden)])
    ws_resumen.append(['No encontrados en BD', len(no_encontrados)])
    ws_resumen.append(['Faltan en el PDF', len(faltantes)])
    ws_resumen.column_dimensions['A'].width = 25
    ws_resumen.column_dimensions['B'].width = 15

    if coinciden:
        ws_coinciden = wb.create_sheet('Coincidencias')
        ws_coinciden.append(['Nombre', 'Apellido', 'DNI', 'Estado'])
        for p in coinciden:
            ws_coinciden.append([p.nombre, p.apellido, p.dni, 'Vigente' if p.esta_vigente else 'No Vigente'])

    if no_encontrados:
        ws_no = wb.create_sheet('No encontrados')
        ws_no.append(['DNI'])
        for d in no_encontrados:
            ws_no.append([d])

    if faltantes:
        ws_faltan = wb.create_sheet('Faltan en PDF')
        ws_faltan.append(['Nombre', 'Apellido', 'DNI', 'Estado'])
        for p in faltantes:
            ws_faltan.append([p.nombre, p.apellido, p.dni, 'Vigente' if p.esta_vigente else 'No Vigente'])

    request.session['pdf_reporte_generado'] = True

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reporte_verificacion_dni.xlsx"'
    wb.save(response)
    return response


@contratista_required
@require_POST
def personal_guardar_cambios(request):
    empresa = get_empresa(request)
    vigentes_ids = request.POST.getlist('vigente_ids')
    TecnicoPuesto.objects.filter(empresa=empresa, pk__in=vigentes_ids).update(esta_vigente=True)
    TecnicoPuesto.objects.filter(empresa=empresa).exclude(pk__in=vigentes_ids).filter(
        pk__in=request.POST.getlist('faltante_ids')
    ).update(esta_vigente=False)
    messages.success(request, 'Cambios guardados correctamente.')
    return redirect('portalsub:personal_verificar_pdf')


@contratista_required
@require_POST
def personal_completar_verificacion(request):
    empresa = get_empresa(request)
    periodo = request.session.get('pdf_periodo', '')
    from datetime import datetime
    try:
        if '/' in periodo:
            parts = periodo.split('/')
            if len(parts) == 2:
                mes, anio = parts
                fecha_verif = datetime.strptime(f'01/{mes}/{anio}', '%d/%m/%Y')
            elif len(parts) == 3:
                fecha_verif = datetime.strptime(periodo, '%d/%m/%Y')
            else:
                fecha_verif = timezone.now()
        else:
            fecha_verif = timezone.now()
    except (ValueError, IndexError):
        fecha_verif = timezone.now()
    TecnicoPuesto.objects.filter(empresa=empresa, esta_vigente=True).update(ultima_verificacion=fecha_verif)
    messages.success(request, 'Verificación completada. Personal actualizado.')
    return redirect('portalsub:personal_reporte_verificacion')


@contratista_required
def personal_reporte_verificacion(request):
    empresa = get_empresa(request)
    dnis_pdf = request.session.get('pdf_dnis')
    if not dnis_pdf:
        messages.error(request, 'No hay datos de verificación. Debes verificar un PDF primero.')
        return redirect('portalsub:personal_verificar_pdf')

    def norm(val):
        return val.replace('-', '').replace(' ', '').lower()

    pdf_norm = {norm(d): d for d in dnis_pdf}

    vigentes_db = TecnicoPuesto.objects.filter(
        empresa=empresa, esta_vigente=True, dni__isnull=False
    ).exclude(dni__exact='')
    db_map = {}
    for p in vigentes_db:
        db_map[norm(p.dni)] = p

    coinciden = []
    no_encontrados = []
    for dni in dnis_pdf:
        key = norm(dni)
        if key in db_map:
            coinciden.append(db_map[key])
        else:
            no_encontrados.append(dni)

    faltantes = [p for p in vigentes_db if norm(p.dni) not in pdf_norm]

    ahora = timezone.now()

    response = render(request, 'portalsub/personal_reporte_verificacion.html', {
        'active_tab': 'personal',
        'empresa': empresa,
        'coinciden': coinciden,
        'no_encontrados': no_encontrados,
        'faltantes': faltantes,
        'total_pdf': len(dnis_pdf),
        'fecha': ahora,
        'periodo': request.session.get('pdf_periodo', ''),
    })
    return response


@contratista_required
def personal_importar_dni(request):
    empresa = get_empresa(request)
    resultado = None

    if request.method == 'POST':
        archivo = request.FILES.get('archivo')
        if not archivo:
            messages.error(request, 'Debes seleccionar un archivo.')
            return redirect('portalsub:personal_importar_dni')

        if not archivo.name.endswith('.xlsx'):
            messages.error(request, 'Solo se admiten archivos .xlsx')
            return redirect('portalsub:personal_importar_dni')

        try:
            wb = openpyxl.load_workbook(io.BytesIO(archivo.read()), read_only=True, data_only=True)
            ws = wb.active

            def norm(val):
                return val.replace('-', '').replace(' ', '').lower()

            dnis_archivo = []
            for row in ws.iter_rows(min_row=1, values_only=True):
                for cell in row:
                    if cell is not None:
                        val = str(cell).strip()
                        if val:
                            dnis_archivo.append(val)
                    break

            if not dnis_archivo:
                messages.error(request, 'El archivo está vacío o no contiene datos en la primera columna.')
                return redirect('portalsub:personal_importar_dni')

            dnis_archivo_norm = {norm(d): d for d in dnis_archivo}

            vigentes_db = TecnicoPuesto.objects.filter(
                empresa=empresa, esta_vigente=True, dni__isnull=False
            ).exclude(dni__exact='')
            db_map = {}
            for p in vigentes_db:
                db_map[norm(p.dni)] = p

            coinciden = []
            no_encontrados = []
            for dni_file in dnis_archivo:
                key = norm(dni_file)
                if key in db_map:
                    coinciden.append(db_map[key])
                else:
                    no_encontrados.append(dni_file)

            faltantes_en_archivo = [p for p in vigentes_db if norm(p.dni) not in dnis_archivo_norm]

            resultado = {
                'total_archivo': len(dnis_archivo),
                'coinciden': len(coinciden),
                'coinciden_lista': coinciden,
                'no_encontrados': no_encontrados,
                'faltantes_en_archivo': faltantes_en_archivo,
                'total_faltantes': len(faltantes_en_archivo),
            }

        except Exception as e:
            messages.error(request, f'Error al leer el archivo: {e}')
            return redirect('portalsub:personal_importar_dni')

    return render(request, 'portalsub/personal_importar_dni.html', {
        'active_tab': 'personal',
        'empresa': empresa,
        'resultado': resultado,
    })


@contratista_required
def personal_crear(request):
    empresa = get_empresa(request)
    puestos = PuestoTrabajo.objects.all().order_by('nombre')
    error = None
    dni_inicial = request.GET.get('dni', '')
    if request.method == 'POST':
        nombre = request.POST.get('nombre', '').strip()
        apellido = request.POST.get('apellido', '').strip()
        dni = request.POST.get('dni', '').strip()
        fecha_nacimiento = request.POST.get('fecha_nacimiento') or None
        tipo_sangre = request.POST.get('tipo_sangre', '').strip()
        fecha_alta = request.POST.get('fecha_alta') or None
        telefono = request.POST.get('telefono', '').strip()
        telefono_emergencia = request.POST.get('telefono_emergencia', '').strip()
        puesto_id = request.POST.get('puesto_id')
        puesto_en_empresa = request.POST.get('puesto_en_empresa', '').strip()
        foto = request.FILES.get('foto')

        if not nombre or not apellido:
            error = 'Nombre y apellido son obligatorios.'
        else:
            empleado = TecnicoPuesto.objects.create(
                empresa=empresa,
                nombre=nombre,
                apellido=apellido,
                dni=dni or None,
                fecha_nacimiento=fecha_nacimiento,
                tipo_sangre=tipo_sangre or None,
                fecha_alta=fecha_alta,
                telefono=telefono or None,
                telefono_emergencia=telefono_emergencia or None,
                puesto_id=puesto_id or None,
                puesto_en_empresa=puesto_en_empresa or None,
                foto=foto,
                disponible=True,
                esta_vigente=True,
            )
            HistorialPersonal.objects.create(
                tecnico=empleado, tipo='ALTA', usuario=request.user,
                detalle=f'Alta registrada por {request.user.get_full_name() or request.user.username}'
            )
            return redirect('portalsub:personal_detalle', pk=empleado.id)

    return render(request, 'portalsub/personal_form.html', {
        'active_tab': 'personal',
        'empresa': empresa,
        'puestos': puestos,
        'accion': 'Agregar',
        'error': error,
        'dni_inicial': dni_inicial,
    })


@contratista_required
def personal_editar(request, pk):
    empresa = get_empresa(request)
    empleado = get_object_or_404(TecnicoPuesto, pk=pk, empresa=empresa)
    old_empresa_id = empleado.empresa_id
    puestos = PuestoTrabajo.objects.all().order_by('nombre')
    empresas = Empresa.objects.all().order_by('nombre')
    error = None

    if request.method == 'POST':
        nombre = request.POST.get('nombre', '').strip()
        apellido = request.POST.get('apellido', '').strip()
        dni = request.POST.get('dni', '').strip()
        empresa_id = request.POST.get('empresa_id')
        fecha_nacimiento = request.POST.get('fecha_nacimiento') or None
        tipo_sangre = request.POST.get('tipo_sangre', '').strip()
        fecha_alta = request.POST.get('fecha_alta') or None
        telefono = request.POST.get('telefono', '').strip()
        telefono_emergencia = request.POST.get('telefono_emergencia', '').strip()
        puesto_id = request.POST.get('puesto_id')
        puesto_en_empresa = request.POST.get('puesto_en_empresa', '').strip()
        foto = request.FILES.get('foto')
        esta_vigente = request.POST.get('esta_vigente') == 'on'

        if not nombre or not apellido:
            error = 'Nombre y apellido son obligatorios.'
        else:
            empleado.nombre = nombre
            empleado.apellido = apellido
            empleado.dni = dni or None
            if empresa_id:
                empleado.empresa_id = empresa_id
            empleado.puesto_en_empresa = puesto_en_empresa or None
            empleado.fecha_nacimiento = fecha_nacimiento
            empleado.tipo_sangre = tipo_sangre or None
            empleado.fecha_alta = fecha_alta
            empleado.telefono = telefono or None
            empleado.telefono_emergencia = telefono_emergencia or None
            old_vigente = empleado.esta_vigente
            old_puesto = empleado.puesto_id
            empleado.puesto_id = puesto_id or None
            if foto:
                empleado.foto = foto
            empleado.esta_vigente = esta_vigente
            empleado.save()

            if old_vigente != esta_vigente:
                if esta_vigente:
                    HistorialPersonal.objects.create(
                        tecnico=empleado, tipo='REINGRESO', usuario=request.user,
                        detalle='Cambió de No Vigente a Vigente'
                    )
                else:
                    HistorialPersonal.objects.create(
                        tecnico=empleado, tipo='BAJA', usuario=request.user,
                        detalle='Cambió de Vigente a No Vigente'
                    )
            if old_puesto != empleado.puesto_id:
                HistorialPersonal.objects.create(
                    tecnico=empleado, tipo='CAMBIO_PUESTO', usuario=request.user,
                    detalle=f'Puesto anterior ID {old_puesto} → nuevo ID {empleado.puesto_id}'
                )
            if str(old_empresa_id) != str(empleado.empresa_id):
                HistorialPersonal.objects.create(
                    tecnico=empleado, tipo='CAMBIO_PUESTO', usuario=request.user,
                    detalle=f'Empresa anterior ID {old_empresa_id} → nueva ID {empleado.empresa_id}'
                )
            return redirect('portalsub:personal_detalle', pk=empleado.id)

    return render(request, 'portalsub/personal_form.html', {
        'active_tab': 'personal',
        'empresa': empresa,
        'empleado': empleado,
        'puestos': puestos,
        'empresas': empresas,
        'accion': 'Editar',
        'error': error,
        'dni_inicial': '',
    })


@contratista_required
@require_POST
def personal_toggle_vigente(request, pk):
    empresa = get_empresa(request)
    empleado = get_object_or_404(TecnicoPuesto, pk=pk)
    empleado.esta_vigente = not empleado.esta_vigente
    empleado.save(update_fields=['esta_vigente'])
    HistorialPersonal.objects.create(
        tecnico=empleado, tipo='REINGRESO' if empleado.esta_vigente else 'BAJA',
        usuario=request.user,
        detalle=f'Cambió a {"Vigente" if empleado.esta_vigente else "No Vigente"} desde importación DNI'
    )
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({'vigente': empleado.esta_vigente})
    return redirect(request.META.get('HTTP_REFERER', 'portalsub:personal_importar_dni'))


@contratista_required
@require_POST
def personal_eliminar(request, pk):
    empresa = get_empresa(request)
    empleado = get_object_or_404(TecnicoPuesto, pk=pk)
    HistorialPersonal.objects.create(
        tecnico=empleado, tipo='BAJA', usuario=request.user,
        detalle=f'Eliminado por {request.user.get_full_name() or request.user.username}'
    )
    empleado.delete()
    return JsonResponse({'status': 'success', 'message': 'Empleado eliminado.'})


@contratista_required
def reporte_altas_bajas(request):
    empresa = get_empresa(request)
    today = date.today()
    altas = TecnicoPuesto.objects.filter(
        empresa=empresa, fecha_alta__year=today.year, fecha_alta__month=today.month
    ).order_by('-fecha_alta')
    bajas = HistorialPersonal.objects.filter(
        tecnico__empresa=empresa, tipo='BAJA',
        fecha__year=today.year, fecha__month=today.month
    ).select_related('tecnico').order_by('-fecha')
    return render(request, 'portalsub/reporte_altas_bajas.html', {
        'empresa': empresa,
        'altas': altas,
        'bajas': bajas,
        'today': timezone.now(),
    })


@contratista_required
def personal_detalle(request, pk):
    empresa = get_empresa(request)
    empleado = get_object_or_404(TecnicoPuesto, pk=pk)
    documentos = DocumentoPersonal.objects.filter(tecnico=empleado).select_related('tipo')
    tipos_doc = TipoDocumentoPersonal.objects.filter(activo=True)

    docs_status = {}
    for t in tipos_doc:
        docs_status[t.id] = {
            'label': t.nombre,
            'documento': documentos.filter(tipo=t).first(),
        }

    historial = HistorialPersonal.objects.filter(tecnico=empleado)[:20]

    return render(request, 'portalsub/personal_detalle.html', {
        'active_tab': 'personal',
        'empresa': empresa,
        'empleado': empleado,
        'docs_status': docs_status,
        'documentos': documentos,
        'historial': historial,
        'tipos_doc': tipos_doc,
    })


@contratista_required
@require_POST
def subir_documento_personal(request, pk):
    empresa = get_empresa(request)
    empleado = get_object_or_404(TecnicoPuesto, pk=pk)
    tipo_id = request.POST.get('tipo')
    archivo = request.FILES.get('archivo')

    if not tipo_id or not archivo:
        return JsonResponse({'status': 'error', 'message': 'Faltan datos requeridos.'}, status=400)

    tipo = get_object_or_404(TipoDocumentoPersonal, pk=tipo_id, activo=True)

    doc, created = DocumentoPersonal.objects.update_or_create(
        tecnico=empleado,
        tipo=tipo,
        defaults={
            'archivo': archivo,
            'subido_por': request.user,
            'es_valido': True,
        }
    )
    accion = 'Subió' if created else 'Reemplazó'
    HistorialPersonal.objects.create(
        tecnico=empleado, tipo='DOCUMENTO', usuario=request.user,
        detalle=f'{accion}: {tipo.nombre}'
    )
    return JsonResponse({
        'status': 'success',
        'message': 'Documento subido correctamente.',
        'documento': {
            'id': doc.id,
            'tipo': tipo.nombre,
            'nombre': doc.archivo.name.split('/')[-1],
            'url': doc.archivo.url,
        }
    })


@contratista_required
@require_POST
def eliminar_documento_personal(request, pk, doc_id):
    empresa = get_empresa(request)
    empleado = get_object_or_404(TecnicoPuesto, pk=pk)
    doc = get_object_or_404(DocumentoPersonal, pk=doc_id, tecnico=empleado)
    doc.archivo.delete(save=False)
    doc.delete()
    return JsonResponse({'status': 'success', 'message': 'Documento eliminado.'})


@contratista_required
def ordenes_list(request):
    empresa = get_empresa(request)
    estado = request.GET.get('estado', '')

    ordenes = OrdenCompra.objects.filter(proveedor=empresa).order_by('-fecha_creacion')
    if estado:
        ordenes = ordenes.filter(estado=estado)

    estados_oc = OrdenCompra.ESTADO_OC_CHOICES
    return render(request, 'portalsub/ordenes_list.html', {
        'active_tab': 'ordenes',
        'empresa': empresa,
        'ordenes': ordenes,
        'estados_oc': estados_oc,
        'filtro_estado': estado,
    })


@contratista_required
def orden_detalle(request, oc_id):
    empresa = get_empresa(request)
    oc = get_object_or_404(OrdenCompra, pk=oc_id, proveedor=empresa)
    documentos = DocumentoOrdenCompra.objects.filter(orden_compra=oc).order_by('tipo')
    puede_subir = oc.estado in ('CONFIRMADA', 'RECIBIDA')

    documentos_dict = {doc.tipo: doc for doc in documentos}
    tipos_doc = DocumentoOrdenCompra.TIPO_DOC_CHOICES
    documentos_items = [(key, label, documentos_dict.get(key)) for key, label in tipos_doc]

    context = {
        'active_tab': 'ordenes',
        'empresa': empresa,
        'oc': oc,
        'documentos': documentos,
        'documentos_items': documentos_items,
        'puede_subir': puede_subir,
        'tipos_doc': tipos_doc,
    }
    return render(request, 'portalsub/orden_detalle.html', context)


@contratista_required
@require_POST
def subir_documento_oc(request, oc_id):
    empresa = get_empresa(request)
    oc = get_object_or_404(OrdenCompra, pk=oc_id, proveedor=empresa)

    if oc.estado not in ('CONFIRMADA', 'RECIBIDA'):
        return JsonResponse({'status': 'error', 'message': 'No se pueden subir documentos en este estado.'}, status=400)

    tipo = request.POST.get('tipo')
    archivo = request.FILES.get('archivo')
    descripcion = request.POST.get('descripcion', '')

    if not tipo or not archivo:
        return JsonResponse({'status': 'error', 'message': 'Faltan datos requeridos.'}, status=400)

    if tipo not in dict(DocumentoOrdenCompra.TIPO_DOC_CHOICES):
        return JsonResponse({'status': 'error', 'message': 'Tipo de documento inválido.'}, status=400)

    doc = DocumentoOrdenCompra.objects.create(
        orden_compra=oc,
        tipo=tipo,
        archivo=archivo,
        descripcion=descripcion,
        subido_por=request.user,
        es_valido=True,
    )
    notificar_a_grupo(
        grupo_nombre='Administradores',
        titulo="Documento de OC Subido",
        mensaje=f"{empresa.nombre} subió un documento {doc.get_tipo_display()} a la OC {oc.numero_oc}.",
        tipo='INFO',
        modulo='PORTAL_SUB',
        enlace=f"/presupuestos/ordenes-compra/{oc.id}/detalle/",
        icono='cloud-upload-outline',
    )
    return JsonResponse({
        'status': 'success',
        'message': 'Documento subido correctamente.',
        'documento': {
            'id': doc.id,
            'tipo': doc.get_tipo_display(),
            'nombre': doc.archivo.name.split('/')[-1],
            'url': doc.archivo.url,
            'descripcion': doc.descripcion or '',
            'creado_en': doc.creado_en.strftime('%d/%m/%Y %H:%M'),
        }
    })


@contratista_required
@require_POST
def eliminar_documento_oc(request, oc_id, doc_id):
    empresa = get_empresa(request)
    oc = get_object_or_404(OrdenCompra, pk=oc_id, proveedor=empresa)
    doc = get_object_or_404(DocumentoOrdenCompra, pk=doc_id, orden_compra=oc)
    doc.archivo.delete(save=False)
    doc.delete()
    return JsonResponse({'status': 'success', 'message': 'Documento eliminado.'})
