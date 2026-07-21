import io, os
from datetime import datetime
from celery import shared_task
from django.core.files.storage import default_storage
from django.core.cache import cache
from django.utils import timezone
from django.core.files.base import ContentFile
from tablib import Dataset
import time
from .resources import RequisicionResource, ItemSolicitudPagoResource, PaqueteMaterialItemResource

def try_decode(content, encodings=['utf-8-sig', 'utf-8', 'windows-1252', 'iso-8859-1', 'utf-16', 'mac_roman']):
    for encoding in encodings:
        try:
            decoded = content.decode(encoding)
            if encoding in ['iso-8859-1', 'windows-1252', 'latin-1'] and '\x00' in decoded:
                continue
            return decoded
        except (UnicodeDecodeError, AttributeError):
            continue
    return content.decode('utf-8', errors='ignore')

@shared_task(bind=True)
def import_requisiciones_task(self, file_path, file_format, user_id=None, verification_mode=False, dry_run=False):
    """
    Tarea Celery para importar o VERIFICAR REQUISICIONES con seguimiento de progreso real.
    """
    from .models import Requisicion
    resource = RequisicionResource()
    cache_key = f"import_requisiciones_progress_{user_id}" if user_id else "import_requisiciones_progress_system"

    try:
        with default_storage.open(file_path, 'rb') as f:
            file_content = f.read()
            if file_format == 'csv':
                dataset = Dataset().load(try_decode(file_content), format='csv')
            elif file_format in ['xls', 'xlsx']:
                dataset = Dataset().load(file_content, format=file_format)
            else:
                raise ValueError(f"Formato no soportado: {file_format}")
    except Exception as e:
        error_res = {'status': 'error', 'message': f'Error al leer archivo: {str(e)}'}
        cache.set(cache_key, error_res, 3600)
        return error_res

    total_rows = len(dataset)
    progress_info = {
        'current': 0, 
        'total': total_rows, 
        'status': 'Iniciando verificacion...' if verification_mode else 'Iniciando importacion...', 
        'percent': 0,
        'new': 0,
        'updated': 0,
        'skipped': 0,
        'errors': 0,
        'found': 0,
        'not_found': 0,
        'verification_mode': verification_mode
    }
    cache.set(cache_key, progress_info, 3600)
    self.update_state(state='PROGRESS', meta=progress_info)

    if verification_mode:
        results = []
        found_count = 0
        not_found_count = 0
        
        # Optimización: Prest-fetch de identificadores existentes para evitar N+1 queries
        # Traemos solo los campos necesarios y los convertimos a string/set para búsqueda O(1)
        existing_reqs = set(Requisicion.objects.exclude(cr8ca_requisicion__isnull=True).values_list('cr8ca_requisicion', flat=True))
        # Para UUIDs, asegurar conversión a string para comparación
        existing_uuids = set(str(uuid_val) for uuid_val in Requisicion.objects.values_list('cr8ca_requisicionid', flat=True))
        
        for i, row in enumerate(dataset.dict, start=1):
            # Identificador de Dynamics (cr8ca_requisicion o ID UUID)
            req_id = str(row.get('cr8ca_requisicion') or '').strip()
            uuid_id = str(row.get('cr8ca_requisicionid') or '').strip()
            
            identifier = req_id if req_id else uuid_id
            
            if not identifier:
                status = "SIN IDENTIFICADOR"
                not_found_count += 1
            else:
                exists = False
                if req_id and req_id in existing_reqs:
                    exists = True
                elif uuid_id and uuid_id in existing_uuids:
                    exists = True
                
                if exists:
                    found_count += 1
                    status = "EXISTE"
                else:
                    not_found_count += 1
                    status = "NO EXISTE"
            
            results.append(f"Fila {i}: '{identifier}' -> {status}")
            
            if i % 10 == 0 or i == total_rows:
                progress_info.update({
                    'current': i,
                    'status': f'Verificando {i}/{total_rows}...',
                    'percent': int((i / total_rows) * 100),
                    'found': found_count,
                    'not_found': not_found_count,
                })
                cache.set(cache_key, progress_info, 3600)
                self.update_state(state='PROGRESS', meta=progress_info)

        final_res = {
            'status': 'completed',
            'status_code': 'completed',
            'total': total_rows,
            'found': found_count,
            'not_found': not_found_count,
            'results': results,
            'verification_mode': True
        }
    else:
        # Modo IMPORTACIÓN real o Dry Run
        try:
            result = resource.import_data(dataset, dry_run=dry_run, raise_errors=False)
            
            detailed_errors = []
            for error in result.base_errors:
                msg = str(error.error) if str(error.error) else repr(error.error)
                detailed_errors.append(f"Error General: {msg}")
            
            for line, errors in result.row_errors():
                for error in errors:
                    err_obj = error.error
                    
                    # Caso especial: Lista de errores o tipo de error críptico
                    # Django-import-export a veces mete el error en una lista
                    if isinstance(err_obj, list) and len(err_obj) > 0:
                        err_obj = err_obj[0]
                    
                    # Extraer el mensaje más humano posible
                    if hasattr(err_obj, 'message') and err_obj.message:
                        msg = err_obj.message
                    elif hasattr(err_obj, 'messages') and err_obj.messages:
                        msg = " | ".join([str(m) for m in err_obj.messages])
                    else:
                        msg = str(err_obj)
                    
                    # Limpiar mensajes técnicos comunes y dar contexto
                    if "decimal.InvalidOperation" in msg or "InvalidOperation" in msg or "ConversionSyntax" in msg:
                        msg = f"Valor numérico inválido. Verifica que el dato sea un número (ej: 57168.00) y no contenga letras."
                    elif "None" == msg or not msg:
                        msg = f"Error de validación ({type(err_obj).__name__})"
                    
                    # Si el mensaje es muy corto, intentar añadir el valor si está disponible en el error
                    if len(msg) < 50 and hasattr(error, 'row'):
                        pass # Podríamos añadir más contexto aquí si fuera necesario
                        
                    detailed_errors.append(f"Fila {line}: {msg}")

            final_res = {
                'status': 'completed',
                'status_code': 'completed',
                'total': total_rows,
                'new': result.totals.get('new', 0),
                'updated': result.totals.get('update', 0),
                'skipped': result.totals.get('skip', 0),
                'errors': len(detailed_errors),
                'error_list': detailed_errors,
                'verification_mode': False,
                'dry_run': dry_run,
                'file_path': file_path
            }
        except Exception as e:
            error_msg = f"Error crítico en importación: {str(e)}"
            progress_info.update({'status': 'error', 'message': error_msg})
            cache.set(cache_key, progress_info, 3600)
            return {'status': 'error', 'message': error_msg}

    # Limpiar archivo original SOLO si no es dry_run
    if not dry_run:
        try:
            if default_storage.exists(file_path):
                default_storage.delete(file_path)
        except:
            pass
        
    cache.set(cache_key, final_res, 3600)
    return final_res

@shared_task(bind=True)
def import_items_solicitud_task(self, file_path, file_format, user_id=None, solicitud_id=None, verification_mode=False, dry_run=False):
    """
    Tarea Celery para importar items de solicitud de pago.
    """
    from .models import ItemSolicitudPago, Requisicion
    resource = ItemSolicitudPagoResource()
    cache_key = f"import_items_pago_progress_{user_id}" if user_id else "import_items_pago_progress_system"

    try:
        with default_storage.open(file_path, 'rb') as f:
            file_content = f.read()
            if file_format == 'csv':
                dataset = Dataset().load(try_decode(file_content), format='csv')
            elif file_format in ['xls', 'xlsx']:
                dataset = Dataset().load(file_content, format=file_format)
            else:
                raise ValueError(f"Formato no soportado: {file_format}")
    except Exception as e:
        error_res = {'status': 'error', 'message': f'Error al leer archivo: {str(e)}'}
        cache.set(cache_key, error_res, 3600)
        return error_res

    total_rows = len(dataset)
    progress_info = {
        'current': 0, 'total': total_rows, 'percent': 0,
        'status': 'Procesando...', 'verification_mode': verification_mode,
        'file_path': file_path
    }
    cache.set(cache_key, progress_info, 3600)

    if verification_mode:
        results = []
        found_count = 0
        existing_reqs = set(Requisicion.objects.values_list('cr8ca_requisicion', flat=True))
        
        for i, row in enumerate(dataset.dict, start=1):
            req_code = str(row.get('codigo_requisicion') or row.get('requisicion_codigo') or '').strip()
            exists = req_code in existing_reqs
            if exists: found_count += 1
            results.append(f"Fila {i}: '{req_code}' -> {'ENCONTRADO' if exists else 'NO EXISTE'}")
            
            if i % 20 == 0 or i == total_rows:
                progress_info.update({'current': i, 'percent': int((i / total_rows) * 100)})
                cache.set(cache_key, progress_info, 3600)

        final_res = {
            'status': 'completed', 'total': total_rows, 'found': found_count,
            'not_found': total_rows - found_count, 'results': results, 
            'verification_mode': True, 'file_path': file_path
        }
    else:
        try:
            # Pasar solicitud_id a través de kwargs para que Resource lo use en before_import_row
            kwargs = {}
            if solicitud_id:
                kwargs['solicitud_id'] = solicitud_id
                
            result = resource.import_data(dataset, dry_run=dry_run, raise_errors=False, **kwargs)
            
            detailed_errors = []
            for line, errors in result.row_errors():
                for error in errors:
                    detailed_errors.append(f"Fila {line}: {str(error.error)}")

            final_res = {
                'status': 'completed', 
                'total': total_rows,
                'new': result.totals.get('new', 0),
                'updated': result.totals.get('update', 0),
                'skipped': result.totals.get('skip', 0),
                'errors': len(detailed_errors),
                'error_list': detailed_errors,
                'dry_run': dry_run
            }
        except Exception as e:
            final_res = {'status': 'error', 'message': str(e)}

    if not dry_run:
        try:
            if default_storage.exists(file_path): default_storage.delete(file_path)
        except: pass

    cache.set(cache_key, final_res, 3600)
    return final_res


@shared_task(bind=True)
def export_paquete_task(self, paquete_id, user_id=None):
    from .models import PaqueteMaterial, PaqueteMaterialItem
    cache_key = f"export_paquete_progress_{user_id}" if user_id else f"export_paquete_progress_{paquete_id}"

    try:
        paquete = PaqueteMaterial.objects.get(pk=paquete_id)
        qs = PaqueteMaterialItem.objects.filter(paquete=paquete).select_related('material')
    except PaqueteMaterial.DoesNotExist:
        err = {'status': 'error', 'message': 'Paquete no encontrado'}
        cache.set(cache_key, err, 3600)
        return err

    total = qs.count()
    cache.set(cache_key, {'status': 'processing', 'percent': 0, 'current': 0, 'total': total}, 3600)
    self.update_state(state='PROGRESS', meta={'percent': 0})

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Materiales"

    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill(start_color="0070F2", end_color="0070F2", fill_type="solid")
    thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    headers = ['SKU', 'Material', 'Descripcion', 'Cantidad', 'Unidad', 'Costo Aprox.', 'Orden']
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal='center')
        c.border = thin

    for ri, item in enumerate(qs.iterator(), 2):
        ws.cell(row=ri, column=1, value=item.material.sku if item.material else '').border = thin
        ws.cell(row=ri, column=2, value=item.material.nombre if item.material else item.descripcion).border = thin
        ws.cell(row=ri, column=3, value=item.descripcion).border = thin
        ws.cell(row=ri, column=4, value=float(item.cantidad)).border = thin
        ws.cell(row=ri, column=5, value=str(item.material.unidad_medida) if item.material and item.material.unidad_medida else '').border = thin
        ws.cell(row=ri, column=6, value=float(item.costo_aproximado) if item.costo_aproximado else 0).border = thin
        ws.cell(row=ri, column=7, value=item.orden).border = thin

        if ri % 20 == 0:
            pct = int((ri / max(total, 1)) * 100)
            prog = {'status': 'processing', 'percent': pct, 'current': ri, 'total': total}
            cache.set(cache_key, prog, 3600)
            self.update_state(state='PROGRESS', meta=prog)

    for col in ws.columns:
        max_len = max((min(len(str(cell.value or '')), 50) for cell in col), default=10) + 3
        ws.column_dimensions[col[0].column_letter].width = max_len

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    safe_name = "".join(c for c in paquete.nombre if c.isalnum() or c in ' _-').strip()[:50]
    file_name = f"paquetes/export_{safe_name}_{ts}.xlsx"
    path = default_storage.save(file_name, ContentFile(output.read()))

    url = default_storage.url(path)
    result = {'status': 'completed', 'file_path': path, 'url': url, 'total': total}
    cache.set(cache_key, result, 3600)
    return result


@shared_task(bind=True)
def import_paquete_items_task(self, file_path, file_format, paquete_id, user_id=None, dry_run=False):
    from .models import PaqueteMaterial
    resource = PaqueteMaterialItemResource()
    cache_key = f"import_paquete_progress_{user_id}" if user_id else f"import_paquete_progress_{paquete_id}"

    try:
        paquete = PaqueteMaterial.objects.get(pk=paquete_id)
    except PaqueteMaterial.DoesNotExist:
        err = {'status': 'error', 'message': 'Paquete no encontrado'}
        cache.set(cache_key, err, 3600)
        return err

    try:
        with default_storage.open(file_path, 'rb') as f:
            content = f.read()
            if file_format == 'csv':
                dataset = Dataset().load(try_decode(content), format='csv')
            elif file_format in ['xls', 'xlsx']:
                dataset = Dataset().load(content, format=file_format)
            else:
                raise ValueError(f"Formato no soportado: {file_format}")
    except Exception as e:
        err = {'status': 'error', 'message': f'Error al leer archivo: {str(e)}'}
        cache.set(cache_key, err, 3600)
        return err

    total_rows = len(dataset)
    prog = {'status': 'processing', 'percent': 0, 'current': 0, 'total': total_rows}
    cache.set(cache_key, prog, 3600)
    self.update_state(state='PROGRESS', meta=prog)

    try:
        result = resource.import_data(dataset, dry_run=dry_run, raise_errors=False, paquete_id=paquete_id)

        detailed_errors = []
        for line, errors in result.row_errors():
            for error in errors:
                detailed_errors.append(f"Fila {line}: {str(error.error)}")

        final = {
            'status': 'completed',
            'total': total_rows,
            'new': result.totals.get('new', 0),
            'updated': result.totals.get('update', 0),
            'skipped': result.totals.get('skip', 0),
            'deleted': result.totals.get('delete', 0),
            'errors': len(detailed_errors),
            'error_list': detailed_errors,
        }
    except Exception as e:
        final = {'status': 'error', 'message': str(e)}

    if not dry_run:
        try:
            if default_storage.exists(file_path):
                default_storage.delete(file_path)
        except:
            pass

    cache.set(cache_key, final, 3600)
    return final
