from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse
from django.db.models import Sum, Q
from django.db import models
from .models import PresupuestoAnual, PartidaPresupuestaria, GastoEjecutado, ItemPresupuesto, PresupuestoAgrupado, Cotizacion, ItemCotizacion, ItemPredefinido, FamiliaItem, ComponenteItem, Moneda
from django.contrib.auth.decorators import login_required, user_passes_test
from django.views.decorators.http import require_POST
from django.contrib import messages
from django.utils import timezone
from datetime import datetime, date

@login_required
def cronograma_presupuesto(request, pk=None):
    if pk:
        presupuesto = get_object_or_404(PresupuestoAnual, pk=pk)
    else:
        presupuesto = PresupuestoAnual.objects.filter(anio=datetime.now().year).first()
        if not presupuesto:
            presupuesto = PresupuestoAnual.objects.order_by('-anio').first()
    
    if not presupuesto:
        return render(request, 'presupuestos/cronograma.html', {'error': 'No hay presupuestos configurados.'})

    data = _get_cronograma_data([presupuesto])
    
    context = {
        'presupuesto': presupuesto,
        'partidas_data': data['presupuestos_data'][0]['partidas'] if data['presupuestos_data'] else [],
        'meses_nombres': data['meses_nombres'],
        'global_proyectado_mes': data['global_proyectado_mes'],
        'global_ejecutado_mes': data['global_ejecutado_mes'],
        'total_general_proyectado': data['total_general_proyectado'],
        'total_general_ejecutado': data['total_general_ejecutado'],
    }

    return render(request, 'presupuestos/cronograma.html', context)

@login_required
def cronograma_grupal(request, pk):
    grupo = get_object_or_404(PresupuestoAgrupado, pk=pk)
    presupuestos = grupo.presupuestos.all()
    
    if not presupuestos:
        return render(request, 'presupuestos/cronograma_grupal.html', {
            'grupo': grupo,
            'error': 'Este grupo no tiene presupuestos vinculados.'
        })

    data = _get_cronograma_data(presupuestos)
    
    context = {
        'grupo': grupo,
        'presupuestos_data': data['presupuestos_data'],
        'meses_nombres': data['meses_nombres'],
        'global_proyectado_mes': data['global_proyectado_mes'],
        'global_ejecutado_mes': data['global_ejecutado_mes'],
        'total_general_proyectado': data['total_general_proyectado'],
        'total_general_ejecutado': data['total_general_ejecutado'],
    }

    return render(request, 'presupuestos/cronograma_grupal.html', context)


def _get_cronograma_data(presupuestos_list):
    """
    Agrega datos de uno o varios presupuestos.
    Retorna datos estructurados por presupuesto.
    Incluye soporte para sub-ítems.
    """
    meses_nombres = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
    global_proyectado_mes = [0.0] * 12
    global_ejecutado_mes = [0.0] * 12
    presupuestos_data = []

    from .models import ItemSolicitudPago
    
    # Pre-cargar todos los pagos de los presupuestos involucrados con estatus PAGADO
    pagos_mapeo = {}
    pagos_qs = ItemSolicitudPago.objects.filter(
        estatus='PAGADO'
    ).filter(
        Q(requisicion__partida__presupuesto_anual__in=presupuestos_list) |
        Q(requisicion__item_presupuesto__partida__presupuesto_anual__in=presupuestos_list)
    ).select_related('requisicion', 'solicitud').distinct()
    
    for p_item in pagos_qs:
        it_id = p_item.requisicion.item_presupuesto_id
        if it_id:
            if it_id not in pagos_mapeo:
                pagos_mapeo[it_id] = [0.0] * 12
            if p_item.solicitud.fecha_solicitud:
                mes = p_item.solicitud.fecha_solicitud.month
                pagos_mapeo[it_id][mes-1] += float(p_item.monto_solicitado)

    for presupuesto in presupuestos_list:
        partidas = presupuesto.partidas.select_related('disciplina').prefetch_related(
            'items', 
            'items__detalles', 
            'items__subitems',
            'gastos'
        ).all()

        partidas_desglose = []
        p_total_proyectado_mensual = [0.0] * 12
        p_total_ejecutado_mensual = [0.0] * 12

        for p in partidas:
            p_disc_nombre = p.disciplina.nombre if p.disciplina else (p.descripcion or "Otros")
            
            # Ejecución de la partida
            ejecucion_partida = [0.0] * 12
            for gasto in p.gastos.all():
                m_idx = gasto.fecha.month - 1
                ejecucion_partida[m_idx] += float(gasto.monto)
                p_total_ejecutado_mensual[m_idx] += float(gasto.monto)
                global_ejecutado_mes[m_idx] += float(gasto.monto)

            # Ítems (solo nivel superior)
            items_tree = []
            proyeccion_partida = [0.0] * 12
            
            # Objetos de items de esta partida
            partida_items = list(p.items.all())
            top_items = [i for i in partida_items if not i.parent_id]
            
            for item in top_items:
                item_data = _get_item_recursive_data(item, partida_items, pagos_mapeo)
                items_tree.append(item_data)
                
                for i in range(12):
                    proyeccion_partida[i] += item_data['proyeccion'][i]
                    p_total_proyectado_mensual[i] += item_data['proyeccion'][i]
                    global_proyectado_mes[i] += item_data['proyeccion'][i]
                    # La ejecución global ya se suma desde gastos (GastoEjecutado)
                    # Pero aquí podríamos agregar la ejecución por ítem si fuera necesario.
                    # Por ahora el usuario pidió una línea debajo del ítem para trazabilidad.

            partidas_desglose.append({
                'partida': p,
                'disciplina': p_disc_nombre,
                'items': items_tree,
                'ejecucion_mensual': ejecucion_partida,
                'proyeccion_total_mensual': proyeccion_partida,
                'total_proyectado': sum(proyeccion_partida),
                'total_ejecutado': sum(ejecucion_partida),
                'gid': f"disc-{len(presupuestos_data) + 1}-{len(partidas_desglose) + 1}"
            })

        # Ordenar partidas por disciplina
        partidas_desglose.sort(key=lambda x: x['disciplina'])

        presupuestos_data.append({
            'presupuesto': presupuesto,
            'partidas': partidas_desglose,
            'total_mensual_proyectado': p_total_proyectado_mensual,
            'total_mensual_ejecutado': p_total_ejecutado_mensual,
            'total_anual_proyectado': sum(p_total_proyectado_mensual),
            'total_anual_ejecutado': sum(p_total_ejecutado_mensual),
            'gid': f"budget-{len(presupuestos_data) + 1}"
        })

    return {
        'presupuestos_data': presupuestos_data,
        'meses_nombres': meses_nombres,
        'global_proyectado_mes': global_proyectado_mes,
        'global_ejecutado_mes': global_ejecutado_mes,
        'total_general_proyectado': sum(global_proyectado_mes),
        'total_general_ejecutado': sum(global_ejecutado_mes),
    }

def _get_item_recursive_data(item, all_items, pagos_mapeo):
    """
    Obtiene datos de un ítem y sus sub-ítems recursivamente del cache 'all_items'.
    """
    proyeccion = [0.0] * 12
    for detalle in item.detalles.all():
        if 1 <= detalle.mes <= 12:
            proyeccion[detalle.mes - 1] += float(detalle.monto)
    
    # Obtener ejecución (pagos) para este ítem
    ejecucion = list(pagos_mapeo.get(item.id, [0.0] * 12))
    
    subitems_data = []
    # Buscar subitems en la lista precargada
    item_subitems = [i for i in all_items if i.parent_id == item.id]
    
    for subitem in item_subitems:
        sub_data = _get_item_recursive_data(subitem, all_items, pagos_mapeo)
        subitems_data.append(sub_data)
        for i in range(12):
            proyeccion[i] += sub_data['proyeccion'][i]
            ejecucion[i] += sub_data['ejecucion'][i]

    return {
        'id': item.id,
        'concepto': item.concepto,
        'proyeccion': proyeccion,
        'ejecucion': ejecucion,
        'total_anual': sum(proyeccion),
        'total_ejecutado_anual': sum(ejecucion),
        'subitems': subitems_data
    }

@login_required
def exportar_cronograma_excel(request, pk):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse

    presupuesto = get_object_or_404(PresupuestoAnual, pk=pk)
    data = _get_cronograma_data([presupuesto])
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Cronograma {presupuesto.anio}"
    
    # Estilos
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
    sub_header_font = Font(bold=True, size=11, color="000000")
    sub_header_fill = PatternFill(start_color="f1f5f9", end_color="f1f5f9", fill_type="solid")
    
    # 1. Título General
    ws['A1'] = f"PRESUPUESTO: {presupuesto.nombre} ({presupuesto.anio})"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:O1')
    
    # 2. Encabezados de Tabla
    headers = ["Disciplina / Item"] + data['meses_nombres'] + ["TOTAL", "EJECUTADO"]
    ws.append([]) # Espacio
    ws.append(headers)
    
    # Aplicar estilo al encabezado
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        
        # Ajustar ancho columnas
        if col_num == 1:
            ws.column_dimensions[get_column_letter(col_num)].width = 40
        else:
            ws.column_dimensions[get_column_letter(col_num)].width = 15

    # 3. Datos
    current_row = 4
    # Obtener las partidas del primer (y único) presupuesto
    partidas_list = data['presupuestos_data'][0]['partidas'] if data['presupuestos_data'] else []
    
    for pd in partidas_list:
        # Fila Partida
        ws.cell(row=current_row, column=1, value=pd['disciplina']).font = sub_header_font
        ws.cell(row=current_row, column=1).fill = sub_header_fill
        
        for m_idx, val in enumerate(pd['proyeccion_total_mensual']):
            c = ws.cell(row=current_row, column=m_idx + 2, value=val)
            c.number_format = '#,##0'
            c.font = Font(bold=True)
            c.fill = sub_header_fill
            
        # Total Partida
        c_total = ws.cell(row=current_row, column=14, value=pd['total_proyectado'])
        c_total.number_format = '#,##0'
        c_total.font = Font(bold=True)
        c_total.fill = sub_header_fill
        
        current_row += 1
        
        # Filas Items
        for item in pd['items']:
            current_row += 1
            
            # Fila Pagado del Item (Trazabilidad)
            ws.cell(row=current_row, column=1, value=f"      (Pagado)").font = Font(italic=True, size=9, color="3b82f6")
            
            for m_idx, val in enumerate(item['ejecucion']):
                c = ws.cell(row=current_row, column=m_idx + 2, value=val if val > 0 else "")
                c.number_format = '#,##0'
                c.font = Font(italic=True, size=9, color="3b82f6")
                
            c_annual_exec = ws.cell(row=current_row, column=14, value=item['total_ejecutado_anual'])
            c_annual_exec.number_format = '#,##0'
            c_annual_exec.font = Font(italic=True, size=9, color="3b82f6", bold=True)
            
            current_row += 1

    # 4. Totales Finales
    current_row += 1
    ws.cell(row=current_row, column=1, value="TOTAL MENSUAL").font = header_font
    ws.cell(row=current_row, column=1).fill = header_fill
    
    for m_idx, val in enumerate(data['global_proyectado_mes']):
        c = ws.cell(row=current_row, column=m_idx + 2, value=val)
        c.font = header_font
        c.fill = header_fill
        c.number_format = '#,##0'
        
    c_grand = ws.cell(row=current_row, column=14, value=data['total_general_proyectado'])
    c_grand.font = header_font
    c_grand.fill = header_fill
    c_grand.number_format = '#,##0'

    # Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Presupuesto_{presupuesto.anio}.xlsx'
    
    wb.save(response)
    return response

@login_required
def exportar_cronograma_pdf(request, pk):
    from django.template.loader import get_template
    from xhtml2pdf import pisa
    from django.http import HttpResponse

    presupuesto = get_object_or_404(PresupuestoAnual, pk=pk)
    data = _get_cronograma_data([presupuesto])
    
    context = {
        'presupuesto': presupuesto,
        'partidas_data': data['presupuestos_data'][0]['partidas'] if data['presupuestos_data'] else [],
        'meses_nombres': data['meses_nombres'],
        'global_proyectado_mes': data['global_proyectado_mes'],
        'global_ejecutado_mes': data['global_ejecutado_mes'],
        'total_general_proyectado': data['total_general_proyectado'],
        'total_general_ejecutado': data['total_general_ejecutado'],
    }
    
    template_path = 'presupuestos/cronograma_pdf.html'
    template = get_template(template_path)
    html = template.render(context)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Presupuesto_{presupuesto.anio}.pdf"'
    
    pisa_status = pisa.CreatePDF(
       html, dest=response
    )
    
    if pisa_status.err:
       return HttpResponse('Error creating PDF', status=500)
       
    return response

@login_required
def exportar_cronograma_grupal_excel(request, pk):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse

    grupo = get_object_or_404(PresupuestoAgrupado, pk=pk)
    presupuestos = grupo.presupuestos.all()

    if not presupuestos:
        return HttpResponse('El grupo no tiene presupuestos.', status=400)

    data = _get_cronograma_data(presupuestos)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Análisis Grupal"

    # Estilos
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
    budget_font = Font(bold=True, size=11, color="FFFFFF")
    budget_fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
    disciplina_font = Font(bold=True, size=11)
    disciplina_fill = PatternFill(start_color="f1f5f9", end_color="f1f5f9", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 1. Título General
    ws['A1'] = f"ANÁLISIS GRUPAL: {grupo.nombre}"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:O1')

    # 2. Totales Generales
    ws['A3'] = "TOTAL PROYECTADO:"
    ws['A3'].font = Font(bold=True)
    ws['B3'] = data['total_general_proyectado']
    ws['B3'].number_format = '#,##0.00'

    ws['A4'] = "TOTAL EJECUTADO:"
    ws['A4'].font = Font(bold=True)
    ws['B4'] = data['total_general_ejecutado']
    ws['B4'].number_format = '#,##0.00'

    if data['total_general_proyectado'] > 0:
        eficiencia = (data['total_general_ejecutado'] / data['total_general_proyectado']) * 100
        ws['A5'] = "EFICIENCIA:"
        ws['A5'].font = Font(bold=True)
        ws['B5'] = f"{eficiencia:.1f}%"
        ws['B5'].number_format = '0.0%'

    # 3. Encabezados de Tabla
    headers = ["Jerarquía"] + data['meses_nombres'] + ["TOTAL", "EJECUTADO"]
    ws.append([])  # Espacio
    ws.append(headers)

    # Aplicar estilo al encabezado
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

        # Ajustar ancho columnas
        if col_num == 1:
            ws.column_dimensions[get_column_letter(col_num)].width = 40
        else:
            ws.column_dimensions[get_column_letter(col_num)].width = 14

    # 4. Datos
    current_row = 8

    # Totales mensuales agrupados (fila principal)
    ws.cell(row=current_row, column=1, value="TOTAL MENSUAL AGRUPADO").font = budget_font
    ws.cell(row=current_row, column=1).fill = budget_fill
    ws.cell(row=current_row, column=1).border = thin_border

    for m_idx, val in enumerate(data['global_proyectado_mes']):
        c = ws.cell(row=current_row, column=m_idx + 2, value=val)
        c.number_format = '#,##0'
        c.font = budget_font
        c.fill = budget_fill
        c.alignment = Alignment(horizontal='right')
        c.border = thin_border

    c_total = ws.cell(row=current_row, column=14, value=data['total_general_proyectado'])
    c_total.number_format = '#,##0'
    c_total.font = budget_font
    c_total.fill = budget_fill
    c_total.border = thin_border

    c_ejec = ws.cell(row=current_row, column=15, value=data['total_general_ejecutado'])
    c_ejec.number_format = '#,##0'
    c_ejec.font = budget_font
    c_ejec.fill = budget_fill
    c_ejec.border = thin_border

    current_row += 1

    # Por cada presupuesto
    for bd in data['presupuestos_data']:
        # Fila Presupuesto
        ws.cell(row=current_row, column=1, value=bd['presupuesto'].nombre).font = budget_font
        ws.cell(row=current_row, column=1).fill = budget_fill
        ws.cell(row=current_row, column=1).border = thin_border

        for m_idx, val in enumerate(bd['total_mensual_proyectado']):
            c = ws.cell(row=current_row, column=m_idx + 2, value=val if val > 0 else 0)
            c.number_format = '#,##0'
            c.font = budget_font
            c.fill = budget_fill
            c.alignment = Alignment(horizontal='right')
            c.border = thin_border

        c_total = ws.cell(row=current_row, column=14, value=bd['total_anual_proyectado'])
        c_total.number_format = '#,##0'
        c_total.font = budget_font
        c_total.fill = budget_fill
        c_total.border = thin_border

        c_ejec = ws.cell(row=current_row, column=15, value=bd['total_anual_ejecutado'])
        c_ejec.number_format = '#,##0'
        c_ejec.font = budget_font
        c_ejec.fill = budget_fill
        c_ejec.border = thin_border

        current_row += 1

        # Por cada partida (disciplina)
        for pd in bd['partidas']:
            # Fila Disciplina
            ws.cell(row=current_row, column=1, value=pd['disciplina']).font = disciplina_font
            ws.cell(row=current_row, column=1).fill = disciplina_fill
            ws.cell(row=current_row, column=1).border = thin_border

            for m_idx, val in enumerate(pd['proyeccion_total_mensual']):
                c = ws.cell(row=current_row, column=m_idx + 2, value=val if val > 0 else 0)
                c.number_format = '#,##0'
                c.font = disciplina_font
                c.fill = disciplina_fill
                c.alignment = Alignment(horizontal='right')
                c.border = thin_border

            c_total = ws.cell(row=current_row, column=14, value=pd['total_proyectado'])
            c_total.number_format = '#,##0'
            c_total.font = disciplina_font
            c_total.fill = disciplina_fill
            c_total.border = thin_border

            c_ejec = ws.cell(row=current_row, column=15, value=pd['total_ejecutado'])
            c_ejec.number_format = '#,##0'
            c_ejec.font = disciplina_font
            c_ejec.fill = disciplina_fill
            c_ejec.border = thin_border

            current_row += 1

            # Items
            for item in pd['items']:
                ws.cell(row=current_row, column=1, value=f"   {item['concepto']}")
                ws.cell(row=current_row, column=1).border = thin_border

                for m_idx, val in enumerate(item['proyeccion']):
                    c = ws.cell(row=current_row, column=m_idx + 2, value=val if val > 0 else 0)
                    c.number_format = '#,##0'
                    c.alignment = Alignment(horizontal='right')
                    c.border = thin_border

                c_annual = ws.cell(row=current_row, column=14, value=item['total_anual'])
                c_annual.number_format = '#,##0'
                c_annual.border = thin_border

                c_ejec = ws.cell(row=current_row, column=15, value=item.get('total_ejecutado_anual', 0))
                c_ejec.number_format = '#,##0'
                c_ejec.border = thin_border

                current_row += 1
                
                # Fila Pagado para este Item
                ws.cell(row=current_row, column=1, value=f"      Pagado").font = Font(italic=True, size=9, color="3b82f6")
                ws.cell(row=current_row, column=1).border = thin_border
                
                for m_idx, pval in enumerate(item['ejecucion']):
                    c = ws.cell(row=current_row, column=m_idx + 2, value=pval if pval > 0 else 0)
                    c.number_format = '#,##0'
                    c.font = Font(italic=True, size=9, color="3b82f6")
                    c.border = thin_border
                    
                c_total_exec = ws.cell(row=current_row, column=14, value=item['total_ejecutado_anual'])
                c_total_exec.number_format = '#,##0'
                c_total_exec.font = Font(italic=True, size=9, color="3b82f6", bold=True)
                c_total_exec.border = thin_border
                
                ws.cell(row=current_row, column=15, value=item['total_ejecutado_anual']).border = thin_border
                
                current_row += 1

    # Ajustar anchos de columnas adicionales
    ws.column_dimensions['N'].width = 16
    ws.column_dimensions['O'].width = 16

    # Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"Analisis_Grupal_{grupo.nombre.replace(' ', '_')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response

@login_required
def exportar_cronograma_grupal_excel_pivot(request, pk):
    """
    Exporta los datos en formato plano para tabla pivote.
    Estructura: Presupuesto | Disciplina | Item | Concepto | Mes | Proyectado | Ejecutado
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse

    grupo = get_object_or_404(PresupuestoAgrupado, pk=pk)
    presupuestos = grupo.presupuestos.all()

    if not presupuestos:
        return HttpResponse('El grupo no tiene presupuestos.', status=400)

    data = _get_cronograma_data(presupuestos)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Datos Pivote"

    # Encabezados
    headers = ["Presupuesto", "Disciplina", "Item", "Concepto", "Mes", "Proyectado", "Ejecutado"]
    ws.append(headers)

    # Estilo encabezado
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Ancho columnas
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15

    meses_nombres = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                     'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']

    # Datos - por cada presupuesto
    for bd in data['presupuestos_data']:
        presupuesto_nombre = bd['presupuesto'].nombre
        presupuesto_ejecutado = bd['total_anual_ejecutado']

        # Por cada partida (disciplina)
        for pd in bd['partidas']:
            disciplina = pd['disciplina']
            partida_proyeccion = pd['proyeccion_total_mensual']
            partida_ejecutado = pd['total_ejecutado']

            # Fila de disciplina (sin item)
            for m_idx, mes_nombre in enumerate(meses_nombres):
                ws.append([
                    presupuesto_nombre,
                    disciplina,
                    "",
                    f"TOTAL {disciplina.upper()}",
                    mes_nombre,
                    partida_proyeccion[m_idx] if partida_proyeccion[m_idx] > 0 else 0,
                    0  # La ejecución se muestra a nivel de item, no de partida
                ])

            # Por cada item
            for item in pd['items']:
                item_concepto = item['concepto']
                item_proyeccion = item['proyeccion']
                item_ejecutado = item.get('total_ejecutado', 0)

                for m_idx, mes_nombre in enumerate(meses_nombres):
                    ws.append([
                        presupuesto_nombre,
                        disciplina,
                        item.get('codigo', ''),
                        item_concepto,
                        mes_nombre,
                        item_proyeccion[m_idx] if item_proyeccion[m_idx] > 0 else 0,
                        0  # Ejecutado mensual no está en item_data, sería 0
                    ])

                # Fila de total anual del item
                ws.append([
                    presupuesto_nombre,
                    disciplina,
                    item.get('codigo', ''),
                    f"TOTAL {item_concepto}",
                    "ANUAL",
                    item['total_anual'],
                    item_ejecutado
                ])

    # Totales mensuales agrupados
    ws.append([])
    ws.append(["TOTALES AGRUPADOS", "", "", "", "", "", ""])
    total_row = ws.max_row + 1

    for m_idx, mes_nombre in enumerate(meses_nombres):
        ws.append([
            "",
            "",
            "",
            f"TOTAL {mes_nombre.upper()}",
            "",
            data['global_proyectado_mes'][m_idx],
            data['global_ejecutado_mes'][m_idx]
        ])

    # Total anual
    ws.append([
        "",
        "",
        "",
        "TOTAL GENERAL",
        "",
        data['total_general_proyectado'],
        data['total_general_ejecutado']
    ])

    # Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"Analisis_Grupal_Pivot_{grupo.nombre.replace(' ', '_')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response

@login_required
def exportar_cronograma_grupal_excel_pivot(request, pk):
    """
    Exporta cronograma grupal en formato plano para tabla pivote.
    Estructura: Presupuesto | Disciplina | Item | Concepto | Mes | Valor
    Una fila por cada mes con valor.
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse

    grupo = get_object_or_404(PresupuestoAgrupado, pk=pk)
    presupuestos = grupo.presupuestos.all()

    if not presupuestos:
        return HttpResponse('El grupo no tiene presupuestos.', status=400)

    data = _get_cronograma_data(presupuestos)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Datos Pivote"

    # Estilos
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 1. Título
    ws['A1'] = f"ANÁLISIS GRUPAL: {grupo.nombre} - Datos para Pivote"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:E1')

    # 2. Encabezados - formato plano (una fila por mes)
    headers = ["Presupuesto", "Disciplina", "Item", "Concepto", "Mes", "Valor"]
    ws.append([])  # Espacio
    ws.append(headers)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15

    # 3. Datos - formato plano (una fila por cada mes con valor)
    current_row = 4

    for bd in data['presupuestos_data']:
        presupuesto_nombre = bd['presupuesto'].nombre

        for pd in bd['partidas']:
            disciplina = pd['disciplina']

            # Items
            for item in pd['items']:
                item_id = item.get('id', '')
                concepto = item['concepto']

                # Una fila por cada mes
                for m_idx, val in enumerate(item['proyeccion']):
                    # Columna A: Presupuesto
                    ws.cell(row=current_row, column=1, value=presupuesto_nombre).border = thin_border

                    # Columna B: Disciplina
                    ws.cell(row=current_row, column=2, value=disciplina).border = thin_border

                    # Columna C: Item ID
                    ws.cell(row=current_row, column=3, value=str(item_id)).border = thin_border

                    # Columna D: Concepto
                    ws.cell(row=current_row, column=4, value=concepto).border = thin_border

                    # Columna E: Mes
                    ws.cell(row=current_row, column=5, value=data['meses_nombres'][m_idx]).border = thin_border

                    # Columna F: Valor (Proyectado)
                    c = ws.cell(row=current_row, column=6, value=val if val > 0 else 0)
                    c.number_format = '#,##0'
                    c.alignment = Alignment(horizontal='right')
                    c.border = thin_border
                    
                    # Columna G: Pagado (Ejecutado de Solicitudes)
                    pval = item['ejecucion'][m_idx]
                    c2 = ws.cell(row=current_row, column=7, value=pval if pval > 0 else 0)
                    c2.number_format = '#,##0'
                    c2.alignment = Alignment(horizontal='right')
                    c2.border = thin_border

                    current_row += 1

    # Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"Analisis_Grupal_{grupo.nombre.replace(' ', '_')}_Pivote.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response


@login_required
def api_update_monto_mensual(request):
    if request.method == "POST":
        import json
        from .models import DetallePeriodico, ItemPresupuesto
        from django.http import JsonResponse
        
        try:
            data = json.loads(request.body)
            item_id = data.get('item_id')
            mes = int(data.get('mes'))
            monto = float(data.get('monto')) # Permitir float
            
            # Buscar Item Padre
            item = ItemPresupuesto.objects.get(pk=item_id)
            
            # Buscar o Crear detalle 
            # (Aunque usualmente ya existe, si el usuario pone monto > 0 en un mes donde no habia, lo creamos)
            detalle, created = DetallePeriodico.objects.get_or_create(
                item=item,
                mes=mes,
                defaults={'monto': monto}
            )
            
            if not created:
                detalle.monto = monto
                detalle.save()
            
            # Si el monto es 0, podríamos optar por borrar el detalle para limpiar la BD, 
            # pero por ahora mantenerlo es más seguro para la integridad histórica.
            
            return JsonResponse({'status': 'ok', 'new_total': float(item.total_anual)})
            
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
            
    return JsonResponse({'status': 'error', 'message': 'Invalid method'}, status=405)

@login_required
def api_create_item(request):
    if request.method == "POST":
        import json
        from .models import ItemPresupuesto, PartidaPresupuestaria
        from django.http import JsonResponse
        
        try:
            data = json.loads(request.body)
            partida_id = data.get('partida_id')
            concepto = data.get('concepto')
            
            if not concepto:
                return JsonResponse({'status': 'error', 'message': 'El concepto es obligatorio'}, status=400)
                
            partida = PartidaPresupuestaria.objects.get(pk=partida_id)
            
            # Crear Item con defaults manuales
            item = ItemPresupuesto.objects.create(
                partida=partida,
                concepto=concepto,
                frecuencia='MANUAL',
                es_recurrente=False
            )
            
            return JsonResponse({'status': 'ok', 'message': 'Item creado'})
            
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
            
    return JsonResponse({'status': 'error', 'message': 'Invalid method'}, status=405)

@login_required
def api_create_partida(request):
    if request.method == "POST":
        import json
        from .models import PartidaPresupuestaria, PresupuestoAnual
        from django.http import JsonResponse
        
        try:
            data = json.loads(request.body)
            presupuesto_id = data.get('presupuesto_id')
            nombre = data.get('nombre')
            
            if not nombre:
                return JsonResponse({'status': 'error', 'message': 'El nombre es obligatorio'}, status=400)
                
            presupuesto = PresupuestoAnual.objects.get(pk=presupuesto_id)
            
            # Crear Partida genérica (Sin disciplina vinculada, usando descripcion)
            PartidaPresupuestaria.objects.create(
                presupuesto_anual=presupuesto,
                disciplina=None,
                descripcion=nombre,
                monto_proyectado=0 
            )
            
            return JsonResponse({'status': 'ok', 'message': 'Partida creada'})
            
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
            
    return JsonResponse({'status': 'error', 'message': 'Invalid method'}, status=405)

@login_required
def api_update_item(request):
    if request.method == "POST":
        import json
        from .models import ItemPresupuesto
        from django.http import JsonResponse
        from django.shortcuts import get_object_or_404
        
        try:
            data = json.loads(request.body)
            item_id = data.get('item_id')
            concepto = data.get('concepto')
            
            if not item_id or not concepto:
                 return JsonResponse({'status': 'error', 'message': 'Faltan datos'}, status=400)

            item = get_object_or_404(ItemPresupuesto, pk=item_id)
            item.concepto = concepto
            item.save()
            
            return JsonResponse({'status': 'ok', 'message': 'Item actualizado'})
            
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
            
    return JsonResponse({'status': 'error', 'message': 'Invalid method'}, status=405)

@login_required
def api_delete_item(request):
    if request.method == "POST":
        import json
        from .models import ItemPresupuesto
        from django.http import JsonResponse
        
        try:
            data = json.loads(request.body)
            item_id = data.get('item_id')
            
            item = get_object_or_404(ItemPresupuesto, pk=item_id)
            item.delete()
            
            return JsonResponse({'status': 'ok', 'message': 'Item eliminado'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Invalid method'}, status=405)

@login_required
def api_delete_partida(request):
    if request.method == "POST":
        import json
        from .models import PartidaPresupuestaria
        from django.http import JsonResponse
        
        try:
            data = json.loads(request.body)
            partida_id = data.get('partida_id')
            
            partida = get_object_or_404(PartidaPresupuestaria, pk=partida_id)
            
            # Opcional: Validar si tiene items o gastos antes de borrar
            # Por ahora permitimos borrar y Django manejará cascadas si las hay
            partida.delete()
            
            return JsonResponse({'status': 'ok', 'message': 'Partida eliminada'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Invalid method'}, status=405)


@login_required
def exportar_cronograma_grupal_pdf(request, pk):
    from django.template.loader import get_template
    from xhtml2pdf import pisa
    from django.http import HttpResponse

    grupo = get_object_or_404(PresupuestoAgrupado, pk=pk)
    presupuestos = grupo.presupuestos.all()
    
    if not presupuestos:
        return HttpResponse('El grupo no tiene presupuestos.', status=400)

    data = _get_cronograma_data(presupuestos)
    
    context = {
        'grupo': grupo,
        'presupuestos_data': data['presupuestos_data'],
        'meses_nombres': data['meses_nombres'],
        'global_proyectado_mes': data['global_proyectado_mes'],
        'global_ejecutado_mes': data['global_ejecutado_mes'],
        'total_general_proyectado': data['total_general_proyectado'],
        'total_general_ejecutado': data['total_general_ejecutado'],
    }
    
    template_path = 'presupuestos/cronograma_grupal_pdf.html'
    template = get_template(template_path)
    html = template.render(context)
    
    response = HttpResponse(content_type='application/pdf')
    filename = f"Analisis_Grupal_{grupo.nombre.replace(' ', '_')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    pisa_status = pisa.CreatePDF(
       html, dest=response
    )
    
    if pisa_status.err:
       return HttpResponse('Error creating PDF', status=500)
       
    return response


# ──────────────────────────────────────────────────
# REPEX Cronograma / Visualizador Interactivo
# ──────────────────────────────────────────────────

@login_required
def cronograma_repex(request, pk):
    from .models import REPEX
    repex = get_object_or_404(REPEX, pk=pk)
    # Filtros de fecha
    desde = request.GET.get('desde')
    hasta = request.GET.get('hasta')
    
    data = _get_repex_cronograma_data(repex, fecha_desde=desde, fecha_hasta=hasta)

    # Vista detalle de activos
    items_detalle = []
    items_qs = repex.items.select_related(
        'activo', 'activo__modelo', 'activo__modelo__marca',
        'activo__modelo__categoria', 'activo__ubicacion', 'activo__familia',
        'modelo', 'modelo__marca', 'modelo__categoria'
    )
    
    if desde:
        items_qs = items_qs.filter(fecha_proyectada__gte=desde)
    if hasta:
        items_qs = items_qs.filter(fecha_proyectada__lte=hasta)
    
    items_qs = items_qs.all()

    for item in items_qs:
        activo = item.activo
        if activo:
            # Item vinculado a un activo
            edificio_obj = activo.ubicacion
            nombre_edificio = '-'
            if edificio_obj:
                visited = {edificio_obj.id}
                found_edificio = edificio_obj if edificio_obj.tipo == 'EDIFICIO' else None
                temp_curr = edificio_obj.padre
                while temp_curr:
                    if temp_curr.id in visited: break
                    visited.add(temp_curr.id)
                    if temp_curr.tipo == 'EDIFICIO':
                        found_edificio = temp_curr
                    temp_curr = temp_curr.padre
                nombre_edificio = found_edificio.nombre if found_edificio else edificio_obj.get_root().nombre
            
            ruta_ubicacion = nombre_edificio
            ruta_categoria = '-'
            categoria_principal = '-'
            if activo.modelo and activo.modelo.categoria:
                cat = activo.modelo.categoria
                path_objs = [cat]
                curr = cat.padre
                visited = {cat.id}
                while curr:
                    if curr.id in visited: break
                    visited.add(curr.id)
                    path_objs.append(curr)
                    curr = curr.padre
                ruta_categoria = ' → '.join(reversed([c.nombre for c in path_objs]))
                categoria_principal = path_objs[-1].nombre

            items_detalle.append({
                'id': item.id,
                'activo_nombre': activo.nombre,
                'codigo': activo.codigo_interno or '-',
                'marca': activo.modelo.marca.nombre if activo.modelo and activo.modelo.marca else '-',
                'modelo': activo.modelo.nombre if activo.modelo else '-',
                'ruta_ubicacion': ruta_ubicacion,
                'ruta_categoria': ruta_categoria,
                'categoria_principal': categoria_principal,
                'familia': activo.familia.nombre if activo.familia else '-',
                'costo_reposicion': float(item.costo_reposicion or 0),
                'prioridad': item.prioridad,
                'es_manual': False,
                'cantidad': float(item.cantidad or 1),
                'unidades': item.unidades or '',
                'precio_unitario': float(item.precio_unitario or 0),
            })
        else:
            # Item manual sin activo
            modelo_str = item.modelo.nombre if item.modelo else '-'
            marca_str = item.modelo.marca.nombre if item.modelo and item.modelo.marca else '-'
            
            ruta_categoria = '-'
            categoria_principal = '-'
            if item.modelo and item.modelo.categoria:
                cat = item.modelo.categoria
                path_objs = [cat]
                curr = cat.padre
                visited = {cat.id}
                while curr:
                    if curr.id in visited: break
                    visited.add(curr.id)
                    path_objs.append(curr)
                    curr = curr.padre
                ruta_categoria = ' → '.join(reversed([c.nombre for c in path_objs]))
                categoria_principal = path_objs[-1].nombre
            elif hasattr(item, 'categoria_manual') and item.categoria_manual:
                ruta_categoria = item.categoria_manual
                categoria_principal = item.categoria_manual

            items_detalle.append({
                'id': item.id,
                'activo_nombre': item.nombre_item or 'Ítem manual',
                'codigo': '-',
                'marca': marca_str,
                'modelo': modelo_str,
                'ruta_ubicacion': item.ubicacion_manual or '-',
                'ruta_categoria': ruta_categoria,
                'categoria_principal': categoria_principal,
                'familia': ruta_categoria,
                'costo_reposicion': float(item.costo_reposicion or 0),
                'prioridad': item.prioridad,
                'es_manual': True,
                'cantidad': float(item.cantidad or 1),
                'unidades': item.unidades or '',
                'precio_unitario': float(item.precio_unitario or 0),
            })

    # Agrupar por categoría y luego por edificio para vista jerárquica de 3 niveles
    from collections import OrderedDict
    hierarchical_dict = OrderedDict()
    resumen_dict = OrderedDict()

    for item in items_detalle:
        cat = item['ruta_categoria']
        building = item['ruta_ubicacion']
        item_id = item['id']

        # 1. Estructura para Detalle/Presupuesto (3 niveles)
        if cat not in hierarchical_dict:
            hierarchical_dict[cat] = {
                'nombre': cat, 
                'edificios': OrderedDict(), 
                'total': 0.0, 
                'total_cantidad': 0.0,
                'count': 0
            }
        
        if building not in hierarchical_dict[cat]['edificios']:
            hierarchical_dict[cat]['edificios'][building] = {
                'nombre': building, 
                'items': [], 
                'total_edificio': 0.0,
                'total_cantidad': 0.0
            }
        
        hierarchical_dict[cat]['edificios'][building]['items'].append(item)
        hierarchical_dict[cat]['edificios'][building]['total_edificio'] += item['costo_reposicion']
        hierarchical_dict[cat]['edificios'][building]['total_cantidad'] += item['cantidad']
        hierarchical_dict[cat]['total'] += item['costo_reposicion']
        hierarchical_dict[cat]['total_cantidad'] += item['cantidad']
        hierarchical_dict[cat]['count'] += 1

        # 2. Resumen por Modelo (existente)
        if cat not in resumen_dict:
            resumen_dict[cat] = {'nombre': cat, 'modelos': OrderedDict(), 'total': 0.0, 'total_cantidad': 0.0}
        
        modelo_key = f"{item['marca']} | {item['modelo']}"
        if modelo_key not in resumen_dict[cat]['modelos']:
            resumen_dict[cat]['modelos'][modelo_key] = {
                'marca': item['marca'],
                'modelo': item['modelo'],
                'cantidad': 0,
                'precio_unitario': item['precio_unitario'],
                'total': 0.0
            }
        
        m_data = resumen_dict[cat]['modelos'][modelo_key]
        m_data['cantidad'] += item['cantidad']
        m_data['total'] += item['costo_reposicion']
        resumen_dict[cat]['total'] += item['costo_reposicion']
        resumen_dict[cat]['total_cantidad'] += item['cantidad']

    # Convertir a listas para el template
    presupuesto_jerarquico = []
    for cat_name, c_data in hierarchical_dict.items():
        cat_item = {
            'nombre': cat_name,
            'total': c_data['total'],
            'total_cantidad': c_data['total_cantidad'],
            'promedio_pu': c_data['total'] / c_data['total_cantidad'] if c_data['total_cantidad'] > 0 else 0,
            'count': c_data['count'],
            'edificios': []
        }
        for b_name, b_data in c_data['edificios'].items():
            b_data['promedio_pu'] = b_data['total_edificio'] / b_data['total_cantidad'] if b_data['total_cantidad'] > 0 else 0
            cat_item['edificios'].append(b_data)
        presupuesto_jerarquico.append(cat_item)

    resumen_modelos = []
    for cat_name, r_data in resumen_dict.items():
        total_cat = r_data['total']
        cant_cat = r_data['total_cantidad']
        promedio_pu = total_cat / cant_cat if cant_cat > 0 else 0
        
        resumen_modelos.append({
            'nombre': cat_name,
            'modelos': list(r_data['modelos'].values()),
            'total': total_cat,
            'total_cantidad': cant_cat,
            'promedio_pu': promedio_pu
        })

    # 3. Resumen por Categoría Principal / Ubicación (NUEVO)
    resumen_cat_ub_dict = OrderedDict()
    for item in items_detalle:
        cat = item['categoria_principal']
        ub = item['ruta_ubicacion']
        ckey = (cat, ub)
        if ckey not in resumen_cat_ub_dict:
            resumen_cat_ub_dict[ckey] = {
                'categoria': cat,
                'ubicacion': ub,
                'total_cantidad': 0.0,
                'total': 0.0,
            }
        r_catub = resumen_cat_ub_dict[ckey]
        r_catub['total_cantidad'] += item['cantidad']
        r_catub['total'] += item['costo_reposicion']
    
    # Organizar por jerarquía de 3 NIVELES para asignar Códigos (A.1.1) y habilitar colapsable
    sorted_cat_ubs = sorted(resumen_cat_ub_dict.values(), key=lambda x: (x['categoria'], x['ubicacion']))
    
    hierarchical_resumen = OrderedDict()
    import string
    def get_excel_letter(n):
        result = ""
        while n >= 0:
            result = string.ascii_uppercase[n % 26] + result
            n = n // 26 - 1
        return result

    main_cat_idx = -1
    for r_val in sorted_cat_ubs:
        main_cat_name = r_val['categoria']
        # Usamos ruta_categoria para identificar el nivel intermedio (Hijo)
        # Buscamos el ítem original para obtener la ruta completa
        # (Alternativamente, podríamos haber guardado la ruta completa en r_val)
        # Busquemos en items_detalle la ruta asociada a este ckey
        matching_item = next((i for i in items_detalle if i['categoria_principal'] == main_cat_name and i['ruta_ubicacion'] == r_val['ubicacion']), None)
        sub_cat_name = matching_item['ruta_categoria'] if matching_item else main_cat_name

        if main_cat_name not in hierarchical_resumen:
            main_cat_idx += 1
            hierarchical_resumen[main_cat_name] = {
                'nombre': main_cat_name,
                'codigo': get_excel_letter(main_cat_idx),
                'total_cantidad': 0.0,
                'total': 0.0,
                'hijos': OrderedDict()
            }
        
        main_cat_obj = hierarchical_resumen[main_cat_name]
        
        if sub_cat_name not in main_cat_obj['hijos']:
            sub_cat_idx = len(main_cat_obj['hijos']) + 1
            main_cat_obj['hijos'][sub_cat_name] = {
                'nombre': sub_cat_name,
                'codigo': f"{main_cat_obj['codigo']}.{sub_cat_idx}",
                'total_cantidad': 0.0,
                'total': 0.0,
                'ubicaciones': []
            }
        
        sub_cat_obj = main_cat_obj['hijos'][sub_cat_name]
        loc_idx = len(sub_cat_obj['ubicaciones']) + 1
        
        total_r = r_val['total']
        cant_r = r_val['total_cantidad']
        r_val['promedio_pu'] = total_r / cant_r if cant_r > 0 else 0
        r_val['codigo'] = f"{sub_cat_obj['codigo']}.{loc_idx}"
        
        # Sube los totales al sub-hijo
        sub_cat_obj['total_cantidad'] += cant_r
        sub_cat_obj['total'] += total_r
        sub_cat_obj['ubicaciones'].append(r_val)
        
        # Sube los totales al padre principal
        main_cat_obj['total_cantidad'] += cant_r
        main_cat_obj['total'] += total_r

    # Convertir a lista y calcular promedios finales
    resumen_cat_ub = []
    for main_cat in hierarchical_resumen.values():
        main_cat['promedio_pu'] = main_cat['total'] / main_cat['total_cantidad'] if main_cat['total_cantidad'] > 0 else 0
        main_cat['hijos_list'] = []
        for sub_cat in main_cat['hijos'].values():
            sub_cat['promedio_pu'] = sub_cat['total'] / sub_cat['total_cantidad'] if sub_cat['total_cantidad'] > 0 else 0
            main_cat['hijos_list'].append(sub_cat)
        resumen_cat_ub.append(main_cat)

    context = {
        'repex': repex,
        'desde': desde,
        'hasta': hasta,
        'familias_data': data['familias_data'],
        'anios_nombres': data['anios_nombres'],
        'total_anual': data['total_mensual'], # Renombrando key para template y vista
        'total_general': data['total_general'],
        'total_items': data['total_items'],
        'items_detalle': items_detalle,
        'presupuesto_jerarquico': presupuesto_jerarquico, # Nueva estructura 3 niveles
        'resumen_modelos': resumen_modelos,
        'resumen_cat_ub': resumen_cat_ub, # NUEVO
    }
    return render(request, 'presupuestos/cronograma_repex.html', context)


@login_required
def exportar_repex_excel(request, pk):
    """Exporta el plan REPEX a Excel. Acepta ?vista=detalle|apu|cronograma para cambiar formato."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    from .models import REPEX
    from collections import OrderedDict

    repex = get_object_or_404(REPEX, pk=pk)
    vista = request.GET.get('vista', 'detalle')

    # Styles
    header_fill = PatternFill(start_color='3D5A80', end_color='3D5A80', fill_type='solid')
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    cat_fill = PatternFill(start_color='C8D8E8', end_color='C8D8E8', fill_type='solid')
    cat_font = Font(name='Calibri', bold=True, color='1E3A5F', size=11)
    item_font = Font(name='Calibri', size=10, color='333333')
    total_fill = PatternFill(start_color='2C4A6E', end_color='2C4A6E', fill_type='solid')
    total_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    money_fmt = '#,##0.00'
    thin_border = Border(bottom=Side(style='thin', color='C8CED8'))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.sheet_properties.outlinePr.summaryBelow = False

    if vista == 'cronograma':
        # ── CRONOGRAMA MENSUAL (MATRIZ) ──
        ws.title = f"Cronograma REPEX {repex.anio}"
        data = _get_repex_cronograma_data(repex)
        
        # Title (Col A to G, since 1+1+5 = 7)
        last_col = get_column_letter(7)
        ws.merge_cells(f'A1:{last_col}1')
        ws['A1'].value = f"REPEX {repex.anio} — {repex.nombre} (Cronograma Mensual)"
        ws['A1'].font = Font(name='Calibri', bold=True, size=14, color='2C4A6E')
        ws.row_dimensions[1].height = 30
        ws.append([])

        # Headers
        headers = ['FAMILIA / ITEM', 'TOTAL'] + data['anios_nombres']
        ws.append(headers)
        hr = ws.max_row
        for col, _ in enumerate(headers, 1):
            cell = ws.cell(row=hr, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[hr].height = 25
        ws.column_dimensions['A'].width = 45
        ws.column_dimensions['B'].width = 15
        for col in range(3, 8):
            ws.column_dimensions[get_column_letter(col)].width = 10

        for fam in data['familias_data']:
            # Familia row
            row_data = [fam['familia_nombre'].upper(), fam['total']] + fam['mensual']
            ws.append(row_data)
            r_fam = ws.max_row
            for col in range(1, 8): 
                ws.cell(row=r_fam, column=col).fill = total_fill
                ws.cell(row=r_fam, column=col).font = total_font
                if col >= 2:
                    ws.cell(row=r_fam, column=col).number_format = money_fmt
            
            group_fam_start = ws.max_row + 1
            for cat in fam['categorias']:
                # Categoria Row
                row_cat = [cat['nombre'].upper(), cat['total']] + cat['mensual']
                ws.append(row_cat)
                r_cat = ws.max_row
                ws.cell(row=r_cat, column=1).alignment = Alignment(indent=2)
                for col in range(1, 8): 
                    ws.cell(row=r_cat, column=col).fill = cat_fill
                    ws.cell(row=r_cat, column=col).font = cat_font
                    if col >= 2:
                        ws.cell(row=r_cat, column=col).number_format = money_fmt
                
                group_cat_start = ws.max_row + 1
                for ub in cat['ubicaciones']:
                    # Ubicacion Row
                    row_ub = [ub['nombre'], ub['total']] + ub['mensual']
                    ws.append(row_ub)
                    r_ub = ws.max_row
                    ws.cell(row=r_ub, column=1).alignment = Alignment(indent=4)
                    for col in range(1, 8): 
                        ws.cell(row=r_ub, column=col).font = Font(name='Calibri', bold=True, size=10)
                        if col >= 2:
                            ws.cell(row=r_ub, column=col).number_format = money_fmt
                    
                    group_ub_start = ws.max_row + 1
                    for item in ub['items']:
                        row_item = [item['activo_nombre'], item['total']] + item['mensual']
                        ws.append(row_item)
                        ir = ws.max_row
                        ws.cell(row=ir, column=1).alignment = Alignment(indent=6)
                        for col in range(1, 8): 
                            ws.cell(row=ir, column=col).font = item_font
                            ws.cell(row=ir, column=col).border = thin_border
                            if col >= 2:
                                ws.cell(row=ir, column=col).number_format = money_fmt
                                ws.cell(row=ir, column=col).alignment = Alignment(horizontal='right')
                    
                    group_ub_end = ws.max_row
                    if group_ub_end >= group_ub_start:
                        ws.row_dimensions.group(group_ub_start, group_ub_end, outline_level=3, hidden=True)
                
                group_cat_end = ws.max_row
                if group_cat_end >= group_cat_start:
                    ws.row_dimensions.group(group_cat_start, group_cat_end, outline_level=2, hidden=True)
            
            group_fam_end = ws.max_row
            if group_fam_end >= group_fam_start:
                ws.row_dimensions.group(group_fam_start, group_fam_end, outline_level=1, hidden=True)

        # Totales mensuales
        ws.append([])
        total_row = ['TOTAL ANUAL', data['total_general']] + data['total_mensual']
        ws.append(total_row)
        tr = ws.max_row
        for col in range(1, 8): 
            ws.cell(row=tr, column=col).fill = total_fill
            ws.cell(row=tr, column=col).font = total_font
            if col >= 2:
                ws.cell(row=tr, column=col).number_format = money_fmt

    elif vista == 'flujo':
        # ── FLUJO DETALLADO 24 MESES ──
        ws.title = f"Flujo 24 Meses {repex.anio}"
        data = _get_repex_cronograma_data(repex)
        anio1, anio2 = data['anios_nombres'][0], data['anios_nombres'][1]
        
        # Title (Col A to AB, 28 columns)
        last_col = get_column_letter(28)
        ws.merge_cells(f'A1:{last_col}1')
        ws['A1'].value = f"REPEX {repex.anio} — {repex.nombre} (Flujo Detallado 24 Meses)"
        ws['A1'].font = Font(name='Calibri', bold=True, size=14, color='2C4A6E')
        ws.row_dimensions[1].height = 30
        ws.append([])

        # Headers
        headers = ['FAMILIA / ACTIVO'] + data['meses_year_1'] + [f'TOTAL {anio1}'] + data['meses_year_2'] + [f'TOTAL {anio2}'] + ['TOTAL GENERAL']
        ws.append(headers)
        hr = ws.max_row
        for col, _ in enumerate(headers, 1):
            cell = ws.cell(row=hr, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[hr].height = 25
        ws.column_dimensions['A'].width = 45
        for col in range(2, 29):
            ws.column_dimensions[get_column_letter(col)].width = 12

        for fam in data['familias_data']:
            m24 = fam['mensual_24']
            row_fam = [fam['familia_nombre'].upper()] + m24[:12] + [fam['total_anio_1']] + m24[12:] + [fam['total_anio_2']] + [fam['total']]
            ws.append(row_fam)
            r_fam = ws.max_row
            for col in range(1, 29):
                ws.cell(row=r_fam, column=col).fill = total_fill
                ws.cell(row=r_fam, column=col).font = total_font
                if col >= 2: ws.cell(row=r_fam, column=col).number_format = money_fmt

            for cat in fam['categorias']:
                m24c = cat['mensual_24']
                row_cat = [cat['nombre'].upper()] + m24c[:12] + [cat['total_anio_1']] + m24c[12:] + [cat['total_anio_2']] + [cat['total']]
                ws.append(row_cat)
                r_cat = ws.max_row
                ws.cell(row=r_cat, column=1).alignment = Alignment(indent=2)
                for col in range(1, 29):
                    ws.cell(row=r_cat, column=col).fill = cat_fill
                    ws.cell(row=r_cat, column=col).font = cat_font
                    if col >= 2: ws.cell(row=r_cat, column=col).number_format = money_fmt

                for ub in cat['ubicaciones']:
                    m24u = ub['mensual_24']
                    row_ub = [ub['nombre']] + m24u[:12] + [ub['total_anio_1']] + m24u[12:] + [ub['total_anio_2']] + [ub['total']]
                    ws.append(row_ub)
                    r_ub = ws.max_row
                    ws.cell(row=r_ub, column=1).alignment = Alignment(indent=4)
                    for col in range(1, 29):
                        ws.cell(row=r_ub, column=col).font = Font(name='Calibri', bold=True, size=10)
                        if col >= 2: ws.cell(row=r_ub, column=col).number_format = money_fmt

                    for item in ub['items']:
                        m24i = item['mensual_24']
                        row_item = [item['activo_nombre']] + m24i[:12] + [item['total_anio_1']] + m24i[12:] + [item['total_anio_2']] + [item['total']]
                        ws.append(row_item)
                        r_item = ws.max_row
                        ws.cell(row=r_item, column=1).alignment = Alignment(indent=6)
                        for col in range(1, 29):
                            if col >= 2: ws.cell(row=r_item, column=col).number_format = money_fmt
        
        # Final Row
        m24g = data['total_mensual_24']
        t1, t2 = sum(m24g[:12]), sum(m24g[12:])
        row_final = ['TOTAL GENERAL'] + m24g[:12] + [t1] + m24g[12:] + [t2] + [data['total_general']]
        ws.append(row_final)
        r_fin = ws.max_row
        for col in range(1, 29):
            cell = ws.cell(row=r_fin, column=col)
            cell.fill = header_fill
            cell.font = header_font
            if col >= 2: cell.number_format = money_fmt

    elif vista == 'modelo':
        # ── RESUMEN POR MODELO (Agrupado por Categoría y Modelo) ──
        ws.title = "Resumen por Modelo"
        items_qs = repex.items.select_related(
            'activo', 'activo__modelo', 'activo__modelo__marca',
            'activo__modelo__categoria', 'activo__ubicacion', 'activo__familia',
            'modelo', 'modelo__marca', 'modelo__categoria'
        ).all()

        consolidated = OrderedDict()
        for item in items_qs:
            # Category name logic
            activo = item.activo
            cat_name = '-'
            if activo:
                if activo.modelo and activo.modelo.categoria:
                    cat = activo.modelo.categoria
                    path = [cat.nombre]
                    curr = cat.padre
                    visited = {cat.id}
                    while curr:
                        if curr.id in visited: break
                        visited.add(curr.id)
                        path.append(curr.nombre)
                        curr = curr.padre
                    cat_name = ' → '.join(reversed(path))
            else:
                cat_name = item.categoria_manual
                if not cat_name and item.modelo and item.modelo.categoria:
                    cat = item.modelo.categoria
                    path = [cat.nombre]
                    curr = cat.padre
                    visited = {cat.id}
                    while curr:
                        if curr.id in visited: break
                        visited.add(curr.id)
                        path.append(curr.nombre)
                        curr = curr.padre
                    cat_name = ' → '.join(reversed(path))
                
                if not cat_name:
                    cat_name = 'Sin Categoría'

            if cat_name not in consolidated:
                consolidated[cat_name] = OrderedDict()
            
            # Model identity
            modelo_str = '-'
            marca_str = '-'
            if activo and activo.modelo:
                modelo_str = activo.modelo.nombre
                if activo.modelo.marca:
                    marca_str = activo.modelo.marca.nombre
            elif item.modelo:
                modelo_str = item.modelo.nombre
                if hasattr(item.modelo, 'marca') and item.modelo.marca:
                    marca_str = item.modelo.marca.nombre
            
            modelo_key = f"{marca_str} | {modelo_str}"
            if modelo_key not in consolidated[cat_name]:
                consolidated[cat_name][modelo_key] = {
                    'marca': marca_str,
                    'modelo': modelo_str,
                    'cantidad': 0,
                    'total': 0.0,
                    'items': []
                }
            
            m_data = consolidated[cat_name][modelo_key]
            m_data['cantidad'] += float(item.cantidad or 1)
            m_data['total'] += float(item.costo_reposicion or 0)
            m_data['items'].append({
                'nombre': item.display_nombre,
                'ubicacion': (activo.ubicacion.nombre if activo and activo.ubicacion else (item.ubicacion_manual or '-')),
                'cantidad': float(item.cantidad or 1),
                'costo': float(item.costo_reposicion or 0)
            })

        headers = ['CATEGORÍA / MODELO', 'MARCA', 'UBICACIÓN ESPECÍFICA', 'CANTIDAD', 'P.U.', 'IMPORTE']
        ws.append(headers)
        hr = ws.max_row
        for col_idx in range(1, 7):
            cell = ws.cell(row=hr, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        ws.column_dimensions['A'].width = 45
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 18

        total_general = 0
        for cat_name, models_dict in consolidated.items():
            cat_total = sum(m['total'] for m in models_dict.values())
            cat_cant = sum(m['cantidad'] for m in models_dict.values())
            total_general += cat_total
            
            # Category Row
            ws.append([cat_name.upper(), '', '', cat_cant, '', cat_total])
            r = ws.max_row
            for col in range(1, 7):
                ws.cell(row=r, column=col).fill = total_fill
                ws.cell(row=r, column=col).font = total_font
            ws.cell(row=r, column=6).number_format = money_fmt
            
            for m_key, m_info in models_dict.items():
                # Model Row
                pu = m_info['total'] / m_info['cantidad'] if m_info['cantidad'] > 0 else 0
                ws.append([m_info['modelo'], m_info['marca'], '', m_info['cantidad'], pu, m_info['total']])
                mr = ws.max_row
                ws.cell(row=mr, column=1).alignment = Alignment(indent=2)
                for col in range(1, 7):
                    ws.cell(row=mr, column=col).fill = cat_fill
                    ws.cell(row=mr, column=col).font = cat_font
                ws.cell(row=mr, column=5).number_format = money_fmt
                ws.cell(row=mr, column=6).number_format = money_fmt
                
                # Items breakdown
                group_start = ws.max_row + 1
                for itm in m_info['items']:
                    pu_item = itm['costo'] / itm['cantidad'] if itm['cantidad'] > 0 else 0
                    ws.append(['', '', itm['ubicacion'], itm['cantidad'], pu_item, itm['costo']])
                    ir = ws.max_row
                    ws.cell(row=ir, column=3).font = item_font
                    for col in range(3, 7): ws.cell(row=ir, column=col).border = thin_border
                    ws.cell(row=ir, column=5).number_format = money_fmt
                    ws.cell(row=ir, column=6).number_format = money_fmt
                
                group_end = ws.max_row
                if group_end >= group_start:
                    ws.row_dimensions.group(group_start, group_end, outline_level=1, hidden=True)

        ws.append([])
        ws.append(['TOTAL GENERAL', '', '', '', '', total_general])
        tr = ws.max_row
        for col in range(1, 7):
            ws.cell(row=tr, column=col).fill = total_fill
            ws.cell(row=tr, column=col).font = total_font
        ws.cell(row=tr, column=6).number_format = money_fmt

    else:
        # ── DETALLE / APU with 3-level hierarchy ──
        items_qs = repex.items.select_related(
            'activo', 'activo__modelo', 'activo__modelo__marca',
            'activo__modelo__categoria', 'activo__ubicacion', 'activo__familia',
            'modelo', 'modelo__marca', 'modelo__categoria'
        ).all()

        hierarchical_obj = OrderedDict()
        for item in items_qs:
            # Category name logic
            activo = item.activo
            cat_name = '-'
            if activo:
                if activo.modelo and activo.modelo.categoria:
                    cat = activo.modelo.categoria
                    path = [cat.nombre]
                    curr = cat.padre
                    visited = {cat.id}
                    while curr:
                        if curr.id in visited: break
                        visited.add(curr.id)
                        path.append(curr.nombre)
                        curr = curr.padre
                    cat_name = ' → '.join(reversed(path))
            else:
                cat_name = item.categoria_manual
                if not cat_name and item.modelo and item.modelo.categoria:
                    cat = item.modelo.categoria
                    path = [cat.nombre]
                    curr = cat.padre
                    visited = {cat.id}
                    while curr:
                        if curr.id in visited: break
                        visited.add(curr.id)
                        path.append(curr.nombre)
                        curr = curr.padre
                    cat_name = ' → '.join(reversed(path))
                
                if not cat_name:
                    cat_name = 'Sin Categoría'
            
            # Logic for Building grouping
            if activo and activo.ubicacion:
                edificio_obj = activo.ubicacion
                visited_ub = {edificio_obj.id}
                found_edificio = edificio_obj if edificio_obj.tipo == 'EDIFICIO' else None
                temp_curr = edificio_obj.padre
                while temp_curr:
                    if temp_curr.id in visited_ub: break
                    visited_ub.add(temp_curr.id)
                    if temp_curr.tipo == 'EDIFICIO':
                        found_edificio = temp_curr
                    temp_curr = temp_curr.padre
                
                building_name = found_edificio.nombre if found_edificio else edificio_obj.get_root().nombre
            else:
                building_name = item.ubicacion_manual or '-'
            
            if cat_name not in hierarchical_obj:
                hierarchical_obj[cat_name] = OrderedDict()
            
            if building_name not in hierarchical_obj[cat_name]:
                hierarchical_obj[cat_name][building_name] = {'items': [], 'total_edificio': 0.0}
            
            # Common metadata
            modelo_str = '-'
            marca_str = '-'
            if activo and activo.modelo:
                modelo_str = activo.modelo.nombre
                if activo.modelo.marca:
                    marca_str = activo.modelo.marca.nombre
            elif item.modelo:
                modelo_str = item.modelo.nombre
                if hasattr(item.modelo, 'marca') and item.modelo.marca:
                    marca_str = item.modelo.marca.nombre

            if vista == 'apu':
                val = {
                    'codigo': activo.codigo_interno if activo else '-',
                    'nombre': item.display_nombre,
                    'marca': marca_str,
                    'modelo': modelo_str,
                    'unidades': item.unidades or 'Unidad',
                    'cantidad': float(item.cantidad or 1),
                    'precio_unitario': float(item.precio_unitario or 0),
                    'costo': float(item.costo_reposicion or 0),
                }
            else:
                val = {
                    'codigo': activo.codigo_interno if activo else '-',
                    'nombre': item.display_nombre,
                    'marca': marca_str,
                    'modelo': modelo_str,
                    'ubicacion': building_name,
                    'prioridad': item.prioridad,
                    'costo': float(item.costo_reposicion or 0),
                }
            
            hierarchical_obj[cat_name][building_name]['items'].append(val)
            hierarchical_obj[cat_name][building_name]['total_edificio'] += val['costo']
        
        ws.merge_cells('A1:H1')
        ws['A1'].value = f"REPEX {repex.anio} — {repex.nombre}"
        ws['A1'].font = Font(name='Calibri', bold=True, size=14, color='2C4A6E')
        ws.row_dimensions[1].height = 30
        ws.append([])

        if vista == 'apu':
            ws.title = "Presupuesto"
            headers = ['CODIGO', 'UBICACION', 'DESCRIPCION', 'MODELO', 'UNIDAD', 'CNTD', 'P.U.', 'IMPORTE']
            col_widths = [12, 30, 40, 20, 12, 10, 16, 18]
        else:
            ws.title = "Detalle"
            headers = ['Código', 'Activo', 'Marca', 'Modelo', 'Ruta Ubicación', 'Categoría', 'Prioridad', 'Costo Reposición']
            col_widths = [14, 30, 16, 18, 35, 30, 12, 18]

        ws.append(headers)
        hr = ws.max_row
        for col_idx in range(1, 9):
            cell = ws.cell(row=hr, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[hr].height = 28
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        cat_idx = 0
        total_repex = 0.0
        for cat_name, edificiions_dict in hierarchical_obj.items():
            cat_idx += 1
            cat_total = sum(e['total_edificio'] for e in edificiions_dict.values())
            total_repex += cat_total
            
            # Level 1 Heading
            ws.append([f'{cat_idx}.', '', cat_name.upper(), '', '', '', '', cat_total])
            r = ws.max_row
            ws.merge_cells(f'C{r}:G{r}')
            for col in range(1, 9):
                ws.cell(row=r, column=col).fill = total_fill
                ws.cell(row=r, column=col).font = total_font
            ws.cell(row=r, column=8).number_format = money_fmt
            
            group_1_start = ws.max_row + 1
            
            b_idx = 0
            for b_name, b_data in edificiions_dict.items():
                b_idx += 1
                # Level 2 Heading
                ws.append([f'{cat_idx}.{b_idx}', b_name, '', '', '', '', '', b_data['total_edificio']])
                br = ws.max_row
                ws.merge_cells(f'B{br}:G{br}')
                for col in range(1, 9):
                    ws.cell(row=br, column=col).fill = cat_fill
                    ws.cell(row=br, column=col).font = cat_font
                ws.cell(row=br, column=8).number_format = money_fmt
                ws.cell(row=br, column=1).alignment = Alignment(horizontal='right')
                
                group_2_start = ws.max_row + 1
                for i_idx, item in enumerate(b_data['items'], 1):
                    # Level 3 Data
                    if vista == 'apu':
                        row = [f'{cat_idx}.{b_idx}.{i_idx}', item['codigo'], item['nombre'], item['modelo'], item['unidades'], item['cantidad'], item['precio_unitario'], item['costo']]
                    else:
                        row = [f'{cat_idx}.{b_idx}.{i_idx}', item['nombre'], item['marca'], item['modelo'], item['ubicacion'], cat_name, item['prioridad'], item['costo']]
                    
                    ws.append(row)
                    ir = ws.max_row
                    for col in range(1, 9):
                        ws.cell(row=ir, column=col).font = item_font
                        ws.cell(row=ir, column=col).border = thin_border
                    
                    if vista == 'apu':
                        ws.cell(row=ir, column=6).number_format = money_fmt # CNTD
                        ws.cell(row=ir, column=7).number_format = money_fmt # P.U.
                    ws.cell(row=ir, column=8).number_format = money_fmt # IMPORTE
                    ws.cell(row=ir, column=1).alignment = Alignment(horizontal='right')

                group_2_end = ws.max_row
                if group_2_end >= group_2_start:
                    ws.row_dimensions.group(group_2_start, group_2_end, outline_level=2, hidden=False)

            group_1_end = ws.max_row
            if group_1_end >= group_1_start:
                ws.row_dimensions.group(group_1_start, group_1_end, outline_level=1, hidden=False)

        ws.append([])
        ws.append(['', '', '', '', '', '', 'TOTAL GENERAL', total_repex])
        r = ws.max_row
        for col in range(1, 9):
            ws.cell(row=r, column=col).fill = total_fill
            ws.cell(row=r, column=col).font = total_font
        ws.cell(row=r, column=8).number_format = money_fmt

    # Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    vista_labels = {
        'cronograma': 'Cronograma', 
        'flujo': 'Flujo_24M',
        'modelo': 'Resumen_Modelo',
        'apu': 'Presupuesto', 
        'detalle': 'Detalle'
    }
    filename = f"REPEX_{repex.anio}_{vista_labels.get(vista, 'Export')}_{repex.nombre.replace(' ', '_')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename=\"{filename}\"'
    wb.save(response)
    return response
    return response


def _get_repex_cronograma_data(repex, fecha_desde=None, fecha_hasta=None):
    """
    Genera datos matriciales de un plan REPEX agrupados por Familia del Activo.
    Soporta filtrado opcional por rango de fechas.
    """
    anios_rango = [2026, 2027, 2028, 2029, 2030]
    anios_nombres = [str(a) for a in anios_rango]
    
    # ── NUEVO: Flujo 24 Meses (Año Inicial + Año Siguiente) ──
    year_start = repex.anio
    meses_nombres_24 = []
    meses_year_1 = []
    meses_year_2 = []
    for y in [year_start, year_start + 1]:
        for m in range(1, 13):
            # Ene 26, Feb 26...
            mes_str = f"{datetime(y, m, 1).strftime('%b')} {str(y)[2:]}"
            meses_nombres_24.append(mes_str)
            if y == year_start: meses_year_1.append(mes_str)
            else: meses_year_2.append(mes_str)
    
    # Ampliamos el filtro para cubrir el rango de años solicitado
    items = repex.items.select_related('activo', 'activo__familia', 'modelo', 'modelo__categoria').filter(
        Q(fecha_proyectada__year__in=anios_rango) | Q(fecha_proyectada__isnull=True)
    )
    
    if fecha_desde:
        items = items.filter(fecha_proyectada__gte=fecha_desde)
    if fecha_hasta:
        items = items.filter(fecha_proyectada__lte=fecha_hasta)
    
    items = items.all()

    # Agrupar por Familia > Categoría > Ubicación
    familias = {}
    for item in items:
        # 1. Familia
        if item.activo:
            familia_nombre = item.activo.familia.nombre if item.activo.familia else "Sin Familia"
            familia_id = item.activo.familia.id if item.activo.familia else 0
        else:
            if item.modelo and item.modelo.categoria:
                familia_nombre = item.modelo.categoria.nombre
                familia_id = item.modelo.categoria.id
            else:
                familia_nombre = item.categoria_manual or "Ítems Manuales"
                familia_id = -1
        
        fam_key = (familia_id, familia_nombre)
        if fam_key not in familias:
            familias[fam_key] = {
                'familia_nombre': familia_nombre,
                'familia_id': familia_id,
                'categorias': {},
                'mensual': [0.0] * 5, 
                'mensual_24': [0.0] * 24, # NUEVO: 24 meses
                'total_anio_1': 0.0,
                'total_anio_2': 0.0,
                'total': 0.0,
            }

        # 2. Categoría (Ruta completa o categoría principal)
        if item.activo and item.activo.modelo and item.activo.modelo.categoria:
            cat_nombre = item.activo.modelo.categoria.nombre
        elif item.modelo and item.modelo.categoria:
            cat_nombre = item.modelo.categoria.nombre
        else:
            cat_nombre = item.categoria_manual or "Sin Categoría"
        
        cat_key = cat_nombre
        if cat_key not in familias[fam_key]['categorias']:
            familias[fam_key]['categorias'][cat_key] = {
                'nombre': cat_nombre,
                'ubicaciones': {},
                'mensual': [0.0] * 5,
                'mensual_24': [0.0] * 24,
                'total_anio_1': 0.0,
                'total_anio_2': 0.0,
                'total': 0.0,
            }

        # 3. Ubicación
        if item.activo and item.activo.ubicacion:
            edificio_obj = item.activo.ubicacion
            visited_ub = {edificio_obj.id}
            found_edificio = edificio_obj if edificio_obj.tipo == 'EDIFICIO' else None
            temp_curr = edificio_obj.padre
            while temp_curr:
                if temp_curr.id in visited_ub: break
                visited_ub.add(temp_curr.id)
                if temp_curr.tipo == 'EDIFICIO':
                    found_edificio = temp_curr
                temp_curr = temp_curr.padre
            ubicacion_nombre = found_edificio.nombre if found_edificio else edificio_obj.get_root().nombre
        else:
            ubicacion_nombre = item.ubicacion_manual or "Sin Ubicación"

        ub_key = ubicacion_nombre
        if ub_key not in familias[fam_key]['categorias'][cat_key]['ubicaciones']:
            familias[fam_key]['categorias'][cat_key]['ubicaciones'][ub_key] = {
                'nombre': ubicacion_nombre,
                'items': [],
                'mensual': [0.0] * 5,
                'mensual_24': [0.0] * 24,
                'total_anio_1': 0.0,
                'total_anio_2': 0.0,
                'total': 0.0,
            }

        # 4. Determinar monto anual (horizonte 5 años)
        item_anual = [0.0] * 5
        costo = float(item.costo_reposicion or 0)
        anio_proyectado = None

        if item.fecha_proyectada:
            fp = item.fecha_proyectada
            if fp.year in anios_rango:
                anio_idx = anios_rango.index(fp.year)
                item_anual[anio_idx] = costo

        # Determinar posición en flujo 24 meses
        item_24m = [0.0] * 24
        if item.fecha_proyectada:
            fp = item.fecha_proyectada
            if fp.year == year_start:
                item_24m[fp.month - 1] = costo
            elif fp.year == (year_start + 1):
                item_24m[11 + fp.month] = costo # 12 + month - 1

        # Agregar Item
        familias[fam_key]['categorias'][cat_key]['ubicaciones'][ub_key]['items'].append({
            'id': item.id,
            'activo_nombre': item.display_nombre,
            'descripcion': item.descripcion or '',
            'prioridad': item.prioridad,
            'costo_reposicion': costo,
            'mensual': item_anual, 
            'mensual_24': item_24m,
            'total_anio_1': sum(item_24m[:12]),
            'total_anio_2': sum(item_24m[12:]),
            'total': costo,
            'anio_proyectado': anio_proyectado,
        })

        # Acumular Subtotales
        for i in range(5):
            val = item_anual[i]
            familias[fam_key]['mensual'][i] += val
            familias[fam_key]['categorias'][cat_key]['mensual'][i] += val
            familias[fam_key]['categorias'][cat_key]['ubicaciones'][ub_key]['mensual'][i] += val
        
        for i in range(24):
            val24 = item_24m[i]
            familias[fam_key]['mensual_24'][i] += val24
            familias[fam_key]['categorias'][cat_key]['mensual_24'][i] += val24
            familias[fam_key]['categorias'][cat_key]['ubicaciones'][ub_key]['mensual_24'][i] += val24

        t1, t2 = sum(item_24m[:12]), sum(item_24m[12:])
        familias[fam_key]['total_anio_1'] += t1
        familias[fam_key]['total_anio_2'] += t2
        familias[fam_key]['categorias'][cat_key]['total_anio_1'] += t1
        familias[fam_key]['categorias'][cat_key]['total_anio_2'] += t2
        familias[fam_key]['categorias'][cat_key]['ubicaciones'][ub_key]['total_anio_1'] += t1
        familias[fam_key]['categorias'][cat_key]['ubicaciones'][ub_key]['total_anio_2'] += t2

        familias[fam_key]['total'] += costo
        familias[fam_key]['categorias'][cat_key]['total'] += costo
        familias[fam_key]['categorias'][cat_key]['ubicaciones'][ub_key]['total'] += costo

    # Convertir a listas ordenadas
    familias_data = []
    for f_key in sorted(familias.keys(), key=lambda x: x[1]):
        f_val = familias[f_key]
        cats_list = []
        for c_key in sorted(f_val['categorias'].keys()):
            c_val = f_val['categorias'][c_key]
            ubs_list = []
            for u_key in sorted(c_val['ubicaciones'].keys()):
                ubs_list.append(c_val['ubicaciones'][u_key])
            c_val['ubicaciones'] = ubs_list
            cats_list.append(c_val)
        f_val['categorias'] = cats_list
        familias_data.append(f_val)

    # Totales globales
    total_anual = [0.0] * 5
    total_24m = [0.0] * 24
    total_general = 0.0
    total_items = 0

    for fam in familias_data:
        # Contar items en todos los niveles
        for cat in fam['categorias']:
            for ub in cat['ubicaciones']:
                total_items += len(ub['items'])
        
        for i in range(5):
            total_anual[i] += fam['mensual'][i]
        for i in range(24):
            total_24m[i] += fam['mensual_24'][i]
        total_general += fam['total']

    return {
        'familias_data': familias_data,
        'anios_nombres': anios_nombres,
        'meses_nombres_24': meses_nombres_24,
        'meses_year_1': meses_year_1,
        'meses_year_2': meses_year_2,
        'total_mensual': total_anual, 
        'total_mensual_24': total_24m,
        'total_general': total_general,
        'total_items': total_items,
    }


@login_required
def api_update_repex_item(request):
    """Actualiza costo_reposicion y fecha_proyectada de un REPEXItem."""
    if request.method == "POST":
        import json
        from decimal import Decimal
        from .models import REPEXItem
        from django.http import JsonResponse
        from datetime import date

        try:
            data = json.loads(request.body)
            item_id = data.get('item_id')
            mes = int(data.get('mes') or 0)
            monto = Decimal(str(data.get('monto', 0)))

            item = get_object_or_404(REPEXItem, pk=item_id)
            tipo = data.get('tipo', 'anual') # 'anual' o 'mensual'
            
            if monto > 0:
                if item.cantidad > 0:
                    item.precio_unitario = monto / item.cantidad
                else:
                    item.costo_reposicion = monto
                
                if tipo == 'mensual':
                    # mes es el índice 0 a 23
                    year_start = item.repex.anio
                    target_year = year_start + (mes // 12)
                    target_month = (mes % 12) + 1
                    item.fecha_proyectada = date(target_year, target_month, 1)
                else:
                    # El cronograma anual usa 5 años: 2026 (1), 2027 (2), 2028 (3), 2029 (4), 2030 (5)
                    if 1 <= mes <= 5:
                        anios_rango = [2026, 2027, 2028, 2029, 2030]
                        target_year = anios_rango[mes - 1]
                        current_month = item.fecha_proyectada.month if item.fecha_proyectada else 1
                        item.fecha_proyectada = date(target_year, current_month, 1)
            elif monto == 0:
                item.fecha_proyectada = None
                item.precio_unitario = 0
                item.costo_reposicion = 0

            item.save()
            return JsonResponse({'status': 'ok', 'new_total': float(item.costo_reposicion)})

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

    return JsonResponse({'status': 'error', 'message': 'Invalid method'}, status=405)


@login_required
def api_update_repex_item_apu(request):
    """Actualiza cantidad o precio_unitario de un REPEXItem."""
    if request.method == "POST":
        import json
        from decimal import Decimal
        from .models import REPEXItem
        from django.http import JsonResponse

        try:
            data = json.loads(request.body)
            item_id = data.get('item_id')
            field = data.get('field')  # 'cantidad' o 'precio_unitario'
            value = Decimal(str(data.get('value', 0)))

            item = get_object_or_404(REPEXItem, pk=item_id)
            if field == 'cantidad':
                item.cantidad = value
            elif field == 'precio_unitario':
                item.precio_unitario = value
            else:
                return JsonResponse({'status': 'error', 'message': 'Campo inválido'}, status=400)

            item.save()  # El método save() recalcula costo_reposicion (Decimal * Decimal)
            return JsonResponse({
                'status': 'ok', 
                'new_total': float(item.costo_reposicion),
                'cantidad': float(item.cantidad),
                'precio_unitario': float(item.precio_unitario)
            })

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

    return JsonResponse({'status': 'error', 'message': 'Invalid method'}, status=405)


@login_required
def api_add_repex_item(request):
    """Agrega un activo existente al plan REPEX."""
    if request.method == "POST":
        import json
        from .models import REPEX, REPEXItem
        from activos.models.activo import Activo
        from django.http import JsonResponse

        try:
            data = json.loads(request.body)
            repex_id = data.get('repex_id')
            activo_id = data.get('activo_id')
            costo = float(data.get('costo', 0))

            repex = get_object_or_404(REPEX, pk=repex_id)
            activo = get_object_or_404(Activo, pk=activo_id)

            # Si el costo no viene en el request (es 0), intentar usar el precio promedio del modelo
            if costo <= 0 and activo.modelo and activo.modelo.precio_promedio:
                costo = float(activo.modelo.precio_promedio)

            # Verificar que no exista ya
            if REPEXItem.objects.filter(repex=repex, activo=activo).exists():
                return JsonResponse({'status': 'error', 'message': 'Este activo ya está en el plan REPEX.'}, status=400)

            REPEXItem.objects.create(
                repex=repex,
                activo=activo,
                costo_reposicion=costo,
                costo_original=activo.costo or 0,
            )
            return JsonResponse({'status': 'ok', 'message': 'Activo agregado al plan'})

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

    return JsonResponse({'status': 'error', 'message': 'Invalid method'}, status=405)


@login_required
def api_delete_repex_item(request):
    """Elimina un item del plan REPEX."""
    if request.method == "POST":
        import json
        from .models import REPEXItem
        from django.http import JsonResponse

        try:
            data = json.loads(request.body)
            item_id = data.get('item_id')
            item = get_object_or_404(REPEXItem, pk=item_id)
            item.delete()
            return JsonResponse({'status': 'ok', 'message': 'Ítem eliminado del plan'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

    return JsonResponse({'status': 'error', 'message': 'Invalid method'}, status=405)


@login_required
def api_search_activos(request):
    """Busca activos por nombre o código para el selector del modal REPEX."""
    from activos.models.activo import Activo
    from django.http import JsonResponse

    q = request.GET.get('q', '').strip()
    if len(q) < 2:
        return JsonResponse({'results': []})

    activos = Activo.objects.filter(
        models.Q(nombre__icontains=q) | models.Q(codigo_interno__icontains=q)
    ).select_related('familia')[:20]

    results = [{
        'id': a.id,
        'text': f"{a.nombre} ({a.codigo_interno})",
        'familia': a.familia.nombre if a.familia else 'Sin Familia',
        'costo': float(a.costo or 0),
    } for a in activos]

    return JsonResponse({'results': results})


@login_required
def api_import_repex_items(request):
    """
    Importa items REPEX masivamente desde un archivo Excel.
    Columnas esperadas: Activo (codigo_interno), Costo (costo_reposicion).
    Activos duplicados dentro del mismo plan se omiten.
    """
    if request.method != "POST":
        from django.http import JsonResponse
        return JsonResponse({'status': 'error', 'message': 'Método inválido'}, status=405)

    import openpyxl
    from io import BytesIO
    from .models import REPEX, REPEXItem
    from activos.models.activo import Activo
    from django.http import JsonResponse

    try:
        repex_id = request.POST.get('repex_id')
        archivo = request.FILES.get('archivo')

        if not archivo:
            return JsonResponse({'status': 'error', 'message': 'No se envió ningún archivo.'}, status=400)

        repex = get_object_or_404(REPEX, pk=repex_id)

        # Leer Excel
        wb = openpyxl.load_workbook(BytesIO(archivo.read()), read_only=True, data_only=True)
        ws = wb.active

        # Obtener encabezados de la primera fila
        headers = [str(cell.value or '').strip().lower() for cell in next(ws.iter_rows(min_row=1, max_row=1))]

        # Buscar índices de las columnas
        col_activo = None
        col_costo = None
        for i, h in enumerate(headers):
            if h in ('activo', 'codigo', 'codigo_interno', 'código', 'código_interno'):
                col_activo = i
            if h in ('costo', 'costo_reposicion', 'costo reposicion', 'costo_reposición', 'precio', 'monto'):
                col_costo = i

        if col_activo is None:
            return JsonResponse({
                'status': 'error',
                'message': f'No se encontró la columna "Activo". Columnas encontradas: {", ".join(headers)}'
            }, status=400)

        if col_costo is None:
            return JsonResponse({
                'status': 'error',
                'message': f'No se encontró la columna "Costo". Columnas encontradas: {", ".join(headers)}'
            }, status=400)

        # Cargar activos existentes en el plan para detectar duplicados
        existing_activo_ids = set(
            REPEXItem.objects.filter(repex=repex).values_list('activo_id', flat=True)
        )

        imported = 0
        skipped = 0
        errors = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row is None or all(c is None for c in row):
                continue

            codigo_raw = str(row[col_activo] or '').strip()
            costo_raw = row[col_costo]

            if not codigo_raw:
                continue

            # Limpiar código (quitar .0 si viene como float de Excel)
            if '.' in codigo_raw:
                try:
                    codigo_raw = str(int(float(codigo_raw)))
                except (ValueError, OverflowError):
                    pass

            # Buscar activo por código interno
            try:
                activo = Activo.objects.get(codigo_interno=codigo_raw)
            except Activo.DoesNotExist:
                errors.append(f"Fila {row_idx}: Activo '{codigo_raw}' no encontrado.")
                continue
            except Activo.MultipleObjectsReturned:
                errors.append(f"Fila {row_idx}: Múltiples activos con código '{codigo_raw}'.")
                continue

            # Verificar duplicado
            if activo.id in existing_activo_ids:
                skipped += 1
                continue

            # Parsear costo
            try:
                costo = float(costo_raw or 0)
            except (ValueError, TypeError):
                costo = 0

            REPEXItem.objects.create(
                repex=repex,
                activo=activo,
                costo_reposicion=costo,
                costo_original=activo.costo or 0,
            )
            existing_activo_ids.add(activo.id)
            imported += 1

        wb.close()

        summary = f"✅ {imported} importados"
        if skipped > 0:
            summary += f", ⏭ {skipped} omitidos (duplicados)"
        if errors:
            summary += f", ❌ {len(errors)} errores"

        return JsonResponse({
            'status': 'ok',
            'message': summary,
            'imported': imported,
            'skipped': skipped,
            'errors': errors[:20],  # Limitar a 20 errores
        })

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@login_required
@require_POST
def api_add_manual_repex_item(request):
    """Agrega un ítem manual al plan REPEX (sin activo vinculado)."""
    from .models import REPEXItem, REPEX
    from activos.models import Modelo
    import json

    try:
        data = json.loads(request.body)
        repex_id = data.get('repex_id')
        modelo_id = data.get('modelo_id')
        nombre = data.get('nombre_item', '').strip()
        ubicacion = data.get('ubicacion_manual', '').strip()
        categoria = data.get('categoria_manual', '').strip()
        unidades_val = data.get('unidades', '').strip()
        cantidad_val = data.get('cantidad', 1)
        precio_val = data.get('precio_unitario', 0)
        prioridad = data.get('prioridad', 'MEDIA')

        if not repex_id:
            return JsonResponse({'status': 'error', 'message': 'El ID del REPEX es requerido.'}, status=400)

        repex = REPEX.objects.get(pk=repex_id)
        
        modelo_obj = None
        if modelo_id:
            modelo_obj = Modelo.objects.filter(pk=modelo_id).first()
            if modelo_obj:
                if not nombre:
                    nombre = f"{modelo_obj.marca.nombre} {modelo_obj.nombre}" if hasattr(modelo_obj, 'marca') else modelo_obj.nombre
                
                if not categoria and hasattr(modelo_obj, 'categoria') and modelo_obj.categoria:
                    categoria = modelo_obj.categoria.nombre
                    
                if not unidades_val and hasattr(modelo_obj, 'unidad_medida') and modelo_obj.unidad_medida:
                    unidades_val = modelo_obj.unidad_medida.nombre

                if (not precio_val or float(precio_val) == 0) and modelo_obj.precio_promedio:
                    precio_val = float(modelo_obj.precio_promedio)

        if not nombre:
             return JsonResponse({'status': 'error', 'message': 'Nombre del ítem es requerido.'}, status=400)

        item = REPEXItem(
            repex=repex,
            activo=None,
            modelo=modelo_obj,
            nombre_item=nombre,
            ubicacion_manual=ubicacion,
            categoria_manual=categoria,
            unidades=unidades_val,
            cantidad=cantidad_val,
            precio_unitario=precio_val,
            prioridad=prioridad,
        )
        item.save()  # save() auto-calcula costo_reposicion

        return JsonResponse({
            'status': 'ok',
            'message': f'Ítem "{nombre}" agregado (${item.costo_reposicion:,.2f})',
            'item_id': item.id,
        })

    except REPEX.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Plan REPEX no encontrado.'}, status=404)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

@login_required
def requisicion_documento_proxy(request, doc_id):
    from .models import DocumentoRequisicion
    from django.http import FileResponse, Http404
    import mimetypes
    
    doc = get_object_or_404(DocumentoRequisicion, id=doc_id)
    if not doc.archivo:
        raise Http404("Documento no tiene archivo")
        
    try:
        file_handle = doc.archivo.open("rb")
        content_type, _ = mimetypes.guess_type(doc.archivo.name)
        response = FileResponse(file_handle, content_type=content_type)
        filename = doc.archivo.name.split('/')[-1]
        
        # Si se solicita descarga forzada
        if request.GET.get('download') == '1':
            response["Content-Disposition"] = f"attachment; filename=\"{filename}\""
        else:
            response["Content-Disposition"] = f"inline; filename=\"{filename}\""
        
        # Cabeceras de seguridad
        response["X-Frame-Options"] = "SAMEORIGIN"
        response["Content-Security-Policy"] = "frame-ancestors 'self'"
        return response
    except Exception as e:
        raise Http404(f"Error al acceder al archivo: {str(e)}")


# ──────────────────────────────────────────────
# Cotizaciones
# ──────────────────────────────────────────────

def staff_required(view):
    return user_passes_test(lambda u: u.is_staff)(view)


@staff_required
def lista_cotizaciones(request):
    cotizaciones = Cotizacion.objects.select_related('proyecto', 'disciplina', 'creado_por').all()
    disciplina_id = request.GET.get('disciplina')
    estado = request.GET.get('estado')
    q = request.GET.get('q')
    if disciplina_id:
        cotizaciones = cotizaciones.filter(disciplina_id=disciplina_id)
    if estado:
        cotizaciones = cotizaciones.filter(estado=estado)
    if q:
        cotizaciones = cotizaciones.filter(Q(numero__icontains=q) | Q(proyecto__nombre__icontains=q) | Q(notas__icontains=q))
    from documentos.models import Disciplina
    return render(request, 'presupuestos/lista_cotizaciones.html', {
        'cotizaciones': cotizaciones,
        'disciplinas': Disciplina.objects.all(),
        'ESTADOS': Cotizacion.ESTADOS,
    })


@staff_required
def crear_cotizacion(request):
    from documentos.models import Disciplina
    from proyectos.models import Proyecto
    if request.method == 'POST':
        proyecto_id = request.POST.get('proyecto')
        disciplina_id = request.POST.get('disciplina')
        fecha = request.POST.get('fecha')
        version = request.POST.get('version') or 1
        valida_hasta = request.POST.get('valida_hasta') or None
        notas = request.POST.get('notas', '')
        items_json = request.POST.get('items', '[]')
        import json
        items = json.loads(items_json)

        if not fecha:
            messages.error(request, 'La fecha es requerida.')
            return redirect('presupuestos:crear_cotizacion')

        ultimo = Cotizacion.objects.aggregate(models.Max('id'))['id__max'] or 0
        numero = f"COT-{datetime.now().year}-{ultimo + 1:04d}"

        cotizacion = Cotizacion.objects.create(
            numero=numero,
            proyecto_id=proyecto_id or None,
            disciplina=None,
            fecha=fecha,
            version=int(version),
            valida_hasta=valida_hasta,
            notas=notas,
            creado_por=request.user,
        )

        for idx, item in enumerate(items):
                predef_id = item.get('item_predefinido_id') or None
                ItemCotizacion.objects.create(
                    cotizacion=cotizacion,
                    item_predefinido_id=predef_id,
                    disciplina_id=item.get('disciplina_id') or None,
                    area=item.get('area', '') or '',
                    descripcion=item['descripcion'],
                    unidad_medida=item['unidad_medida'],
                    cantidad=item['cantidad'],
                    precio_unitario=item['precio_unitario'],
                    descuento_porcentaje=item.get('descuento_porcentaje', 0),
                    orden=idx,
                )

        messages.success(request, f'Cotización {numero} creada exitosamente.')
        return redirect('presupuestos:ver_cotizacion', pk=cotizacion.pk)

    disciplinas = Disciplina.objects.all()
    proyectos = Proyecto.objects.all()
    # Proyecto pre-seleccionado desde ?proyecto=ID (viene del detalle de proyecto)
    proyecto_id_preselect = request.GET.get('proyecto')
    from activos.models import Ubicacion
    ubicaciones = Ubicacion.objects.all().order_by('nombre')
    return render(request, 'presupuestos/form_cotizacion.html', {
        'disciplinas': disciplinas,
        'proyectos': proyectos,
        'ubicaciones': ubicaciones,
        'proyecto_id_preselect': proyecto_id_preselect,
        'es_nuevo': True,
    })


@staff_required
def editar_cotizacion(request, pk):
    from documentos.models import Disciplina
    from proyectos.models import Proyecto
    cotizacion = get_object_or_404(Cotizacion.objects.select_related('proyecto', 'disciplina', 'creado_por'), pk=pk)

    if request.method == 'POST':
        accion = request.POST.get('accion')
        if accion == 'guardar_cabecera':
            cotizacion.proyecto_id = request.POST.get('proyecto') or None
            cotizacion.fecha = request.POST.get('fecha')
            cotizacion.version = int(request.POST.get('version') or 1)
            cotizacion.valida_hasta = request.POST.get('valida_hasta') or None
            cotizacion.notas = request.POST.get('notas', '')
            cotizacion.save()
            messages.success(request, 'Cabecera actualizada.')
        elif accion == 'cambiar_estado':
            nuevo_estado = request.POST.get('estado')
            if nuevo_estado in dict(Cotizacion.ESTADOS):
                cotizacion.estado = nuevo_estado
                cotizacion.save()
                messages.success(request, f'Estado cambiado a {cotizacion.get_estado_display()}.')
        elif accion == 'guardar_items':
            items_json = request.POST.get('items', '[]')
            import json
            items = json.loads(items_json)
            cotizacion.items.all().delete()
            for idx, item in enumerate(items):
                predef_id = item.get('item_predefinido_id') or None
                ItemCotizacion.objects.create(
                    cotizacion=cotizacion,
                    item_predefinido_id=predef_id,
                    disciplina_id=item.get('disciplina_id') or None,
                    area=item.get('area', '') or '',
                    descripcion=item['descripcion'],
                    unidad_medida=item['unidad_medida'],
                    cantidad=item['cantidad'],
                    precio_unitario=item['precio_unitario'],
                    descuento_porcentaje=item.get('descuento_porcentaje', 0),
                    orden=idx,
                )
            messages.success(request, 'Items actualizados.')
        elif accion == 'eliminar':
            cotizacion.delete()
            messages.success(request, 'Cotización eliminada.')
            return redirect('presupuestos:lista_cotizaciones')
        return redirect('presupuestos:editar_cotizacion', pk=cotizacion.pk)

    from collections import defaultdict
    import json as json_mod

    disciplinas = Disciplina.objects.all()
    proyectos = Proyecto.objects.all()
    from activos.models import Ubicacion
    ubicaciones = Ubicacion.objects.all().order_by('nombre')
    items_qs = cotizacion.items.select_related('disciplina').order_by('orden')

    grupos = defaultdict(list)
    for item in items_qs:
        key = (item.area or '', item.disciplina_id)
        grupos[key].append({
            'id': item.id,
            'area': item.area or '',
            'disciplina_id': item.disciplina_id,
            'disciplina_nombre': item.disciplina.nombre if item.disciplina else 'Sin disciplina',
            'item_predefinido_id': item.item_predefinido_id,
            'descripcion': item.descripcion,
            'unidad_medida': item.unidad_medida,
            'cantidad': float(item.cantidad),
            'precio_unitario': float(item.precio_unitario),
            'descuento_porcentaje': float(item.descuento_porcentaje),
        })

    items_por_disciplina = []
    for (area, disc_id), items_list in grupos.items():
        items_por_disciplina.append({
            'area': area,
            'disciplina_id': disc_id,
            'disciplina_nombre': items_list[0]['disciplina_nombre'],
            'items': items_list,
        })

    return render(request, 'presupuestos/form_cotizacion.html', {
        'cotizacion': cotizacion,
        'disciplinas': disciplinas,
        'proyectos': proyectos,
        'ubicaciones': ubicaciones,
        'items_por_disciplina_json': json_mod.dumps(items_por_disciplina),
        'es_nuevo': False,
        'ESTADOS': Cotizacion.ESTADOS,
    })


@staff_required
def ver_cotizacion(request, pk):
    cotizacion = get_object_or_404(Cotizacion.objects.select_related('proyecto', 'disciplina', 'creado_por'), pk=pk)
    items = cotizacion.items.select_related('disciplina').order_by('area', 'disciplina__nombre', 'orden')
    
    # Calcular subtotales por área
    from itertools import groupby
    items_list = list(items)
    subtotales_area = {}
    for area, area_items in groupby(items_list, key=lambda x: x.area):
        subtotales_area[area or ''] = sum(item.total for item in area_items)
    
    return render(request, 'presupuestos/ver_cotizacion.html', {
        'cotizacion': cotizacion,
        'items': items,
        'subtotales_area': subtotales_area,
    })


@staff_required
def cotizacion_pdf(request, pk):
    cotizacion = get_object_or_404(Cotizacion.objects.select_related('proyecto', 'disciplina', 'creado_por'), pk=pk)
    items = cotizacion.items.select_related('disciplina').order_by('area', 'disciplina__nombre', 'orden')
    
    # Calcular subtotales por área para el PDF (no tiene JS)
    from itertools import groupby
    from collections import OrderedDict
    items_list = list(items)
    
    areas_data = OrderedDict()
    area_counter = 0
    for area, area_items_iter in groupby(items_list, key=lambda x: x.area):
        area_counter += 1
        area_items_list = list(area_items_iter)
        area_total = sum(item.total for item in area_items_list)
        
        # Agrupar por disciplina dentro del área
        disc_counter = 0
        disciplinas_data = []
        for disc, disc_items_iter in groupby(area_items_list, key=lambda x: x.disciplina):
            disc_counter += 1
            disc_items = list(disc_items_iter)
            disciplinas_data.append({
                'disciplina': disc,
                'num': f"{area_counter}.{disc_counter}",
                'items': [{'item': it, 'num': f"{area_counter}.{disc_counter}.{i+1}"} for i, it in enumerate(disc_items)],
            })
        
        areas_data[area or ''] = {
            'nombre': area or 'Sin área',
            'num': area_counter,
            'total': area_total,
            'disciplinas': disciplinas_data,
        }
    
    return render(request, 'presupuestos/cotizacion_pdf.html', {
        'cotizacion': cotizacion,
        'items': items,
        'areas_data': areas_data,
    })


@staff_required
def cotizacion_excel(request, pk):
    """Exportar cotización a Excel con filas colapsables por Nivel (área) y Disciplina."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from itertools import groupby

    cotizacion = get_object_or_404(Cotizacion.objects.select_related('proyecto'), pk=pk)
    items = cotizacion.items.select_related('disciplina').order_by('area', 'disciplina__nombre', 'orden')
    items_list = list(items)

    wb = Workbook()
    ws = wb.active
    ws.title = "Cotización"

    # Estilos
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    area_font = Font(bold=True, size=11, color="1B5E20")
    area_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    disc_font = Font(bold=True, size=10, color="0D47A1")
    disc_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    money_fmt = '#,##0.00'
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Encabezado del documento
    ws.merge_cells('A1:G1')
    ws['A1'] = f"Cotización #{cotizacion.id} — {cotizacion.proyecto or 'Sin proyecto'}"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A2:G2')
    ws['A2'] = f"Fecha: {cotizacion.fecha} | Estado: {cotizacion.get_estado_display()} | Total: L. {cotizacion.total:,.2f}"
    ws['A2'].font = Font(size=10, italic=True)

    # Headers de columnas
    headers = ['#', 'Descripción', 'Unidad', 'Cantidad', 'Precio Unit.', 'Desc. %', 'Total']
    row_num = 4
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Anchos de columna
    col_widths = [6, 55, 10, 12, 15, 10, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row_num += 1
    area_counter = 0

    for area, area_items_iter in groupby(items_list, key=lambda x: x.area):
        area_counter += 1
        area_items_list = list(area_items_iter)
        area_total = sum(float(item.total) for item in area_items_list)

        # Fila de Área (Nivel 1 - outline level 1)
        area_name = area or 'Sin área'
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=6)
        cell = ws.cell(row=row_num, column=1, value=f"{area_counter}. {area_name}")
        cell.font = area_font
        cell.fill = area_fill
        for c in range(1, 8):
            ws.cell(row=row_num, column=c).fill = area_fill
            ws.cell(row=row_num, column=c).border = thin_border
        ws.cell(row=row_num, column=7, value=area_total)
        ws.cell(row=row_num, column=7).font = area_font
        ws.cell(row=row_num, column=7).number_format = money_fmt
        ws.cell(row=row_num, column=7).fill = area_fill
        row_num += 1

        # Agrupar por disciplina dentro del área
        disc_counter = 0
        for disc, disc_items_iter in groupby(area_items_list, key=lambda x: x.disciplina):
            disc_counter += 1
            disc_items = list(disc_items_iter)
            disc_total = sum(float(item.total) for item in disc_items)
            disc_name = disc.nombre if disc else 'Sin disciplina'

            # Fila de Disciplina (Nivel 2 - outline level 1)
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=6)
            cell = ws.cell(row=row_num, column=1, value=f"  {area_counter}.{disc_counter} {disc_name}")
            cell.font = disc_font
            cell.fill = disc_fill
            for c in range(1, 8):
                ws.cell(row=row_num, column=c).fill = disc_fill
                ws.cell(row=row_num, column=c).border = thin_border
            ws.cell(row=row_num, column=7, value=disc_total)
            ws.cell(row=row_num, column=7).font = disc_font
            ws.cell(row=row_num, column=7).number_format = money_fmt
            ws.cell(row=row_num, column=7).fill = disc_fill
            ws.row_dimensions[row_num].outline_level = 1
            row_num += 1

            # Items individuales (outline level 2 - colapsables)
            for i, item in enumerate(disc_items, 1):
                ws.cell(row=row_num, column=1, value=f"{area_counter}.{disc_counter}.{i}")
                ws.cell(row=row_num, column=2, value=item.descripcion)
                ws.cell(row=row_num, column=3, value=item.unidad_medida)
                ws.cell(row=row_num, column=4, value=float(item.cantidad))
                ws.cell(row=row_num, column=5, value=float(item.precio_unitario))
                ws.cell(row=row_num, column=5).number_format = money_fmt
                ws.cell(row=row_num, column=6, value=float(item.descuento_porcentaje))
                ws.cell(row=row_num, column=7, value=float(item.total))
                ws.cell(row=row_num, column=7).number_format = money_fmt
                for c in range(1, 8):
                    ws.cell(row=row_num, column=c).border = thin_border
                ws.row_dimensions[row_num].outline_level = 2
                row_num += 1

    # Fila de TOTAL GENERAL
    row_num += 1
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=6)
    ws.cell(row=row_num, column=1, value="TOTAL GENERAL")
    ws.cell(row=row_num, column=1).font = Font(bold=True, size=12)
    ws.cell(row=row_num, column=7, value=float(cotizacion.total))
    ws.cell(row=row_num, column=7).font = Font(bold=True, size=12)
    ws.cell(row=row_num, column=7).number_format = money_fmt
    for c in range(1, 8):
        ws.cell(row=row_num, column=c).border = thin_border

    # Configurar outline para que se muestre colapsado
    ws.sheet_properties.outlinePr.summaryBelow = False

    # Response
    from django.http import HttpResponse
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"Cotizacion_{cotizacion.id}_{cotizacion.fecha}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@staff_required
def api_items_por_disciplina(request, disciplina_id):
    items = ItemPredefinido.objects.filter(disciplina_id=disciplina_id, activo=True).select_related('moneda')
    data = [{
        'id': i.id,
        'codigo': i.codigo,
        'descripcion': i.descripcion,
        'unidad_medida': i.unidad_medida,
        'precio_unitario': float(i.precio_unitario),
        'moneda': i.moneda.codigo,
    } for i in items]
    return JsonResponse(data, safe=False)


# ──────────────────────────────────────────────
# Items Predefinidos
# ──────────────────────────────────────────────

@staff_required
def lista_items_predefinidos(request):
    from documentos.models import Disciplina

    disciplina_id = request.GET.get('disciplina')
    familia_id = request.GET.get('familia')
    q = request.GET.get('q', '').strip()
    solo_activos = request.GET.get('activos', '1')

    items = ItemPredefinido.objects.select_related('disciplina', 'familia', 'moneda').all()
    if disciplina_id:
        items = items.filter(disciplina_id=disciplina_id)
    if familia_id:
        items = items.filter(familia_id=familia_id)
    if solo_activos == '1':
        items = items.filter(activo=True)
    if q:
        items = items.filter(Q(descripcion__icontains=q) | Q(codigo__icontains=q) | Q(notas__icontains=q))

    familias = FamiliaItem.objects.select_related('disciplina').all()
    if disciplina_id:
        familias = familias.filter(disciplina_id=disciplina_id)

    return render(request, 'presupuestos/catalogo_articulos.html', {
        'items': items,
        'disciplinas': Disciplina.objects.all(),
        'familias': familias,
        'monedas': Moneda.objects.all(),
        'disciplina_sel': disciplina_id,
        'familia_sel': familia_id,
        'q': q,
        'solo_activos': solo_activos,
    })


@staff_required
def crear_item_predefinido(request):
    from documentos.models import Disciplina
    if request.method == 'POST':
        import json as _json
        disc_id = request.POST.get('disciplina')
        familia_id = request.POST.get('familia') or None
        es_compuesto = request.POST.get('es_compuesto') == 'on'
        componentes_raw = request.POST.get('componentes_json', '[]')

        try:
            componentes_data = _json.loads(componentes_raw)
        except Exception:
            componentes_data = []

        # Si compuesto, calcular precio desde los componentes_data
        precio = request.POST.get('precio_unitario') or '0'
        if es_compuesto and componentes_data:
            total = 0
            for c in componentes_data:
                if c.get('esNuevo') and c.get('datos_nuevo'):
                    total += float(c['datos_nuevo'].get('precio_unitario', 0)) * float(c.get('cantidad', 1))
                elif c.get('id'):
                    try:
                        comp = ItemPredefinido.objects.get(pk=c['id'])
                        total += float(comp.precio_unitario) * float(c.get('cantidad', 1))
                    except ItemPredefinido.DoesNotExist:
                        pass
            precio = str(total)

        item = ItemPredefinido.objects.create(
            disciplina_id=disc_id,
            familia_id=familia_id,
            codigo=request.POST.get('codigo') or None,
            descripcion=request.POST['descripcion'],
            unidad_medida=request.POST['unidad_medida'],
            precio_unitario=precio or '0',
            moneda_id=request.POST['moneda'],
            activo=request.POST.get('activo') == 'on',
            notas=request.POST.get('notas', ''),
            es_compuesto=es_compuesto,
        )

        # Crear componentes
        if es_compuesto:
            _guardar_componentes(item, componentes_data, disc_id)
            item.recalcular_precio()

        messages.success(request, 'Artículo creado.')
        next_url = request.POST.get('next') or 'presupuestos:lista_items_predefinidos'
        return redirect(next_url)

    disc_id = request.GET.get('disciplina')
    familia_id = request.GET.get('familia')
    familias = FamiliaItem.objects.select_related('disciplina').all()
    if disc_id:
        familias = familias.filter(disciplina_id=disc_id)
    from documentos.models import Disciplina
    return render(request, 'presupuestos/form_articulo.html', {
        'disciplinas': Disciplina.objects.all(),
        'familias': familias,
        'monedas': Moneda.objects.all(),
        'disc_presel': disc_id,
        'familia_presel': familia_id,
        'es_nuevo': True,
        'componentes_iniciales_json': '[]',
    })


@staff_required
def editar_item_predefinido(request, pk):
    from documentos.models import Disciplina
    item = get_object_or_404(ItemPredefinido.objects.select_related('disciplina', 'familia', 'moneda'), pk=pk)
    if request.method == 'POST':
        if request.POST.get('accion') == 'eliminar':
            item.delete()
            messages.success(request, 'Artículo eliminado.')
            return redirect('presupuestos:lista_items_predefinidos')

        import json as _json
        es_compuesto = request.POST.get('es_compuesto') == 'on'
        componentes_raw = request.POST.get('componentes_json', '[]')
        try:
            componentes_data = _json.loads(componentes_raw)
        except Exception:
            componentes_data = []

        item.disciplina_id = request.POST['disciplina']
        item.familia_id = request.POST.get('familia') or None
        item.codigo = request.POST.get('codigo') or None
        item.descripcion = request.POST['descripcion']
        item.unidad_medida = request.POST['unidad_medida']
        item.precio_unitario = request.POST.get('precio_unitario') or '0'
        item.moneda_id = request.POST['moneda']
        item.activo = request.POST.get('activo') == 'on'
        item.notas = request.POST.get('notas', '')
        item.es_compuesto = es_compuesto
        item.save()

        if es_compuesto:
            item.componentes.all().delete()
            _guardar_componentes(item, componentes_data, item.disciplina_id)
            item.recalcular_precio()

        messages.success(request, 'Artículo actualizado.')
        return redirect('presupuestos:lista_items_predefinidos')

    familias = FamiliaItem.objects.select_related('disciplina').filter(disciplina=item.disciplina)
    # Pasar componentes existentes como JSON para el template
    import json as _json
    comps_existentes = []
    for i, c in enumerate(item.componentes.select_related('componente').order_by('orden'), 1):
        comps_existentes.append({
            'tempId': i,
            'id': c.componente.id,
            'codigo': c.componente.codigo or '',
            'descripcion': c.componente.descripcion,
            'um': c.componente.unidad_medida,
            'precio': float(c.componente.precio_unitario),
            'cantidad': float(c.cantidad),
            'esNuevo': False,
            'datos_nuevo': None,
        })
    return render(request, 'presupuestos/form_articulo.html', {
        'item': item,
        'disciplinas': Disciplina.objects.all(),
        'familias': familias,
        'monedas': Moneda.objects.all(),
        'es_nuevo': False,
        'componentes_iniciales_json': _json.dumps(comps_existentes),
    })


def _guardar_componentes(padre, componentes_data, disciplina_id_padre):
    """
    Procesa la lista de componentes del JSON del formulario.
    Crea artículos nuevos si esNuevo=True, luego crea los ComponenteItem.
    """
    for orden, c in enumerate(componentes_data):
        cantidad = float(c.get('cantidad', 1) or 1)
        comp_id = c.get('id')

        if c.get('esNuevo') and c.get('datos_nuevo'):
            nd = c['datos_nuevo']
            # Buscar moneda por defecto
            moneda = Moneda.objects.first()
            nuevo = ItemPredefinido.objects.create(
                disciplina_id=disciplina_id_padre,
                codigo=nd.get('codigo') or None,
                descripcion=nd.get('descripcion', 'Componente'),
                unidad_medida=nd.get('unidad_medida', 'und'),
                precio_unitario=nd.get('precio_unitario', 0),
                moneda=moneda,
                activo=True,
                es_compuesto=False,
            )
            comp_id = nuevo.pk

        if comp_id and int(comp_id) != padre.pk:
            try:
                comp_obj = ItemPredefinido.objects.get(pk=comp_id)
                ComponenteItem.objects.get_or_create(
                    padre=padre,
                    componente=comp_obj,
                    defaults={'cantidad': cantidad, 'orden': orden}
                )
            except ItemPredefinido.DoesNotExist:
                pass


# ── Familias CRUD ──────────────────────────────────────────────────────────────

@staff_required
def lista_familias(request):
    from documentos.models import Disciplina
    disciplina_id = request.GET.get('disciplina')
    familias = FamiliaItem.objects.select_related('disciplina').annotate(
        total_items=models.Count('items')
    ).all()
    if disciplina_id:
        familias = familias.filter(disciplina_id=disciplina_id)
    return render(request, 'presupuestos/lista_familias.html', {
        'familias': familias,
        'disciplinas': Disciplina.objects.all(),
        'disciplina_sel': disciplina_id,
    })


@staff_required
def crear_familia(request):
    from documentos.models import Disciplina
    if request.method == 'POST':
        familia = FamiliaItem.objects.create(
            disciplina_id=request.POST['disciplina'],
            nombre=request.POST['nombre'],
            descripcion=request.POST.get('descripcion', ''),
            orden=int(request.POST.get('orden') or 0),
        )
        # Si es AJAX (llamada inline desde form_articulo), devolver JSON
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            from django.http import JsonResponse as JR
            return JR({'id': familia.id, 'nombre': familia.nombre})
        messages.success(request, 'Familia creada.')
        return redirect('presupuestos:lista_familias')
    return render(request, 'presupuestos/form_familia.html', {
        'disciplinas': Disciplina.objects.all(),
        'es_nuevo': True,
    })


@staff_required
def editar_familia(request, pk):
    from documentos.models import Disciplina
    familia = get_object_or_404(FamiliaItem.objects.select_related('disciplina'), pk=pk)
    if request.method == 'POST':
        if request.POST.get('accion') == 'eliminar':
            familia.delete()
            messages.success(request, 'Familia eliminada.')
            return redirect('presupuestos:lista_familias')
        familia.disciplina_id = request.POST['disciplina']
        familia.nombre = request.POST['nombre']
        familia.descripcion = request.POST.get('descripcion', '')
        familia.orden = int(request.POST.get('orden') or 0)
        familia.save()
        messages.success(request, 'Familia actualizada.')
        return redirect('presupuestos:lista_familias')
    return render(request, 'presupuestos/form_familia.html', {
        'familia': familia,
        'disciplinas': Disciplina.objects.all(),
        'es_nuevo': False,
    })


@staff_required
def api_familias_por_disciplina(request, disciplina_id):
    """API: retorna familias de una disciplina para poblar selects dinámicos."""
    familias = FamiliaItem.objects.filter(disciplina_id=disciplina_id).values('id', 'nombre').order_by('orden', 'nombre')
    return JsonResponse(list(familias), safe=False)


# ── BOM: Artículos Compuestos ───────────────────────────────────────────────

@staff_required
def ver_bom(request, pk):
    """Vista principal del BOM de un artículo compuesto."""
    item = get_object_or_404(
        ItemPredefinido.objects.select_related('disciplina', 'familia', 'moneda'),
        pk=pk
    )
    componentes = item.componentes.select_related(
        'componente', 'componente__disciplina', 'componente__familia', 'componente__moneda'
    ).order_by('orden')

    # Todos los artículos disponibles como componentes (excluir el propio padre)
    candidatos = ItemPredefinido.objects.select_related('disciplina', 'familia').filter(
        activo=True
    ).exclude(pk=pk).order_by('disciplina__nombre', 'descripcion')

    total_calculado = sum(c.subtotal for c in componentes)

    return render(request, 'presupuestos/bom_articulo.html', {
        'item': item,
        'componentes': componentes,
        'candidatos': candidatos,
        'total_calculado': total_calculado,
    })


@staff_required
@require_POST
def bom_agregar_componente(request, pk):
    """Agrega un componente al BOM via POST (JSON o form)."""
    padre = get_object_or_404(ItemPredefinido, pk=pk)
    import json as _json
    try:
        data = _json.loads(request.body)
    except Exception:
        data = request.POST

    componente_id = data.get('componente_id')
    cantidad = float(data.get('cantidad') or 1)
    orden = int(data.get('orden') or 0)

    if not componente_id:
        return JsonResponse({'error': 'componente_id requerido'}, status=400)
    if int(componente_id) == pk:
        return JsonResponse({'error': 'Un artículo no puede ser componente de sí mismo'}, status=400)

    componente = get_object_or_404(ItemPredefinido, pk=componente_id)

    comp, created = ComponenteItem.objects.get_or_create(
        padre=padre,
        componente=componente,
        defaults={'cantidad': cantidad, 'orden': orden}
    )
    if not created:
        comp.cantidad = cantidad
        comp.orden = orden
        comp.save()

    # Recalcular precio del padre
    padre.recalcular_precio()

    return JsonResponse({
        'id': comp.id,
        'componente_id': componente.id,
        'componente_codigo': componente.codigo or '',
        'componente_descripcion': componente.descripcion,
        'componente_um': componente.unidad_medida,
        'componente_precio': float(componente.precio_unitario),
        'cantidad': float(comp.cantidad),
        'subtotal': float(comp.subtotal),
        'precio_padre_actualizado': float(padre.precio_unitario),
        'created': created,
    })


@staff_required
@require_POST
def bom_eliminar_componente(request, pk, comp_pk):
    """Elimina un componente del BOM."""
    comp = get_object_or_404(ComponenteItem, pk=comp_pk, padre_id=pk)
    padre = comp.padre
    comp.delete()
    padre.recalcular_precio()
    return JsonResponse({'ok': True, 'precio_padre_actualizado': float(padre.precio_unitario)})


@staff_required
@require_POST
def bom_actualizar_componente(request, pk, comp_pk):
    """Actualiza cantidad/orden de un componente."""
    comp = get_object_or_404(ComponenteItem, pk=comp_pk, padre_id=pk)
    import json as _json
    try:
        data = _json.loads(request.body)
    except Exception:
        data = request.POST

    if 'cantidad' in data:
        comp.cantidad = float(data['cantidad'])
    if 'orden' in data:
        comp.orden = int(data['orden'])
    comp.save()
    comp.padre.recalcular_precio()
    return JsonResponse({
        'ok': True,
        'cantidad': float(comp.cantidad),
        'subtotal': float(comp.subtotal),
        'precio_padre_actualizado': float(comp.padre.precio_unitario),
    })


@staff_required
def api_buscar_articulos(request):
    """API de búsqueda de artículos — soporta q, excluir, disciplina, familia."""
    q = request.GET.get('q', '').strip()
    excluir = request.GET.get('excluir')
    disciplina_id = request.GET.get('disciplina')
    familia_id = request.GET.get('familia')

    items = ItemPredefinido.objects.select_related('disciplina', 'familia').filter(activo=True)
    if excluir:
        items = items.exclude(pk=excluir)
    if disciplina_id:
        items = items.filter(disciplina_id=disciplina_id)
    if familia_id:
        items = items.filter(familia_id=familia_id)
    if q:
        items = items.filter(Q(descripcion__icontains=q) | Q(codigo__icontains=q))
    items = items.order_by('disciplina__nombre', 'familia__nombre', 'codigo')[:60]

    data = [{
        'id': i.id,
        'codigo': i.codigo or '',
        'descripcion': i.descripcion,
        'unidad_medida': i.unidad_medida,
        'precio_unitario': float(i.precio_unitario),
        'disciplina': i.disciplina.nombre,
        'familia': i.familia.nombre if i.familia else '',
        'es_compuesto': i.es_compuesto,
    } for i in items]
    return JsonResponse(data, safe=False)


@staff_required
def api_cotizacion_datos(request, pk):
    """Devuelve JSON con cabecera + items de una cotización."""
    cotizacion = get_object_or_404(
        Cotizacion.objects.select_related('disciplina', 'proyecto', 'creado_por'),
        pk=pk
    )
    items_qs = cotizacion.items.select_related('disciplina').order_by('orden')

    from collections import defaultdict
    grupos = defaultdict(list)
    for item in items_qs:
        key = (item.area or '', item.disciplina_id)
        grupos[key].append({
            'id': item.id,
            'area': item.area or '',
            'disciplina_id': item.disciplina_id,
            'disciplina_nombre': item.disciplina.nombre if item.disciplina else 'Sin disciplina',
            'item_predefinido_id': item.item_predefinido_id,
            'descripcion': item.descripcion,
            'unidad_medida': item.unidad_medida,
            'cantidad': float(item.cantidad),
            'precio_unitario': float(item.precio_unitario),
            'descuento_porcentaje': float(item.descuento_porcentaje),
            'aprobado': item.aprobado,
        })

    secciones = []
    for (area, disc_id), items_list in grupos.items():
        secciones.append({
            'area': area,
            'disciplina_id': disc_id,
            'disciplina_nombre': items_list[0]['disciplina_nombre'],
            'items': items_list,
        })

    return JsonResponse({
        'id': cotizacion.id,
        'numero': cotizacion.numero,
        'fecha': cotizacion.fecha.isoformat() if cotizacion.fecha else None,
        'version': cotizacion.version,
        'valida_hasta': cotizacion.valida_hasta.isoformat() if cotizacion.valida_hasta else None,
        'estado': cotizacion.estado,
        'notas': cotizacion.notas,
        'proyecto_id': cotizacion.proyecto_id,
        'creado_por': cotizacion.creado_por.username if cotizacion.creado_por else None,
        'total': float(cotizacion.total),
        'secciones': secciones,
    })


@staff_required
@require_POST
def api_cotizacion_guardar(request, pk):
    """Guarda cabecera + items de una cotización desde JSON."""
    import json
    cotizacion = get_object_or_404(Cotizacion, pk=pk)
    data = json.loads(request.body)

    if data.get('fecha'):
        cotizacion.fecha = data['fecha']
    if data.get('valida_hasta'):
        cotizacion.valida_hasta = data['valida_hasta']
    if data.get('version'):
        cotizacion.version = int(data['version'])
    if data.get('notas') is not None:
        cotizacion.notas = data['notas']
    if data.get('estado') and data['estado'] in dict(Cotizacion.ESTADOS):
        cotizacion.estado = data['estado']
    cotizacion.save()

    items = data.get('items', [])
    cotizacion.items.all().delete()
    from django.utils import timezone
    nuevos_items = []
    for idx, item in enumerate(items):
        aprobado = item.get('aprobado', False)
        obj = ItemCotizacion.objects.create(
            cotizacion=cotizacion,
            item_predefinido_id=item.get('item_predefinido_id') or None,
            disciplina_id=item.get('disciplina_id') or None,
            area=item.get('area', '') or '',
            descripcion=item['descripcion'],
            unidad_medida=item['unidad_medida'],
            cantidad=item['cantidad'],
            precio_unitario=item['precio_unitario'],
            descuento_porcentaje=item.get('descuento_porcentaje', 0),
            aprobado=aprobado,
            aprobado_en=timezone.now() if aprobado else None,
            orden=idx,
        )
        nuevos_items.append(obj)

    # Auto-crear elementos en el proyecto para items aprobados
    elementos_creados = 0
    if cotizacion.proyecto:
        from proyectos.models import ElementoProyecto
        for item_obj in nuevos_items:
            if not item_obj.aprobado:
                continue
            # Check if already exists by matching description + cotizacion
            ya_existe = ElementoProyecto.objects.filter(
                proyecto=cotizacion.proyecto,
                nombre=item_obj.descripcion[:300],
                item_cotizacion__cotizacion=cotizacion,
            ).exists()
            if not ya_existe:
                ElementoProyecto.objects.create(
                    proyecto=cotizacion.proyecto,
                    item_cotizacion=item_obj,
                    disciplina=item_obj.disciplina,
                    area=item_obj.area or '',
                    nombre=item_obj.descripcion[:300],
                    descripcion='',
                    cantidad=item_obj.cantidad,
                    unidad_medida=item_obj.unidad_medida or '',
                    precio_unitario=item_obj.precio_unitario,
                    estado='PENDIENTE',
                    orden=cotizacion.proyecto.elementos.count() + 1,
                )
                elementos_creados += 1

    return JsonResponse({
        'ok': True,
        'numero': cotizacion.numero,
        'total': float(cotizacion.total),
        'items_count': cotizacion.items.count(),
        'elementos_creados': elementos_creados,
    })


# ──────────────────────────────────────────────
# Partida Presupuestaria - Admin Fiori
# ──────────────────────────────────────────────

@login_required
def partida_admin_fiori(request):
    """Vista principal: solo carga presupuestos (Nivel 0). Hijos se cargan via AJAX."""
    from documentos.models import Disciplina
    from core.models import Departamento
    from django.db.models import Count, Sum

    presupuestos = PresupuestoAnual.objects.select_related('departamento').annotate(
        num_partidas=Count('partidas')
    ).filter(num_partidas__gt=0).order_by('-anio', 'nombre')

    presupuestos_data = []
    anios_set = set()
    total_partidas = 0

    for pres in presupuestos:
        anios_set.add(pres.anio)
        total_partidas += pres.num_partidas
        # Aggregate totals at presupuesto level
        agg = pres.partidas.aggregate(total_original=Sum('monto_proyectado'))
        presupuestos_data.append({
            'id': pres.id,
            'nombre': str(pres),
            'anio': pres.anio,
            'departamento': pres.departamento.nombre if pres.departamento else None,
            'num_partidas': pres.num_partidas,
            'total_original': float(agg['total_original'] or 0),
        })

    disciplinas = Disciplina.objects.all().order_by('nombre')
    departamentos = Departamento.objects.all().order_by('nombre')
    anios = sorted(anios_set, reverse=True)

    return render(request, 'admin/presupuestos/partida_admin_fiori.html', {
        'presupuestos_data': presupuestos_data,
        'disciplinas': disciplinas,
        'departamentos': departamentos,
        'anios': anios,
        'total_partidas': total_partidas,
        'title': 'Administrar Partidas Presupuestarias',
    })


@login_required
def partida_admin_api(request):
    """API CRUD para partidas presupuestarias (Fiori admin) + lazy loading."""
    import json
    from documentos.models import Disciplina
    from core.models import Departamento

    if request.method == 'GET':
        action = request.GET.get('action')

        if action == 'get':
            partida_id = request.GET.get('id')
            partida = get_object_or_404(PartidaPresupuestaria, pk=partida_id)
            return JsonResponse({
                'status': 'ok',
                'partida': {
                    'id': partida.id,
                    'presupuesto_id': partida.presupuesto_anual_id,
                    'disciplina_id': partida.disciplina_id,
                    'descripcion': partida.descripcion,
                    'monto_proyectado': float(partida.monto_proyectado),
                    'departamentos_ids': list(partida.departamentos.values_list('id', flat=True)),
                }
            })

        elif action == 'children_presupuesto':
            # Lazy-load partidas de un presupuesto
            pres_id = request.GET.get('id')
            partidas = PartidaPresupuestaria.objects.filter(
                presupuesto_anual_id=pres_id
            ).select_related('disciplina').prefetch_related('departamentos')

            data = []
            for p in partidas:
                nombre = p.disciplina.nombre if p.disciplina else (p.descripcion or "Partida General")
                num_items = p.items.count()
                data.append({
                    'id': p.id,
                    'nombre': nombre,
                    'descripcion': p.descripcion,
                    'monto_original': float(p.monto_proyectado),
                    'vigente': float(p.presupuesto_vigente),
                    'comprometido': float(p.total_comprometido),
                    'disponible': float(p.pendiente_comprometer),
                    'departamentos': [d.nombre for d in p.departamentos.all()],
                    'num_items': num_items,
                })
            return JsonResponse({'status': 'ok', 'partidas': data})

        elif action == 'children_partida':
            # Lazy-load items de una partida
            partida_id = request.GET.get('id')
            items = ItemPresupuesto.objects.filter(
                partida_id=partida_id, parent__isnull=True
            )
            data = []
            for item in items:
                sub_count = item.subitems.count()
                data.append({
                    'id': item.id,
                    'concepto': item.concepto,
                    'total_anual': float(item.total_anual),
                    'es_recurrente': item.es_recurrente,
                    'frecuencia': item.get_frecuencia_display() if item.es_recurrente else 'Manual',
                    'num_subitems': sub_count,
                })
            return JsonResponse({'status': 'ok', 'items': data})

        elif action == 'children_item':
            # Lazy-load sub-items de un item
            item_id = request.GET.get('id')
            subs = ItemPresupuesto.objects.filter(parent_id=item_id)
            data = []
            for s in subs:
                data.append({
                    'id': s.id,
                    'concepto': s.concepto,
                    'total_anual': float(s.total_anual),
                    'es_recurrente': s.es_recurrente,
                    'frecuencia': s.get_frecuencia_display() if s.es_recurrente else 'Manual',
                })
            return JsonResponse({'status': 'ok', 'subitems': data})

        return JsonResponse({'status': 'error', 'message': 'Acción no válida'}, status=400)

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            action = data.get('action')

            if action == 'create':
                partida = PartidaPresupuestaria.objects.create(
                    presupuesto_anual_id=data['presupuesto_id'],
                    disciplina_id=data.get('disciplina_id'),
                    descripcion=data.get('descripcion', ''),
                    monto_proyectado=data.get('monto_proyectado', 0),
                )
                if data.get('departamentos'):
                    partida.departamentos.set(data['departamentos'])
                return JsonResponse({'status': 'ok', 'message': 'Partida creada', 'id': partida.id})

            elif action == 'update':
                partida = get_object_or_404(PartidaPresupuestaria, pk=data['id'])
                partida.presupuesto_anual_id = data['presupuesto_id']
                partida.disciplina_id = data.get('disciplina_id')
                partida.descripcion = data.get('descripcion', '')
                partida.monto_proyectado = data.get('monto_proyectado', 0)
                partida.save()
                if 'departamentos' in data:
                    partida.departamentos.set(data['departamentos'])
                return JsonResponse({'status': 'ok', 'message': 'Partida actualizada'})

            elif action == 'delete':
                partida = get_object_or_404(PartidaPresupuestaria, pk=data['id'])
                partida.delete()
                return JsonResponse({'status': 'ok', 'message': 'Partida eliminada'})

            return JsonResponse({'status': 'error', 'message': 'Acción no reconocida'}, status=400)

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

    return JsonResponse({'status': 'error', 'message': 'Método no permitido'}, status=405)
