from django.shortcuts import render, get_object_or_404
from django.contrib import admin
from django.contrib.admin.views.decorators import staff_member_required
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Q, Count, OuterRef, Subquery, DecimalField, Max
from django.db.models.functions import Coalesce
from .models import SolicitudPago, Requisicion, ItemSolicitudPago
from mantenimiento.models import Empresa
from datetime import datetime
from django.http import HttpResponse, JsonResponse
from django.core.files.storage import default_storage
from django.core.cache import cache
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import json
from openpyxl.utils import get_column_letter
import json
import os

@staff_member_required
@login_required
def dashboard_pagos(request):
    """
    Dashboard para visualizar Solicitudes de Pago.
    """
    # Filtros capturados de GET
    estado_filter = request.GET.get('estado')
    search_query = request.GET.get('q')

    # QuerySet Base optimized for items
    solicitudes = SolicitudPago.objects.prefetch_related('items', 'items__requisicion').all().order_by('-fecha_solicitud')

    # Aplicar filtros
    if estado_filter:
        solicitudes = solicitudes.filter(estado=estado_filter)
    
    if search_query:
        solicitudes = solicitudes.filter(
            Q(descripcion__icontains=search_query) | 
            Q(usuario_solicitante__username__icontains=search_query) |
            Q(pk__icontains=search_query)
        )

    # Métricas Globales (Cards Superiores)
    total_solicitudes = solicitudes.count()
    
    # Calculamos totales en Python para aprovechar las properties (ya que son campos calculados complejos)
    # Si fueran muchos registros, esto debería optimizarse con anotaciones DB, 
    # pero por ahora usaremos las properties ya definidas en el modelo para consistencia.
    monto_global_solicitado = sum(s.total_solicitado for s in solicitudes)
    monto_global_aprobado = sum(s.total_aprobado for s in solicitudes)
    monto_global_pagado = sum(s.total_pagado for s in solicitudes)

    # Conteo por estados para badges rápidos
    conteo_estados = {
        'ABIERTA': solicitudes.filter(estado='ABIERTA').count(),
        'EN_REVISION': solicitudes.filter(estado='EN_REVISION').count(),
        'CERRADA': solicitudes.filter(estado='CERRADA').count(),
    }

    context = {
        **admin.site.each_context(request),
        'solicitudes': solicitudes,
        'metrics': {
            'total_count': total_solicitudes,
            'total_solicitado': monto_global_solicitado,
            'total_aprobado': monto_global_aprobado,
            'total_pagado': monto_global_pagado,
            'total_pendiente': monto_global_aprobado - monto_global_pagado,
        },
        'conteo_estados': conteo_estados,
        'current_filter': estado_filter,
        'search_query': search_query,
        'today': datetime.now(),
        'title': 'Dashboard de Pagos',
    }

    return render(request, 'admin/presupuestos/solicitudpago/dashboard.html', context)

@login_required
def api_search_requisiciones(request):
    """
    Endpoint para búsqueda asíncrona de requisiciones (Select2).
    """
    q = request.GET.get('q', '')
    
    # Anotamos el monto pagado para mostrarlo en el buscador
    pagos_reales = ItemSolicitudPago.objects.filter(requisicion=OuterRef('pk'), estatus='PAGADO')
    
    requisiciones = Requisicion.objects.annotate(
        monto_pagado_db=Coalesce(
            Subquery(pagos_reales.values('requisicion').annotate(total=Sum('monto_solicitado')).values('total')),
            0.0,
            output_field=DecimalField()
        )
    ).select_related('proveedor')

    if q:
        requisiciones = requisiciones.filter(
            Q(cr8ca_requisicion__icontains=q) |
            Q(cr8ca_asunto__icontains=q) |
            Q(proveedor__nombre__icontains=q)
        )

    # Limitar resultados para velocidad
    requisiciones = requisiciones.order_by('-cr8ca_requisicion')[:20]

    results = []
    for r in requisiciones:
        results.append({
            'id': r.pk,
            'text': f"{r.cr8ca_requisicion} - {r.cr8ca_asunto[:40]}",
            'total': float(r.cr8ca_totalenarticulos or 0),
            'pagado': float(r.monto_pagado_db or 0),
            'asunto': r.cr8ca_asunto,
        })

    return JsonResponse({'results': results})

@login_required
def detalle_solicitud_pago(request, pk):
    """
    Vista detallada de una Solicitud de Pago.
    """
    solicitud = get_object_or_404(SolicitudPago.objects.prefetch_related('items', 'items__requisicion'), pk=pk)
    
    from .models import Requisicion, ItemSolicitudPago
    from django.db.models import Sum, DecimalField
    from django.db.models.functions import Coalesce

    # Ya no cargamos todas las requisiciones aquí para evitar lentitud
    
    # Agregadores para gráficos y tablas
    items_por_proveedor = {}
    
    prov_totals = {}
    rutina_totals = {}
    partida_totals = {}
    item_budget_totals = {}

    detalle_prov = {}
    detalle_rutina = {}
    detalle_partida = {}
    detalle_item_budget = {}

    for item in solicitud.items.all().select_related('requisicion__proveedor', 'requisicion__tipo_rutina', 'requisicion__partida__disciplina', 'requisicion__item_presupuesto'):
        # --- Datos para la tabla principal ---
        p_name = item.requisicion.proveedor.nombre if item.requisicion.proveedor else "Sin Proveedor"
        if p_name not in items_por_proveedor:
            items_por_proveedor[p_name] = {
                'id': item.requisicion.proveedor.id if item.requisicion.proveedor else None,
                'lista_items': [], 
                'total': 0, 
                'total_pagado': 0
            }
        items_por_proveedor[p_name]['lista_items'].append(item)
        
        # Solo sumar si tiene estatus PAGADO o SOLICITADO
        if item.estatus in ['PAGADO', 'SOLICITADO']:
            items_por_proveedor[p_name]['total'] += (item.monto_solicitado or 0)
        
        if item.estatus == 'PAGADO':
            items_por_proveedor[p_name]['total_pagado'] += (item.monto_solicitado or 0)

        # --- Agregación para Gráficos ---
        if item.estatus in ['PAGADO', 'SOLICITADO']:
            prov_totals[p_name] = prov_totals.get(p_name, 0) + float(item.monto_solicitado or 0)
            
            r_name = item.requisicion.tipo_rutina.nombre if item.requisicion.tipo_rutina else "No Asignada"
            rutina_totals[r_name] = rutina_totals.get(r_name, 0) + float(item.monto_solicitado or 0)

            partida_obj = item.requisicion.partida
            pa_name = "Sin Partida"
            if partida_obj and partida_obj.disciplina:
                pa_name = partida_obj.disciplina.nombre
            elif partida_obj:
                pa_name = partida_obj.descripcion or "Partida General"
            partida_totals[pa_name] = partida_totals.get(pa_name, 0) + float(item.monto_solicitado or 0)

            ib_obj = item.requisicion.item_presupuesto
            ib_name = ib_obj.concepto if ib_obj else "Sin Ítem Budget"
            item_budget_totals[ib_name] = item_budget_totals.get(ib_name, 0) + float(item.monto_solicitado or 0)

            # Totales para el desglose del modal
            total_rq = float(item.requisicion.cr8ca_totalenarticulos or 0)
            pagado_rq = float(item.requisicion.monto_pagado or 0)
            item_dict = {
                'req': item.requisicion.cr8ca_requisicion,
                'pk': str(item.requisicion.pk),
                'asunto': item.requisicion.cr8ca_asunto,
                'monto': float(item.monto_solicitado or 0),
                'total_rq': total_rq,
                'pagado_rq': pagado_rq,
                'pendiente_rq': total_rq - pagado_rq
            }

            # Llenar detalles por categoría para el modal interactivo
            if p_name not in detalle_prov: detalle_prov[p_name] = []
            detalle_prov[p_name].append(item_dict)

            if r_name not in detalle_rutina: detalle_rutina[r_name] = []
            detalle_rutina[r_name].append(item_dict)

            if pa_name not in detalle_partida: detalle_partida[pa_name] = []
            detalle_partida[pa_name].append(item_dict)

            if ib_name not in detalle_item_budget: detalle_item_budget[ib_name] = []
            detalle_item_budget[ib_name].append(item_dict)

    # Preparar data final
    graph_data = {
        'proveedores': sorted([{'nombre': k, 'total': v} for k, v in prov_totals.items()], key=lambda x: x['total'], reverse=True),
        'rutinas': sorted([{'nombre': k, 'total': v} for k, v in rutina_totals.items()], key=lambda x: x['total'], reverse=True),
        'partidas': sorted([{'nombre': k, 'total': v} for k, v in partida_totals.items()], key=lambda x: x['total'], reverse=True),
        'items_budget': sorted([{'nombre': k, 'total': v} for k, v in item_budget_totals.items()], key=lambda x: x['total'], reverse=True),
        'detalle_proveedores': detalle_prov,
        'detalle_rutinas': detalle_rutina,
        'detalle_partidas': detalle_partida,
        'detalle_items_budget': detalle_item_budget,
    }

    es_procura = request.user.groups.filter(name__in=['Procura', 'PROCURA', 'Procura_Tecnica', 'PROCURA_TECNICA']).exists()

    # Mapa de requisición -> primera OC asociada (para menú contextual)
    req_oc_map = {}
    if es_procura:
        from .models import OrdenCompra
        req_ids = set()
        for prov_data in items_por_proveedor.values():
            for item in prov_data['lista_items']:
                req_ids.add(str(item.requisicion.pk))
        ocs = OrdenCompra.objects.filter(
            requisicion_id__in=req_ids
        ).values_list('requisicion_id', 'id')
        for req_id, oc_id in ocs:
            req_oc_map[str(req_id)] = oc_id

    context = {
        'solicitud': solicitud,
        'items_por_proveedor': items_por_proveedor,
        'graph_data': graph_data,
        'ESTATUS_CHOICES': ItemSolicitudPago.ESTATUS_CHOICES,
        'CONDICION_CHOICES': ItemSolicitudPago.CONDICION_PAGO_CHOICES,
        'proveedores_todos': Empresa.objects.all().order_by('nombre'),
        'es_procura': es_procura,
        'req_oc_map_json': json.dumps(req_oc_map),
    }
    
    return render(request, 'presupuestos/solicitudes_pago/detalle.html', context)

from django.http import JsonResponse
import json

@login_required
@csrf_exempt
def api_update_requisicion_fields(request, pk):
    """
    Actualiza campos específicos de una requisición (proveedor, monto, asunto).
    """
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            requisicion = get_object_or_404(Requisicion, pk=pk)
            
            if 'proveedor_id' in data:
                prov_id = data.get('proveedor_id')
                if prov_id:
                    proveedor = get_object_or_404(Empresa, pk=prov_id)
                    requisicion.proveedor = proveedor
                else:
                    requisicion.proveedor = None
            
            if 'monto' in data:
                monto_val = data.get('monto')
                if monto_val is not None:
                    requisicion.cr8ca_totalenarticulos = monto_val
                
            if 'asunto' in data:
                requisicion.cr8ca_asunto = data.get('asunto')
                
            requisicion.save()
            return JsonResponse({'status': 'success'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Solo POST'}, status=405)

@login_required
@csrf_exempt
def api_update_item_pago(request):
    """
    Actualiza monto o descripción de un item de solicitud de pago.
    """
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            item_id = data.get('id')
            campo = data.get('campo')
            valor = data.get('valor')
            
            from .models import ItemSolicitudPago
            item = get_object_or_404(ItemSolicitudPago, pk=item_id)
            
            if campo == 'monto_solicitado':
                item.monto_solicitado = valor
            elif campo == 'descripcion':
                item.descripcion = valor
            elif campo == 'estatus':
                item.estatus = valor
            elif campo == 'condicion_pago':
                item.condicion_pago = valor
                
            item.save()
            return JsonResponse({'status': 'success'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Solo POST'}, status=405)

@login_required
def api_add_requisicion_pago(request):
    """
    Agrega una requisición a una solicitud de pago existente.
    """
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            solicitud_id = data.get('solicitud_id')
            requisicion_id = data.get('requisicion_id')
            monto = data.get('monto', 0)
            descripcion = data.get('descripcion', '')
            
            from .models import ItemSolicitudPago, Requisicion, SolicitudPago
            solicitud = get_object_or_404(SolicitudPago, pk=solicitud_id)
            requisicion = get_object_or_404(Requisicion, pk=requisicion_id)
            
            # Validar duplicados en LA MISMA solicitud
            if ItemSolicitudPago.objects.filter(solicitud=solicitud, requisicion=requisicion).exists():
                return JsonResponse({
                    'status': 'error', 
                    'message': f'La requisición {requisicion.cr8ca_requisicion} ya existe en esta solicitud.'
                }, status=400)
            
            new_item = ItemSolicitudPago.objects.create(
                solicitud=solicitud,
                requisicion=requisicion,
                monto_solicitado=monto,
                descripcion=descripcion
            )
            return JsonResponse({'status': 'success', 'item_id': new_item.pk})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Solo POST'}, status=405)

@login_required
def api_delete_item_pago(request):
    """
    Elimina un item de la solicitud de pago.
    """
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            item_id = data.get('id')
            from .models import ItemSolicitudPago
            item = get_object_or_404(ItemSolicitudPago, pk=item_id)
            item.delete()
            return JsonResponse({'status': 'success'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Solo POST'}, status=405)

@login_required
def api_requisicion_detalle(request, pk):
    """
    Retorna detalles de una requisición y su historial de pagos en JSON.
    """
    from .models import Requisicion
    requisicion = get_object_or_404(Requisicion.objects.prefetch_related(
        'articulos', 
        'items_pago', 
        'items_pago__solicitud'
    ), pk=pk)
    
    articulos = []
    for art in requisicion.articulos.all():
        articulos.append({
            'descripcion': art.cr8ca_articulo,
            'cantidad': float(art.cr8ca_cantidad),
            'costo': float(art.cr8ca_costoaproximado or 0),
            'subtotal': float(art.subtotal)
        })
        
    pagos = []
    for item in requisicion.items_pago.all().order_by('-creado_en'):
        pagos.append({
            'item_id': item.pk,
            'solicitud_id': item.solicitud.pk,
            'descripcion': item.descripcion,
            'monto': float(item.monto_solicitado),
            'estatus': item.estatus,
            'fecha': item.creado_en.strftime("%d/%m/%Y")
        })
        
    documentos = []
    for doc in requisicion.documentos.all():
        documentos.append({
            'id': doc.id,
            'nombre': doc.nombre or doc.archivo.name.split('/')[-1],
            'url': doc.archivo.url,
            'fecha': doc.creado_en.strftime("%d/%m/%Y"),
            'ext': doc.archivo.name.split('.')[-1].lower() if doc.archivo else ""
        })
        
    data = {
        'id': requisicion.pk,
        'codigo': requisicion.cr8ca_requisicion,
        'asunto': requisicion.cr8ca_asunto,
        'proveedor': requisicion.proveedor.nombre if requisicion.proveedor else "No asignado",
        'proveedor_id': requisicion.proveedor.id if requisicion.proveedor else None,
        'total_estimado': float(requisicion.cr8ca_totalenarticulos or 0),
        'total_pagado': float(requisicion.monto_pagado),
        'comentarios': requisicion.cr8ca_comentarios or "",
        'partida_id': requisicion.partida_id,
        'item_id': requisicion.item_presupuesto_id,
        'articulos': articulos,
        'pagos': pagos,
        'documentos': documentos
    }
    
    return JsonResponse(data)

@login_required
@csrf_exempt
def api_add_nota_requisicion(request, pk):
    """
    Agrega una nueva nota al timeline de la requisición.
    """
    if request.method == 'POST':
        try:
            from .models import Requisicion, NotaRequisicion
            data = json.loads(request.body)
            texto = data.get('texto', '').strip()
            if not texto:
                return JsonResponse({'status': 'error', 'message': 'La nota no puede estar vacía.'}, status=400)
            requisicion = get_object_or_404(Requisicion, pk=pk)
            nota = NotaRequisicion.objects.create(
                requisicion=requisicion,
                texto=texto,
                usuario=request.user
            )
            return JsonResponse({
                'status': 'success',
                'nota': {
                    'id': nota.pk,
                    'texto': nota.texto,
                    'usuario': nota.usuario.username if nota.usuario else 'Sistema',
                    'nombre': nota.usuario.get_full_name() or nota.usuario.username if nota.usuario else 'Sistema',
                    'creado_en': nota.creado_en.strftime('%d/%m/%Y %H:%M'),
                }
            })
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Solo POST'}, status=405)


@login_required
def api_get_notas_requisicion(request, pk):
    """
    Retorna todas las notas del timeline de una requisición.
    """
    from .models import Requisicion, NotaRequisicion
    requisicion = get_object_or_404(Requisicion, pk=pk)
    notas = NotaRequisicion.objects.filter(requisicion=requisicion).order_by('-creado_en')
    data = []
    for nota in notas:
        data.append({
            'id': nota.pk,
            'texto': nota.texto,
            'usuario': nota.usuario.username if nota.usuario else 'Sistema',
            'nombre': nota.usuario.get_full_name() or nota.usuario.username if nota.usuario else 'Sistema',
            'creado_en': nota.creado_en.strftime('%d/%m/%Y %H:%M'),
        })
    return JsonResponse({'notas': data})

@login_required
def exportar_solicitud_pago_excel(request, pk):
    """
    Genera un archivo Excel con el detalle de la solicitud de pago.
    """
    solicitud = get_object_or_404(SolicitudPago.objects.prefetch_related('items', 'items__requisicion'), pk=pk)
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"Solicitud_{solicitud.pk}"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
    center_align = Alignment(horizontal="center")
    
    # Encabezado de la solicitud
    ws.merge_cells('A1:G1')
    ws['A1'] = f"SOLICITUD DE PAGO #{solicitud.pk}"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = center_align
    
    ws.append([])
    ws.append(["Fecha:", solicitud.fecha_solicitud.strftime("%d/%m/%Y")])
    ws.append(["Solicitante:", solicitud.usuario_solicitante.get_full_name() if solicitud.usuario_solicitante else "N/A"])
    ws.append(["Descripción:", solicitud.descripcion or ""])
    ws.append(["Estado:", solicitud.get_estado_display()])
    ws.append([])
    
    # Tabla de items
    headers = ["N° REQUISICIÓN", "PROVEEDOR", "ASUNTO", "DESCRIPCIÓN PAGO", "MONTO SOLICITADO", "PAGADO RQ", "AVANCE %"]
    ws.append(headers)
    
    # Aplicar estilos a headers
    for cell in ws[ws.max_row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    for item in solicitud.items.all():
        total_rq = item.requisicion.cr8ca_totalenarticulos or 0
        pagado_rq = item.requisicion.monto_pagado or 0
        porcentaje = 0
        if total_rq > 0:
            porcentaje = ((pagado_rq + (item.monto_solicitado or 0)) / total_rq) * 100
        
        row = [
            item.requisicion.cr8ca_requisicion,
            item.requisicion.proveedor.nombre if item.requisicion.proveedor else "N/A",
            item.requisicion.cr8ca_asunto,
            item.descripcion,
            float(item.monto_solicitado or 0),
            float(pagado_rq),
            f"{porcentaje:.2f}%"
        ]
        ws.append(row)
        
    # Totales al final
    ws.append([])
    ws.append(["", "", "", "TOTAL SOLICITADO:", float(solicitud.total_solicitado or 0)])
    
    # Ajustar anchos
    for col in ws.columns:
        max_length = 0
        column_index = col[0].column
        column_letter = get_column_letter(column_index)
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Solicitud_Pago_{solicitud.pk}.xlsx'
    wb.save(response)
    return response

@login_required
def import_items_pago_background(request, pk):
    """
    Vista para subir archivo e iniciar la importación asíncrona de items.
    """
    solicitud = get_object_or_404(SolicitudPago, pk=pk)
    
    if request.method == 'POST' and request.FILES.get('archivo'):
        archivo = request.FILES['archivo']
        ext = os.path.splitext(archivo.name)[1].lower().replace('.', '')
        
        if ext not in ['csv', 'xls', 'xlsx']:
            return render(request, 'presupuestos/solicitudes_pago/import_items.html', {
                'solicitud': solicitud,
                'error': 'Formato no soportado. Use CSV, XLS o XLSX.'
            })
            
        # Guardar temporalmente
        path = default_storage.save(f'tmp/import_items_{request.user.id}_{datetime.now().timestamp()}.{ext}', archivo)
        
        # Iniciar tarea Celery
        from .tasks import import_items_solicitud_task
        import_items_solicitud_task.delay(
            path, ext, 
            user_id=request.user.id, 
            solicitud_id=solicitud.pk,
            verification_mode=True # Iniciar con verificación
        )
        
        return render(request, 'presupuestos/solicitudes_pago/import_items.html', {
            'solicitud': solicitud,
            'importing': True,
            'cache_key': f"import_items_pago_progress_{request.user.id}"
        })
        
    return render(request, 'presupuestos/solicitudes_pago/import_items.html', {'solicitud': solicitud})

@login_required
def import_items_pago_process(request):
    """
    API para confirmar la ejecución real (Import o Dry Run) tras la verificación.
    """
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            solicitud_id = data.get('solicitud_id')
            file_path = data.get('file_path')
            dry_run = data.get('dry_run', False)
            
            ext = os.path.splitext(file_path)[1].lower().replace('.', '')
            
            from .tasks import import_items_solicitud_task
            import_items_solicitud_task.delay(
                file_path, ext, 
                user_id=request.user.id, 
                solicitud_id=solicitud_id,
                verification_mode=False,
                dry_run=dry_run
            )
            return JsonResponse({'status': 'success'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Solo POST'}, status=405)

@login_required
def import_items_pago_progress(request):
    """
    API para consultar el progreso de la importación desde el caché.
    """
    cache_key = f"import_items_pago_progress_{request.user.id}"
    data = cache.get(cache_key)
    return JsonResponse(data or {'status': 'waiting'})

@login_required
def dashboard_proveedores(request):
    """
    Lista de proveedores que tienen requisiciones, con métricas resumidas.
    """
    search_query = request.GET.get('q', '')
    
    # Proveedores con al menos una requisición
    proveedores = Empresa.objects.annotate(
        total_requisiciones=Count('requisiciones_asignadas'),
        monto_total_solicitado=Sum('requisiciones_asignadas__cr8ca_totalenarticulos'),
        ultima_fecha_requisicion=Max('requisiciones_asignadas__fecha')
    ).filter(total_requisiciones__gt=0).order_by('-monto_total_solicitado')

    if search_query:
        proveedores = proveedores.filter(nombre__icontains=search_query)

    context = {
        'proveedores': proveedores,
        'search_query': search_query,
    }
    return render(request, 'presupuestos/proveedores/dashboard_prov.html', context)

@login_required
def detalle_proveedor(request, empresa_id):
    """
    Detalle estadístico de un proveedor específico.
    """
    empresa = get_object_or_404(Empresa, pk=empresa_id)
    
    requisiciones = empresa.requisiciones_asignadas.all().order_by('-fecha')
    
    # Requisiciones Pendientes (Estado PENDIENTE)
    pendientes = requisiciones.filter(estado_requisicion='PENDIENTE')
    count_pendientes = pendientes.count()
    monto_pendiente = pendientes.aggregate(total=Sum('cr8ca_totalenarticulos'))['total'] or 0
    
    # Pagos realizados (Items de Solicitud de Pago con estatus PAGADO)
    pagos = ItemSolicitudPago.objects.filter(
        requisicion__proveedor=empresa,
        estatus='PAGADO'
    ).aggregate(total=Sum('monto_solicitado'))['total'] or 0

    context = {
        'empresa': empresa,
        'requisiciones': requisiciones[:10],  # Últimas 10
        'count_pendientes': count_pendientes,
        'monto_pendiente': monto_pendiente,
        'total_pagado': pagos,
    }
    return render(request, 'presupuestos/proveedores/detalle_prov.html', context)

@login_required
def api_get_budget_selection_data(request):
    """Retorna todas las partidas y facultativamente ítems de una partida."""
    from .models import PartidaPresupuestaria, ItemPresupuesto
    partida_id = request.GET.get('partida_id')
    
    partidas = PartidaPresupuestaria.objects.select_related('disciplina', 'presupuesto_anual').all().order_by('-presupuesto_anual__anio', 'disciplina__nombre')
    
    data_partidas = [
        {
            'id': p.id,
            'nombre': str(p)
        } for p in partidas
    ]
    
    items = []
    if partida_id:
        items = list(ItemPresupuesto.objects.filter(partida_id=partida_id).values('id', 'concepto'))
        
    return JsonResponse({
        'partidas': data_partidas,
        'items': items
    })

@login_required
def api_requisicion_update_budget(request):
    """Actualiza la partida e ítem de una requisición."""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            requisicion_id = data.get('requisicion_id')
            partida_id = data.get('partida_id')
            item_id = data.get('item_id')
            
            from .models import Requisicion, PartidaPresupuestaria, ItemPresupuesto
            requisicion = get_object_or_404(Requisicion, pk=requisicion_id)
            
            if partida_id:
                requisicion.partida_id = partida_id
            else:
                requisicion.partida = None
                
            if item_id:
                requisicion.item_presupuesto_id = item_id
            else:
                requisicion.item_presupuesto = None
                
            requisicion.save()
            return JsonResponse({'status': 'success'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Solo POST'}, status=405)
